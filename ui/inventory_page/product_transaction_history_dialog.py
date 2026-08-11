# ui/inventory_page/product_transaction_history_dialog.py
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QTableWidget,
    QTableWidgetItem, QHeaderView, QPushButton, QDateEdit,
    QComboBox, QMessageBox, QFileDialog, QFrame, QWidget
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QIcon, QColor
from models.database import connect_db
from utils.currency import format_money
from utils.excel_exporter import ExcelExporter
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.date_range_widget import DateRangeWidget
from ui.widgets.modern_button import ModernButton
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime


class ProductTransactionHistoryDialog(QDialog):
    """Product Transaction History Dialog - Theme-aware"""
    
    def __init__(self, product_id, product_name, parent=None):
        super().__init__(parent)
        self.product_id = product_id
        self.product_name = product_name
        self.current_page = 1
        self.page_size = 25
        self._is_dark = is_dark_theme()
        
        self.setWindowTitle(f"Transaction History - {product_name}")
        self.setMinimumSize(1100, 580)
        self.setWindowIcon(QIcon("assets/icons/zaypos.png"))
        self.setModal(True)
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)

        layout = QVBoxLayout()
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header info
        info_layout = QHBoxLayout()
        info_layout.setSpacing(15)
        
        self.info_label = QLabel(f"<b>Product:</b> {product_name}")
        self.info_label.setStyleSheet("font-size: 12pt;")
        info_layout.addWidget(self.info_label)
        
        info_layout.addStretch()
        
        self.stock_label = QLabel()
        self.stock_label.setStyleSheet("font-size: 11pt; font-weight: bold;")
        info_layout.addWidget(self.stock_label)
        
        layout.addLayout(info_layout)

        # Filter section - Using DateRangeWidget and ModernButton
        filter_frame = QFrame()
        filter_frame.setObjectName("filter_frame")
        colors = get_theme_colors()
        filter_frame.setStyleSheet(self._get_filter_frame_style(colors))
        
        filter_layout = QHBoxLayout(filter_frame)
        filter_layout.setSpacing(12)
        filter_layout.setContentsMargins(15, 8, 15, 8)
        
        # Date Range Widget
        date_label = QLabel("📅 Date:")
        date_label.setStyleSheet(f"color: {colors['text']}; font-size: 10pt;")
        filter_layout.addWidget(date_label)
        
        self.date_range = DateRangeWidget(self)
        self.date_range.date_range_changed.connect(self.on_filter_changed)
        filter_layout.addWidget(self.date_range)
        
        # Action filter
        action_label = QLabel("📌 Action:")
        action_label.setStyleSheet(f"color: {colors['text']}; font-size: 10pt;")
        filter_layout.addWidget(action_label)
        
        self.action_filter = QComboBox()
        self.action_filter.addItems(["All", "Stock In", "Stock Out", "Adjustment", "Sale"])
        self.action_filter.currentTextChanged.connect(self.on_filter_changed)
        self.action_filter.setStyleSheet(self._get_combobox_style(colors))
        filter_layout.addWidget(self.action_filter)
        
        # Refresh Button
        self.btn_refresh = ModernButton("🔄 Refresh", ModernButton.SECONDARY)
        self.btn_refresh.set_compact(True)
        self.btn_refresh.clicked.connect(self.load_history)
        filter_layout.addWidget(self.btn_refresh)
        
        # Export Button
        self.btn_export = ModernButton("📊 Export Excel", ModernButton.SECONDARY)
        self.btn_export.set_compact(True)
        self.btn_export.clicked.connect(self.export_to_excel)
        filter_layout.addWidget(self.btn_export)
        
        filter_layout.addStretch()
        layout.addWidget(filter_frame)

        # Transaction table
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        
        # Apply table style
        self._update_table_style(colors)
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Apply initial theme
        self._apply_theme()
        
        self.load_current_stock()
        self.load_history()

    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
        self.load_history()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        
        # Dialog background
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {colors['bg']};
            }}
        """)
        
        # Update info label
        if hasattr(self, 'info_label'):
            self.info_label.setStyleSheet(f"font-size: 12pt; color: {colors['text']};")
        
        # Update filter frame
        filter_frame = self.findChild(QFrame, "filter_frame")
        if filter_frame:
            filter_frame.setStyleSheet(self._get_filter_frame_style(colors))
        
        # Update filter labels
        for child in self.findChildren(QLabel):
            if child.parent() and child.parent().objectName() == "filter_frame":
                child.setStyleSheet(f"color: {colors['text']}; font-size: 10pt;")
        
        # Update combobox
        if hasattr(self, 'action_filter'):
            self.action_filter.setStyleSheet(self._get_combobox_style(colors))
        
        # Update table
        self._update_table_style(colors)
        
        # Update stock label
        self.load_current_stock()
    
    def _get_filter_frame_style(self, colors):
        return f"""
            QFrame#filter_frame {{
                background: {colors['bg_hover']};
                border-radius: 8px;
                padding: 5px;
            }}
        """
    
    def _get_combobox_style(self, colors):
        return f"""
            QComboBox {{
                padding: 6px 12px;
                border: 1px solid {colors['border']};
                border-radius: 6px;
                background: {colors['card_bg']};
                color: {colors['text']};
                font-size: 10pt;
                min-width: 120px;
            }}
            QComboBox:focus {{
                border-color: #5865f2;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox QAbstractItemView {{
                background-color: {colors['card_bg']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                color: {colors['text']};
                selection-background-color: #5865f2;
                selection-color: white;
                padding: 4px;
            }}
            QComboBox QAbstractItemView::item {{
                padding: 4px 8px;
                border-radius: 2px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: {colors['bg_hover']};
            }}
            QComboBox QAbstractItemView::item:selected {{
                background-color: #5865f2;
                color: white;
            }}
        """
    
    def _update_table_style(self, colors):
        """Update table style based on theme"""
        is_dark = is_dark_theme()
        
        if is_dark:
            table_style = """
                QTableWidget {
                    background-color: #2f3136;
                    alternate-background-color: #36393f;
                    selection-background-color: #40444b;
                    selection-color: #dcddde;
                    gridline-color: #40444b;
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    color: #dcddde;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #dcddde;
                }
                QTableWidget::item:selected {
                    background-color: #40444b;
                    color: #dcddde;
                }
                QHeaderView::section {
                    background-color: #202225;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #40444b;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #b9bbbe;
                }
                QTableWidget::item:hover {
                    background-color: #40444b;
                }
            """
        else:
            table_style = """
                QTableWidget {
                    background-color: white;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e9ecef;
                    selection-color: #212529;
                    gridline-color: #dee2e6;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    color: #212529;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #212529;
                }
                QTableWidget::item:selected {
                    background-color: #e9ecef;
                    color: #212529;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #dee2e6;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #2c3e50;
                }
                QTableWidget::item:hover {
                    background-color: #f1f3f5;
                }
            """
        
        self.table.setStyleSheet(table_style)
    
    def load_current_stock(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT stock FROM products WHERE id = ?", (self.product_id,))
        row = cursor.fetchone()
        conn.close()
        
        colors = get_theme_colors()
        
        if row:
            stock = row[0] if row[0] else 0
            lang = self.get_lang()
            if lang == "my":
                self.stock_label.setText(f"လက်ကျန်စတော့: {stock}")
            else:
                self.stock_label.setText(f"Current Stock: {stock}")
            
            if stock <= 0:
                self.stock_label.setStyleSheet(f"font-size: 11pt; color: #e74c3c; font-weight: bold; background: transparent; border: none;")
            else:
                self.stock_label.setStyleSheet(f"font-size: 11pt; color: #2ecc71; font-weight: bold; background: transparent; border: none;")

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def on_filter_changed(self):
        self.current_page = 1
        self.load_history()

    def on_page_changed(self, page, page_size):
        self.current_page = page
        self.page_size = page_size
        self.load_history()

    def load_history(self):
        from_date = self.date_range.get_from_date()
        to_date = self.date_range.get_to_date()
        action = self.action_filter.currentText()
        lang = self.get_lang()
        
        # Update headers based on language
        if lang == "my":
            headers = ["ရက်စွဲ", "လုပ်ဆောင်ချက်", "မပြောင်းမီ", "ပြောင်းပြီး", "ပြောင်းလဲမှု", "နေရာ", "အသုံးပြုသူ", "မှတ်ချက်"]
        else:
            headers = ["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"]
        self.table.setHorizontalHeaderLabels(headers)
        
        conn = connect_db()
        cursor = conn.cursor()
        
        # Check if location column exists
        cursor.execute("PRAGMA table_info(stock_movements)")
        columns = [col[1] for col in cursor.fetchall()]
        has_location = 'location' in columns
        
        action_map = {
            "Stock In": "in",
            "Stock Out": "out",
            "Adjustment": "adjustment",
            "Sale": "sale"
        }
        
        count_query = """
            SELECT COUNT(*)
            FROM stock_movements sm
            WHERE sm.product_id = ? 
              AND date(sm.created_at) BETWEEN ? AND ?
        """
        count_params = [self.product_id, from_date, to_date]
        
        if action != "All":
            db_action = action_map.get(action, action.lower())
            count_query += " AND sm.type = ?"
            count_params.append(db_action)
        
        cursor.execute(count_query, count_params)
        result = cursor.fetchone()
        total_items = result[0] if result else 0
        
        if total_items is None:
            total_items = 0
        else:
            total_items = int(total_items)
        
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        # Build query based on whether location column exists
        if has_location:
            data_query = """
                SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                       sm.quantity, sm.created_by, sm.reason, sm.notes,
                       sm.location
                FROM stock_movements sm
                WHERE sm.product_id = ? 
                  AND date(sm.created_at) BETWEEN ? AND ?
            """
        else:
            data_query = """
                SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                       sm.quantity, sm.created_by, sm.reason, sm.notes,
                       NULL as location
                FROM stock_movements sm
                WHERE sm.product_id = ? 
                  AND date(sm.created_at) BETWEEN ? AND ?
            """
        
        data_params = [self.product_id, from_date, to_date]
        
        if action != "All":
            db_action = action_map.get(action, action.lower())
            data_query += " AND sm.type = ?"
            data_params.append(db_action)
        
        data_query += " ORDER BY sm.created_at DESC LIMIT ? OFFSET ?"
        
        offset = (self.current_page - 1) * self.page_size
        cursor.execute(data_query, data_params + [self.page_size, offset])
        rows = cursor.fetchall()
        conn.close()
        
        self.table.setRowCount(0)
        
        for row in rows:
            if has_location:
                created_at, action_type, old_stock, new_stock, qty, user, reason, notes, location = row
            else:
                created_at, action_type, old_stock, new_stock, qty, user, reason, notes = row
                location = None
            
            if lang == "my":
                action_display = {
                    "in": "စတော့ဝင်",
                    "out": "စတော့ထွက်",
                    "adjustment": "ပြင်ဆင်ချက်",
                    "sale": "ရောင်းချမှု"
                }.get(action_type, action_type)
            else:
                action_display = {
                    "in": "Stock In",
                    "out": "Stock Out",
                    "adjustment": "Adjustment",
                    "sale": "Sale"
                }.get(action_type, action_type)
            
            date_str = created_at[:16] if created_at else ""
            
            if qty is not None:
                qty_display = abs(qty)
                if action_type in ["out", "sale"]:
                    qty_display = f"-{abs(qty)}"
                elif action_type == "adjustment":
                    if new_stock > old_stock:
                        qty_display = f"+{new_stock - old_stock}"
                    else:
                        qty_display = f"{new_stock - old_stock}"
                else:
                    qty_display = f"+{abs(qty)}"
            else:
                qty_display = ""
            
            remark = reason if reason else (notes if notes else "")
            location_display = location if location else "-"
            
            row_num = self.table.rowCount()
            self.table.insertRow(row_num)
            self.table.setItem(row_num, 0, QTableWidgetItem(date_str))
            
            action_item = QTableWidgetItem(action_display)
            if action_type in ["in"]:
                action_item.setForeground(QColor(46, 204, 113))
            elif action_type in ["out", "sale"]:
                action_item.setForeground(QColor(231, 76, 60))
            else:
                action_item.setForeground(QColor(241, 196, 15))
            self.table.setItem(row_num, 1, action_item)
            
            self.table.setItem(row_num, 2, QTableWidgetItem(str(old_stock) if old_stock is not None else ""))
            self.table.setItem(row_num, 3, QTableWidgetItem(str(new_stock) if new_stock is not None else ""))
            self.table.setItem(row_num, 4, QTableWidgetItem(str(qty_display)))
            self.table.setItem(row_num, 5, QTableWidgetItem(str(location_display)))
            self.table.setItem(row_num, 6, QTableWidgetItem(user or ""))
            self.table.setItem(row_num, 7, QTableWidgetItem(remark or ""))

    def export_to_excel(self):
        from_date = self.date_range.get_from_date()
        to_date = self.date_range.get_to_date()
        action = self.action_filter.currentText()
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"transaction_history_{self.product_name}_{from_date}_to_{to_date}.xlsx",
            "Export Transaction History" if lang != "my" else "ငွေပေးချေမှုမှတ်တမ်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Check if location column exists
            cursor.execute("PRAGMA table_info(stock_movements)")
            columns = [col[1] for col in cursor.fetchall()]
            has_location = 'location' in columns
            
            action_map = {
                "Stock In": "in",
                "Stock Out": "out",
                "Adjustment": "adjustment",
                "Sale": "sale"
            }
            
            if has_location:
                query = """
                    SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                           sm.quantity, sm.created_by, sm.reason, sm.notes,
                           sm.location
                    FROM stock_movements sm
                    WHERE sm.product_id = ? 
                      AND date(sm.created_at) BETWEEN ? AND ?
                """
            else:
                query = """
                    SELECT sm.created_at, sm.type, sm.old_stock, sm.new_stock, 
                           sm.quantity, sm.created_by, sm.reason, sm.notes,
                           NULL as location
                    FROM stock_movements sm
                    WHERE sm.product_id = ? 
                      AND date(sm.created_at) BETWEEN ? AND ?
                """
            
            params = [self.product_id, from_date, to_date]
            
            if action != "All":
                db_action = action_map.get(action, action.lower())
                query += " AND sm.type = ?"
                params.append(db_action)
            
            query += " ORDER BY sm.created_at DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Transaction History"
            
            ws.merge_cells('A1:H1')
            ws['A1'] = f"TRANSACTION HISTORY - {self.product_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            if lang == "my":
                headers = ["ရက်စွဲ", "လုပ်ဆောင်ချက်", "မပြောင်းမီ", "ပြောင်းပြီး", "ပြောင်းလဲမှု", "နေရာ", "အသုံးပြုသူ", "မှတ်ချက်"]
            else:
                headers = ["Date", "Action", "Qty Before", "Qty After", "Quantity", "Location", "User", "Remark"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row_data in enumerate(rows, start=6):
                if has_location:
                    created_at, action_type, old_stock, new_stock, qty, user, reason, notes, location = row_data
                else:
                    created_at, action_type, old_stock, new_stock, qty, user, reason, notes = row_data
                    location = None
                
                if lang == "my":
                    action_display = {
                        "in": "စတော့ဝင်",
                        "out": "စတော့ထွက်",
                        "adjustment": "ပြင်ဆင်ချက်",
                        "sale": "ရောင်းချမှု"
                    }.get(action_type, action_type)
                else:
                    action_display = {
                        "in": "Stock In",
                        "out": "Stock Out",
                        "adjustment": "Adjustment",
                        "sale": "Sale"
                    }.get(action_type, action_type)
                
                ws.cell(row=row_idx, column=1, value=created_at[:16] if created_at else "")
                ws.cell(row=row_idx, column=2, value=action_display)
                ws.cell(row=row_idx, column=3, value=old_stock if old_stock is not None else "")
                ws.cell(row=row_idx, column=4, value=new_stock if new_stock is not None else "")
                
                if qty is not None:
                    if action_type in ["out", "sale"]:
                        ws.cell(row=row_idx, column=5, value=f"-{abs(qty)}")
                    elif action_type == "adjustment":
                        if new_stock > old_stock:
                            ws.cell(row=row_idx, column=5, value=f"+{new_stock - old_stock}")
                        else:
                            ws.cell(row=row_idx, column=5, value=f"{new_stock - old_stock}")
                    else:
                        ws.cell(row=row_idx, column=5, value=f"+{abs(qty)}")
                else:
                    ws.cell(row=row_idx, column=5, value="")
                
                ws.cell(row=row_idx, column=6, value=location if location else "-")
                ws.cell(row=row_idx, column=7, value=user or "")
                
                remark = reason if reason else (notes if notes else "")
                ws.cell(row=row_idx, column=8, value=remark or "")
            
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['H'].width = 25
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def retranslateUi(self):
        """Retranslate UI for language change"""
        lang = self.get_lang()
        self.date_range.retranslateUi(lang)
        
        if lang == "my":
            self.btn_refresh.setText("🔄 ပြန်လည်")
            self.btn_export.setText("📊 Excel ထုတ်မည်")
        else:
            self.btn_refresh.setText("🔄 Refresh")
            self.btn_export.setText("📊 Export Excel")
        
        # Apply theme after language change
        self._apply_theme()
        self.load_current_stock()
        self.load_history()
    
    def showEvent(self, event):
        """Refresh data when dialog becomes visible"""
        self.load_current_stock()
        self.load_history()
        super().showEvent(event)