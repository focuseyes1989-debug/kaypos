# ui/inventory_page/purchase_history_tab.py
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QMessageBox, QFileDialog, QLabel, QComboBox
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QColor, QIcon, QPixmap, QPainter
from models.database import connect_db
from utils.currency import format_money, get_currency_symbol
from utils.translations import tr
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.modern_button import ModernButton
from ui.widgets.search_widget import SearchWidget
from ui.widgets.date_range_widget import DateRangeWidget
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime
import csv
import os


class PurchaseHistoryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        self.current_page = 1
        self.page_size = 50
        self._is_dark = is_dark_theme()
        
        layout = QVBoxLayout()
        layout.setSpacing(12)

        # ============================================================
        # ✅ FILTER ROW - All in one horizontal line
        # ============================================================
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(0, 8, 0, 8)
        
        # 1. SearchWidget (expands to fill available space)
        self.search_widget = SearchWidget(
            placeholder="Search by PO number, supplier or product...",
            show_label=False
        )
        self.search_widget.search_changed.connect(self.reset_pagination)
        self.search_widget.search_cleared.connect(self.reset_pagination)
        filter_layout.addWidget(self.search_widget, 3)  # stretch=3
        
        # 2. Payment status filter
        filter_layout.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Paid", "Unpaid", "Partial"])
        self.status_filter.currentTextChanged.connect(self.reset_pagination)
        self.status_filter.setMinimumWidth(120)
        filter_layout.addWidget(self.status_filter, 1)  # stretch=1
        
        # 3. Date Range Widget
        self.date_range = DateRangeWidget()
        self.date_range.date_range_changed.connect(self.reset_pagination)
        filter_layout.addWidget(self.date_range, 2)  # stretch=2
        
        # 4. Export Button (rightmost)
        self.btn_export = ModernButton(" Export", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(False)
        self.btn_export.clicked.connect(self.export_purchase_history)
        filter_layout.addWidget(self.btn_export, 0)  # no stretch
        
        layout.addLayout(filter_layout)
        
        # ============================================================
        # TABLE - NO custom style, use PyQt6 default
        # ============================================================
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        
        # ✅ NO custom table style
        # self._apply_table_theme()  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        # ============================================================
        # PAGINATION
        # ============================================================
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.retranslateUi()

    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        # ✅ Only update button icons and reload data - no table style update
        self._update_button_icons()
        self.load_data()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_export.set_icon("file_export", size=(16, 16))

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def reset_pagination(self):
        self.current_page = 1
        self.load_data()

    def on_page_changed(self, page: int, page_size: int):
        self.current_page = page
        self.page_size = page_size
        self.load_data(page, page_size)

    def refresh(self):
        self.load_data()
        self.retranslateUi()

    def get_all_purchase_data(self):
        """Get all purchase history data for export"""
        conn = connect_db()
        cursor = conn.cursor()
        
        search_text = self.search_widget.get_text().lower()
        status_filter = self.status_filter.currentText()
        from_date = self.date_range.get_from_date()
        to_date = self.date_range.get_to_date()
        
        query = """
            SELECT 
                po.po_no,
                s.name as supplier_name,
                GROUP_CONCAT(p.name, ' | ') as product_names,
                GROUP_CONCAT(poi.quantity, ' | ') as quantities,
                GROUP_CONCAT(poi.unit_price, ' | ') as unit_prices,
                po.total_amount,
                po.payment_status,
                po.order_date,
                po.received_by,
                po.notes
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN purchase_order_items poi ON po.id = poi.po_id
            LEFT JOIN products p ON poi.product_id = p.id
            WHERE po.order_date BETWEEN ? AND ?
        """
        params = [from_date, to_date]
        
        if search_text:
            query += " AND (LOWER(po.po_no) LIKE ? OR LOWER(s.name) LIKE ? OR LOWER(p.name) LIKE ?)"
            like = f'%{search_text}%'
            params.extend([like, like, like])
        
        if status_filter != "All":
            query += " AND po.payment_status = ?"
            params.append(status_filter)
        
        query += " GROUP BY po.id ORDER BY po.order_date DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_purchase_history(self):
        """Export purchase history to CSV"""
        lang = self.get_lang()
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export Purchase History" if lang != "my" else "ဝယ်ယူမှုမှတ်တမ်း ထုတ်ရန်", 
            f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", 
            "CSV Files (*.csv)"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_purchase_data()
            symbol = get_currency_symbol()
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow(["=" * 90])
                writer.writerow(["PURCHASE HISTORY REPORT"] if lang != "my" else ["ဝယ်ယူမှုမှတ်တမ်း အစီရင်ခံစာ"])
                writer.writerow(["=" * 90])
                writer.writerow([])
                writer.writerow(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                writer.writerow(["Total Purchase Orders:", len(rows)])
                writer.writerow([])
                
                # Column Headers
                if lang == "my":
                    writer.writerow(["ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ပစ္စည်းများ", 
                                   "အရေအတွက်များ", "တစ်ခုချင်းကုန်ကျစရိတ်",
                                   "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", "ဝယ်ယူရက်", 
                                   "လက်ခံသူ", "မှတ်ချက်"])
                else:
                    writer.writerow(["Purchase No", "Supplier", "Products", 
                                   "Quantities", "Unit Costs",
                                   "Total Amount", "Payment Status", "Purchase Date", 
                                   "Received By", "Notes"])
                writer.writerow(["-" * 90])
                
                total_amount = 0
                paid_count = 0
                unpaid_count = 0
                partial_count = 0
                
                for row in rows:
                    po_no, supplier, products, quantities, unit_prices, total, payment_status, order_date, received_by, notes = row
                    
                    writer.writerow([
                        po_no or "",
                        supplier or "",
                        (products[:200] + "...") if products and len(products) > 200 else (products or ""),
                        quantities or "",
                        unit_prices or "",
                        format_money(total, symbol),
                        payment_status or "Unpaid",
                        order_date or "",
                        received_by or "",
                        notes or ""
                    ])
                    
                    total_amount += total if total else 0
                    if payment_status == "Paid":
                        paid_count += 1
                    elif payment_status == "Unpaid":
                        unpaid_count += 1
                    else:
                        partial_count += 1
                
                writer.writerow([])
                writer.writerow(["=" * 90])
                writer.writerow(["SUMMARY"])
                writer.writerow(["-" * 50])
                writer.writerow(["Total Purchase Amount:", format_money(total_amount, symbol)])
                writer.writerow(["Paid Orders:", paid_count])
                writer.writerow(["Unpaid Orders:", unpaid_count])
                writer.writerow(["Partial Payments:", partial_count])
                writer.writerow(["=" * 90])
                writer.writerow(["End of Report"])
            
            msg = f"Purchase history exported successfully to:\n{file_path}" if lang != "my" else f"ဝယ်ယူမှုမှတ်တမ်း အောင်မြင်စွာ ထုတ်ယူပြီးပါပြီ:\n{file_path}"
            QMessageBox.information(self, "Export Complete", msg)
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {e}")

    def load_data(self, page=None, page_size=None):
        if page is None:
            page = self.current_page
        if page_size is None:
            page_size = self.page_size
            
        lang = self.get_lang()
        search_text = self.search_widget.get_text().lower()
        status_filter = self.status_filter.currentText()
        from_date = self.date_range.get_from_date()
        to_date = self.date_range.get_to_date()
        
        # Define headers
        if lang == "my":
            headers = [
                "ID", "ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ပစ္စည်း", "အရေအတွက်", 
                "တစ်ခုချင်းကုန်ကျစရိတ်", "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", 
                "ဝယ်ယူရက်", "လက်ခံသူ", "မှတ်ချက်"
            ]
        else:
            headers = [
                "ID", "Purchase No", "Supplier", "Product", "Quantity", 
                "Unit Cost", "Total Amount", "Payment Status", 
                "Purchase Date", "Received By", "Notes"
            ]
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # Hide ID column (column 0)
        self.table.setColumnHidden(0, True)

        conn = connect_db()
        cursor = conn.cursor()
        
        # Build base query with filters
        base_query = """
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN purchase_order_items poi ON po.id = poi.po_id
            LEFT JOIN products p ON poi.product_id = p.id
            WHERE po.order_date BETWEEN ? AND ?
        """
        params = [from_date, to_date]
        
        if search_text:
            base_query += " AND (LOWER(po.po_no) LIKE ? OR LOWER(s.name) LIKE ? OR LOWER(p.name) LIKE ?)"
            like = f'%{search_text}%'
            params.extend([like, like, like])
        
        if status_filter != "All":
            base_query += " AND po.payment_status = ?"
            params.append(status_filter)
        
        # Get total count for pagination
        count_query = f"SELECT COUNT(DISTINCT po.id) {base_query}"
        cursor.execute(count_query, params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        
        # Get purchase orders with their items
        data_query = f"""
            SELECT 
                po.id,
                po.po_no,
                s.name as supplier_name,
                GROUP_CONCAT(p.name, ', ') as product_names,
                GROUP_CONCAT(poi.quantity, ', ') as quantities,
                GROUP_CONCAT(poi.unit_price, ', ') as unit_prices,
                po.total_amount,
                po.payment_status,
                po.order_date,
                po.received_by,
                po.notes
            {base_query}
            GROUP BY po.id
            ORDER BY po.order_date DESC
            LIMIT ? OFFSET ?
        """
        
        cursor.execute(data_query, params + [page_size, offset])
        rows = cursor.fetchall()
        conn.close()

        self.table.setRowCount(0)
        
        for row in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            
            # Column 0: ID (hidden)
            self.table.setItem(r, 0, QTableWidgetItem(str(row[0])))
            
            # Column 1: Purchase No
            self.table.setItem(r, 1, QTableWidgetItem(str(row[1]) if row[1] else ""))
            
            # Column 2: Supplier
            self.table.setItem(r, 2, QTableWidgetItem(str(row[2]) if row[2] else ""))
            
            # Column 3: Product (combined)
            self.table.setItem(r, 3, QTableWidgetItem(str(row[3]) if row[3] else ""))
            
            # Column 4: Quantity (combined)
            self.table.setItem(r, 4, QTableWidgetItem(str(row[4]) if row[4] else ""))
            
            # Column 5: Unit Cost (combined)
            self.table.setItem(r, 5, QTableWidgetItem(str(row[5]) if row[5] else ""))
            
            # Column 6: Total Amount
            total_amt = row[6] if row[6] else 0
            self.table.setItem(r, 6, QTableWidgetItem(format_money(total_amt)))
            
            # Column 7: Payment Status - ✅ Use hardcoded colors
            payment_status = row[7] if row[7] else "Unpaid"
            status_item = QTableWidgetItem(payment_status)
            if payment_status == "Paid":
                status_item.setForeground(QColor("#28a745"))  # Green
            elif payment_status == "Unpaid":
                status_item.setForeground(QColor("#dc3545"))  # Red
            else:  # Partial
                status_item.setForeground(QColor("#f39c12"))  # Orange
            self.table.setItem(r, 7, status_item)
            
            # Column 8: Purchase Date
            self.table.setItem(r, 8, QTableWidgetItem(str(row[8]) if row[8] else ""))
            
            # Column 9: Received By
            self.table.setItem(r, 9, QTableWidgetItem(str(row[9]) if row[9] else ""))
            
            # Column 10: Notes
            self.table.setItem(r, 10, QTableWidgetItem(str(row[10]) if row[10] else ""))
    
    def retranslateUi(self):
        lang = self.get_lang()
        
        # Update SearchWidget placeholder
        if lang == "my":
            self.search_widget.set_placeholder_text("ဝယ်ယူမှုအမှတ် / ပေးသွင်းသူ / ပစ္စည်းဖြင့် ရှာရန်...")
            self.btn_export.setText(" ထုတ်မည်")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "ပေးပြီး")
            self.status_filter.setItemText(2, "မပေးရသေး")
            self.status_filter.setItemText(3, "တစ်ပိုင်းပေးပြီး")
        else:
            self.search_widget.set_placeholder_text("Search by PO number, supplier or product...")
            self.btn_export.setText(" Export")
            self.status_filter.setItemText(0, "All")
            self.status_filter.setItemText(1, "Paid")
            self.status_filter.setItemText(2, "Unpaid")
            self.status_filter.setItemText(3, "Partial")
        
        # Update button icons
        self._update_button_icons()
        
        self.load_data()
    
    def export_to_excel(self):
        """Export purchase history to Excel"""
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"purchase_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Purchase History" if lang != "my" else "ဝယ်ယူမှုမှတ်တမ်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_purchase_data()
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Purchase History"
            
            # Title
            ws.merge_cells('A1:J1')
            ws['A1'] = "PURCHASE HISTORY REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Purchase Orders: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            # Headers
            if lang == "my":
                headers = ["ဝယ်ယူမှုအမှတ်", "ပေးသွင်းသူ", "ပစ္စည်းများ", 
                          "အရေအတွက်များ", "တစ်ခုချင်းကုန်ကျစရိတ်",
                          "စုစုပေါင်းပမာဏ", "ငွေပေးချေမှုအခြေအနေ", 
                          "ရက်စွဲ", "လက်ခံသူ", "မှတ်ချက်"]
            else:
                headers = ["Purchase No", "Supplier", "Products", 
                          "Quantities", "Unit Costs",
                          "Total Amount", "Payment Status", 
                          "Date", "Received By", "Notes"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_amount = 0
            paid_count = 0
            unpaid_count = 0
            partial_count = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                po_no, supplier, products, quantities, unit_prices, total, payment_status, order_date, received_by, notes = row_data
                
                ws.cell(row=row_idx, column=1, value=po_no or "")
                ws.cell(row=row_idx, column=2, value=supplier or "")
                ws.cell(row=row_idx, column=3, value=(products[:100] + "...") if products and len(products) > 100 else (products or ""))
                ws.cell(row=row_idx, column=4, value=quantities or "")
                ws.cell(row=row_idx, column=5, value=unit_prices or "")
                ws.cell(row=row_idx, column=6, value=format_money(total, symbol))
                ws.cell(row=row_idx, column=7, value=payment_status or "Unpaid")
                ws.cell(row=row_idx, column=8, value=order_date or "")
                ws.cell(row=row_idx, column=9, value=received_by or "")
                ws.cell(row=row_idx, column=10, value=notes or "")
                
                total_amount += total if total else 0
                if payment_status == "Paid":
                    paid_count += 1
                elif payment_status == "Unpaid":
                    unpaid_count += 1
                else:
                    partial_count += 1
            
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=5, value="SUMMARY").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=5, value=f"Total Amount: {format_money(total_amount, symbol)}")
            ws.cell(row=summary_row + 2, column=5, value=f"Paid Orders: {paid_count}")
            ws.cell(row=summary_row + 3, column=5, value=f"Unpaid Orders: {unpaid_count}")
            ws.cell(row=summary_row + 4, column=5, value=f"Partial Payments: {partial_count}")
            
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 18
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def showEvent(self, event):
        """Handle show event - refresh data"""
        self.load_data()
        super().showEvent(event)