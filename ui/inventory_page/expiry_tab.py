# ui/inventory_page/expiry_tab.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, 
    QHeaderView, QMessageBox, QFileDialog, QLabel, 
    QComboBox, QDateEdit, QMenu, QProgressDialog, QDialog, QFormLayout,
    QDialogButtonBox, QCheckBox, QDoubleSpinBox, QLineEdit, QFrame
)
from PyQt6.QtCore import Qt, QDate, QSize, pyqtSignal
from PyQt6.QtGui import QColor, QIcon, QPixmap, QPainter, QAction
from models.database import connect_db
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.summary_card_widget import SummaryCardWidget
from ui.widgets.modern_button import ModernButton
from ui.widgets.search_widget import SearchWidget
from ui.widgets.action_toolbar import ActionToolbar
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import date, datetime, timedelta
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from utils.translations import tr
from models.database.queries import get_all_expired_stock, get_expiring_soon_stock
from ui.print_barcode_dialog import PrintBarcodeDialog
from loguru import logger  # ✅ Added
import os


class ExpiryTab(QWidget):
    """Expiry management tab with write-off capabilities"""
    
    # Signal when data is refreshed
    data_refreshed = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        self.current_page = 1
        self.page_size = 25
        self._is_dark = is_dark_theme()
        self._selected_location_ids = []  # Track selected items
        self._ensure_expiry_discount_columns()
        
        layout = QVBoxLayout()
        layout.setSpacing(12)

        # ====== Filter Section ======
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(0, 8, 0, 8)

        self.search_widget = SearchWidget(
            placeholder="Search by product, batch, or location...",
            show_label=False
        )
        self.search_widget.search_changed.connect(self.on_filter_changed)
        self.search_widget.search_cleared.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_widget, 2)

        filter_layout.addWidget(QLabel("Location:"))
        self.location_filter = QComboBox()
        self.location_filter.addItem("All Locations")
        self.location_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.location_filter, 1)

        filter_layout.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Expired", "Expiring Soon", "OK"])
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.status_filter, 1)

        # ====== Action Toolbar ======
        self.action_toolbar = ActionToolbar(self)
        self.btn_edit_expiry = self.action_toolbar.add_primary(
            " Edit Expiry",
            self.edit_selected_expiry_date,
            "edit",
            ModernButton.SECONDARY,
            width=116,
            stretch=False,
        )
        self.btn_edit_expiry.setEnabled(False)
        self.btn_write_off = self.action_toolbar.add_primary(
            " Write Off",
            self.write_off_selected,
            "delete",
            ModernButton.DANGER,
            width=112,
            stretch=False,
        )
        self.btn_write_off.setEnabled(False)
        self.action_toolbar.add_separator()
        self.action_apply_discount = self.action_toolbar.add_more_action(
            "Apply Expiry Discount",
            self.apply_expiry_discount,
            "percent_discount",
        )
        self.action_clear_discount = self.action_toolbar.add_more_action(
            "Clear Expiry Discount",
            self.clear_expiry_discount,
            "close",
        )
        self.action_print_clearance_label = self.action_toolbar.add_more_action(
            "Print Clearance Label",
            self.print_clearance_label,
            "barcode",
        )
        self.action_toolbar.add_separator()
        self.action_write_off_all = self.action_toolbar.add_more_action(
            "Write Off All Expired",
            self.write_off_all_expired,
            "delete",
        )
        self.action_export = self.action_toolbar.add_more_action(
            "Export Excel",
            self.export_to_excel,
            "file_export",
        )
        self.btn_write_off_all = self.action_write_off_all
        self.btn_export = self.action_export
        self.action_toolbar.finalize()
        filter_layout.addWidget(self.action_toolbar, 0)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # ====== Summary Cards ======
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)

        self.total_card = SummaryCardWidget(
            title="Total Batches",
            value="0",
            icon="inventory",
            color="#3498db",
            icon_is_svg=True
        )
        self.total_card.set_icon("inventory", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.total_card)

        self.expired_card = SummaryCardWidget(
            title="Expired Batches",
            value="0",
            icon="cancel",
            color="#e74c3c",
            icon_is_svg=True
        )
        self.expired_card.set_icon("cancel", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.expired_card)

        self.expiring_card = SummaryCardWidget(
            title="Expiring Soon (7 days)",
            value="0",
            icon="warning",
            color="#f39c12",
            icon_is_svg=True
        )
        self.expiring_card.set_icon("warning", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.expiring_card)

        self.total_qty_card = SummaryCardWidget(
            title="Total Quantity",
            value="0",
            icon="bar_chart",
            color="#2ecc71",
            icon_is_svg=True
        )
        self.total_qty_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.total_qty_card)

        layout.addLayout(card_layout)

        # ====== Main Table ======
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        
        # Enable context menu
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)

        self._set_headers()
        
        # Configure column widths
        header = self.table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
            header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)

        vertical_header = self.table.verticalHeader()
        if vertical_header is not None:
            vertical_header.setDefaultSectionSize(40)
        
        # Hide location_id column
        self.table.setColumnHidden(8, True)
        
        # Selection change
        self.table.itemSelectionChanged.connect(self._update_write_off_button)
        
        layout.addWidget(self.table)

        # ====== Pagination ======
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)

        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)

        # Load initial data
        self.load_locations()
        self.load_data()

    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._update_button_icons()
        self._update_card_icons()
        self.load_data()

    def _ensure_expiry_discount_columns(self):
        """Add expiry discount columns for existing databases."""
        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(product_locations)")
            columns = {row[1] for row in cursor.fetchall()}
            for column, definition in {
                "expiry_discount_enabled": "INTEGER DEFAULT 0",
                "expiry_discount_percent": "REAL DEFAULT 0",
                "expiry_discount_start_date": "TEXT",
                "expiry_discount_end_date": "TEXT",
                "clearance_note": "TEXT",
            }.items():
                if column not in columns:
                    cursor.execute(f"ALTER TABLE product_locations ADD COLUMN {column} {definition}")
            conn.commit()
        except Exception as e:
            logger.warning(f"Could not ensure expiry discount columns: {e}")
        finally:
            if conn:
                conn.close()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_write_off.set_icon("delete", size=(16, 16))
        self.btn_edit_expiry.set_icon("edit", size=(16, 16))
        if hasattr(self, "action_toolbar"):
            self.action_toolbar.update_theme()
    
    def _update_card_icons(self):
        """Update card icons when theme changes"""
        self.total_card.set_icon("inventory", is_svg=True, size=(24, 24))
        self.expired_card.set_icon("cancel", is_svg=True, size=(24, 24))
        self.expiring_card.set_icon("warning", is_svg=True, size=(24, 24))
        self.total_qty_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        
        self.total_card.update_theme()
        self.expired_card.update_theme()
        self.expiring_card.update_theme()
        self.total_qty_card.update_theme()

    def _set_headers(self):
        """Set table headers based on language."""
        lang = self.get_lang()
        if lang == "my":
            headers = [
                "ပစ္စည်းအမည်", "အသုတ်အမှတ်", "နေရာ", "သက်တမ်းကုန်ရက်",
                "ကျန်ရက်များ", "အရေအတွက်", "အခြေအနေ", "ဝင်ရက်အဟောင်း", "Location ID"
            ]
        else:
            headers = [
                "Product Name", "Batch No", "Location", "Expiry Date",
                "Days Left", "Quantity", "Status", "Age", "Location ID"
            ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        # Hide location_id column after setting headers
        self.table.setColumnHidden(8, True)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def load_locations(self):
        """Load locations from product_locations table."""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT location FROM product_locations 
            WHERE location IS NOT NULL AND location != '' 
            ORDER BY location
        """)
        rows = cursor.fetchall()
        conn.close()

        self.location_filter.blockSignals(True)
        current = self.location_filter.currentText()
        self.location_filter.clear()
        self.location_filter.addItem("All Locations")

        for (name,) in rows:
            self.location_filter.addItem(name)

        idx = self.location_filter.findText(current)
        if idx >= 0:
            self.location_filter.setCurrentIndex(idx)
        self.location_filter.blockSignals(False)

    def on_filter_changed(self):
        """Handle filter changes."""
        self.current_page = 1
        self.load_data()

    def on_page_changed(self, page: int, page_size: int):
        """Handle page changes."""
        self.current_page = page
        self.page_size = page_size
        self.load_data()

    def refresh(self):
        """Refresh the tab data."""
        self.load_locations()
        self.load_data()

    def _update_write_off_button(self):
        """Update write off button state based on selection"""
        selected = self.table.selectedIndexes()
        if selected:
            rows = set(idx.row() for idx in selected)
            self.btn_edit_expiry.setEnabled(len(rows) == 1)
            # Check if any selected items are expired
            has_expired = False
            for row in rows:
                status_item = self.table.item(row, 6)
                if status_item:
                    status_text = status_item.text().lower()
                    lang = self.get_lang()
                    if lang == "my":
                        if "သက်တမ်းကုန်ပြီ" in status_text:
                            has_expired = True
                            break
                    else:
                        if "expired" in status_text:
                            has_expired = True
                            break
            self.btn_write_off.setEnabled(has_expired and len(rows) > 0)
        else:
            self.btn_write_off.setEnabled(False)
            self.btn_edit_expiry.setEnabled(False)

    def load_data(self, page=None, page_size=None):
        """Load batch-level expiry data from product_locations table."""
        if page is None:
            page = self.current_page
        if page_size is None:
            page_size = self.page_size

        lang = self.get_lang()
        search_text = self.search_widget.get_text().lower()
        location_filter = self.location_filter.currentText()
        status_filter = self.status_filter.currentText()

        self._set_headers()

        today = date.today()

        conn = connect_db()
        cursor = conn.cursor()

        # Build query from product_locations table
        base_query = """
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND pl.quantity > 0
        """
        params = []

        # Search filter
        if search_text:
            like = f'%{search_text}%'
            base_query += """ AND (
                LOWER(p.name) LIKE ? OR 
                LOWER(pl.batch_no) LIKE ? OR 
                LOWER(pl.location) LIKE ? OR 
                LOWER(pl.batch_no) LIKE ?
            )"""
            params.extend([like, like, like, like])

        # Location filter
        if location_filter != "All Locations":
            base_query += " AND pl.location = ?"
            params.append(location_filter)

        # Status filter
        if status_filter != "All":
            if status_filter == "Expired" or status_filter == "သက်တမ်းကုန်ပြီ":
                base_query += """ AND pl.expire_date IS NOT NULL 
                    AND pl.expire_date != '' 
                    AND date(pl.expire_date) < date('now')"""
            elif status_filter == "Expiring Soon" or status_filter == "သက်တမ်းနီးပြီ":
                base_query += """ AND pl.expire_date IS NOT NULL 
                    AND pl.expire_date != '' 
                    AND date(pl.expire_date) >= date('now') 
                    AND date(pl.expire_date) <= date('now', '+7 days')"""
            elif status_filter == "OK" or status_filter == "ကောင်းသည်":
                base_query += """ AND (pl.expire_date IS NULL 
                    OR pl.expire_date = '' 
                    OR date(pl.expire_date) > date('now', '+7 days'))"""

        # Count total
        count_query = f"SELECT COUNT(*) {base_query}"
        cursor.execute(count_query, params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        # Main query with pagination
        offset = (page - 1) * page_size
        data_query = f"""
            SELECT 
                pl.id as location_id,
                p.name as product_name,
                pl.batch_no,
                pl.location,
                pl.expire_date,
                pl.quantity,
                pl.last_updated,
                p.sku,
                p.category,
                p.price
            {base_query}
            ORDER BY 
                CASE 
                    WHEN pl.expire_date IS NULL OR pl.expire_date = '' THEN 1 
                    ELSE 0 
                END,
                pl.expire_date ASC,
                pl.last_updated ASC
            LIMIT ? OFFSET ?
        """

        cursor.execute(data_query, params + [page_size, offset])
        rows = cursor.fetchall()
        conn.close()

        # Populate table
        self.table.setRowCount(0)

        expired_count = 0
        expiring_count = 0
        total_quantity = 0
        total_batches = len(rows)

        expired_color = "#dc3545"
        expiring_color = "#f39c12"
        ok_color = "#28a745"
        gray_color = "#6c757d"

        self._selected_location_ids = []

        for row in rows:
            location_id, product_name, batch_no, location, expire_date, quantity, last_updated, sku, category, price = row

            # Calculate days left
            days_left = None
            status = "OK"
            status_color = QColor(ok_color)

            if expire_date and expire_date.strip():
                try:
                    exp_d = date.fromisoformat(expire_date)
                    days_left = (exp_d - today).days

                    if days_left < 0:
                        status = "Expired"
                        status_color = QColor(expired_color)
                        expired_count += 1
                    elif days_left <= 7:
                        status = "Expiring Soon"
                        status_color = QColor(expiring_color)
                        expiring_count += 1
                    else:
                        status = "OK"
                        status_color = QColor(ok_color)
                except:
                    status = "Invalid Date"
                    status_color = QColor(gray_color)
            else:
                status = "No Expiry"
                status_color = QColor(gray_color)

            # Calculate age
            age_days = "N/A"
            if last_updated:
                try:
                    if isinstance(last_updated, str):
                        last_updated_str = last_updated[:10] if len(last_updated) >= 10 else last_updated
                        last_date = date.fromisoformat(last_updated_str)
                        age_days = (today - last_date).days
                        if age_days < 0:
                            age_days = 0
                        age_days = f"{age_days} days"
                except:
                    age_days = "N/A"

            # Translate status for display
            lang = self.get_lang()
            if lang == "my":
                status_display = {
                    "Expired": "သက်တမ်းကုန်ပြီ",
                    "Expiring Soon": "သက်တမ်းနီးပြီ",
                    "OK": "ကောင်းသည်",
                    "No Expiry": "သက်တမ်းမရှိ",
                    "Invalid Date": "ရက်စွဲမမှန်ပါ"
                }.get(status, status)
            else:
                status_display = status

            r = self.table.rowCount()
            self.table.insertRow(r)

            # Product Name (with SKU if available)
            name_display = product_name
            if sku:
                name_display = f"{product_name} ({sku})"
            self.table.setItem(r, 0, QTableWidgetItem(name_display))

            # Batch No
            self.table.setItem(r, 1, QTableWidgetItem(batch_no or "-"))

            # Location
            self.table.setItem(r, 2, QTableWidgetItem(location or "-"))

            # Expiry Date
            expiry_item = QTableWidgetItem(expire_date or "-")
            if status == "Expired":
                expiry_item.setForeground(QColor(expired_color))
            elif status == "Expiring Soon":
                expiry_item.setForeground(QColor(expiring_color))
            self.table.setItem(r, 3, expiry_item)

            # Days Left
            days_item = QTableWidgetItem(str(days_left) if days_left is not None else "-")
            if status == "Expired":
                days_item.setForeground(QColor(expired_color))
            elif status == "Expiring Soon":
                days_item.setForeground(QColor(expiring_color))
            self.table.setItem(r, 4, days_item)

            # Quantity
            qty_item = QTableWidgetItem(str(quantity))
            if quantity <= 0:
                qty_item.setForeground(QColor(expired_color))
            self.table.setItem(r, 5, qty_item)
            total_quantity += quantity

            # Status
            status_item = QTableWidgetItem(status_display)
            status_item.setForeground(status_color)
            self.table.setItem(r, 6, status_item)

            # Age
            self.table.setItem(r, 7, QTableWidgetItem(str(age_days)))

            # Location ID (hidden)
            loc_id_item = QTableWidgetItem(str(location_id))
            self.table.setItem(r, 8, loc_id_item)

        # Update summary cards
        self.total_card.set_value(str(total_batches))
        self.expired_card.set_value(str(expired_count))
        self.expiring_card.set_value(str(expiring_count))
        self.total_qty_card.set_value(str(total_quantity))

        if expired_count > 0:
            self.expired_card.set_color(expired_color)
        if expiring_count > 0:
            self.expiring_card.set_color(expiring_color)

        self._update_write_off_button()
        self.data_refreshed.emit()

    # =========================================================================
    # WRITE OFF OPERATIONS
    # =========================================================================

    def _get_selected_location_ids(self):
        """Get location IDs from selected rows"""
        selected = self.table.selectedIndexes()
        if not selected:
            return []
        
        rows = set(idx.row() for idx in selected)
        location_ids = []
        
        for row in rows:
            loc_id_item = self.table.item(row, 8)  # Hidden location_id column
            if loc_id_item:
                try:
                    location_ids.append(int(loc_id_item.text()))
                except ValueError:
                    pass
        
        return location_ids

    def _get_location_row(self, location_id: int):
        """Return product location details for expiry editing."""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                pl.id,
                pl.product_id,
                p.name,
                pl.location,
                pl.batch_no,
                pl.expire_date,
                pl.quantity
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE pl.id = ?
        """, (location_id,))
        row = cursor.fetchone()
        conn.close()
        return row

    def edit_selected_expiry_date(self):
        """Edit expiry date for the selected product-location batch."""
        location_ids = self._get_selected_location_ids()
        if len(location_ids) != 1:
            QMessageBox.warning(self, "Select One Batch", "Please select one batch to edit expiry date.")
            return

        row = self._get_location_row(location_ids[0])
        if not row:
            QMessageBox.warning(self, "Not Found", "Selected batch was not found.")
            self.refresh()
            return

        location_id, _product_id, product_name, location, batch_no, current_expiry, quantity = row
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Expiry Date")
        dialog.setModal(True)
        dialog.resize(360, 180)

        colors = get_theme_colors()
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {colors.get('bg', '#ffffff')};
                color: {colors.get('text', '#212529')};
            }}
            QLabel {{
                color: {colors.get('text', '#212529')};
            }}
        """)

        layout = QVBoxLayout(dialog)
        form = QFormLayout()
        form.addRow("Product:", QLabel(product_name or "-"))
        form.addRow("Location:", QLabel(location or "-"))
        form.addRow("Batch:", QLabel(batch_no or "-"))
        form.addRow("Quantity:", QLabel(str(quantity)))

        no_expiry_check = QCheckBox("No Expiry")
        date_edit = QDateEdit()
        date_edit.setCalendarPopup(True)
        date_edit.setDisplayFormat("yyyy-MM-dd")
        if current_expiry:
            date_value = QDate.fromString(str(current_expiry), "yyyy-MM-dd")
            date_edit.setDate(date_value if date_value.isValid() else QDate.currentDate())
            no_expiry_check.setChecked(False)
        else:
            date_edit.setDate(QDate.currentDate())
            no_expiry_check.setChecked(True)
        date_edit.setEnabled(not no_expiry_check.isChecked())
        no_expiry_check.toggled.connect(lambda checked: date_edit.setEnabled(not checked))

        form.addRow("Expiry Date:", date_edit)
        form.addRow("", no_expiry_check)
        layout.addLayout(form)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        new_expiry = "" if no_expiry_check.isChecked() else date_edit.date().toString("yyyy-MM-dd")
        if (current_expiry or "") == new_expiry:
            return

        if self._update_location_expiry(location_id, new_expiry):
            QMessageBox.information(self, "Expiry Updated", "Expiry date updated successfully.")
            self.refresh()
        else:
            QMessageBox.critical(self, "Update Failed", "Could not update expiry date. Please try again.")

    def _sync_product_expiry_date(self, cursor, product_id: int):
        """Keep product master expiry aligned to the earliest dated active batch."""
        cursor.execute("""
            SELECT expire_date
            FROM product_locations
            WHERE product_id = ?
              AND quantity > 0
              AND expire_date IS NOT NULL
              AND expire_date != ''
            ORDER BY date(expire_date) ASC
            LIMIT 1
        """, (product_id,))
        row = cursor.fetchone()
        master_expiry = row[0] if row else ""
        cursor.execute("""
            UPDATE products
            SET expire_date = ?
            WHERE id = ?
        """, (master_expiry, product_id))

    def _update_location_expiry(self, location_id: int, new_expiry: str) -> bool:
        """Update expiry date, merging rows if the target batch already exists."""
        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE")

            cursor.execute("""
                SELECT product_id, location, batch_no, expire_date, quantity
                FROM product_locations
                WHERE id = ?
            """, (location_id,))
            row = cursor.fetchone()
            if not row:
                conn.rollback()
                return False

            product_id, location, batch_no, old_expiry, quantity = row
            old_expiry = old_expiry or ""
            new_expiry = new_expiry or ""
            if old_expiry == new_expiry:
                conn.rollback()
                return True

            cursor.execute("""
                SELECT id, quantity
                FROM product_locations
                WHERE product_id = ?
                  AND COALESCE(location, '') = COALESCE(?, '')
                  AND COALESCE(batch_no, '') = COALESCE(?, '')
                  AND COALESCE(expire_date, '') = ?
                  AND id != ?
            """, (product_id, location, batch_no, new_expiry, location_id))
            existing = cursor.fetchone()

            if existing:
                target_id, _target_qty = existing
                cursor.execute("""
                    UPDATE product_locations
                    SET quantity = quantity + ?,
                        last_updated = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (quantity, target_id))
                cursor.execute("DELETE FROM product_locations WHERE id = ?", (location_id,))
            else:
                cursor.execute("""
                    UPDATE product_locations
                    SET expire_date = ?,
                        last_updated = CURRENT_TIMESTAMP
                    WHERE id = ?
                """, (new_expiry, location_id))

            self._sync_product_expiry_date(cursor, product_id)
            cursor.execute("""
                INSERT INTO stock_movements
                (product_id, type, quantity, reason, reference, notes, location)
                VALUES (?, 'adjustment', 0, 'Edit Expiry Date', ?, ?, ?)
            """, (
                product_id,
                f"EXPIRY_EDIT_{location_id}",
                f"Expiry date changed from {old_expiry or 'No Expiry'} to {new_expiry or 'No Expiry'}; batch: {batch_no or 'N/A'}",
                location or ""
            ))

            conn.commit()
            logger.info(f"Expiry updated for product_location {location_id}: {old_expiry or 'No Expiry'} -> {new_expiry or 'No Expiry'}")
            return True
        except Exception as e:
            logger.error(f"Failed to update expiry for location {location_id}: {e}", exc_info=True)
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            return False
        finally:
            if conn:
                conn.close()

    def apply_expiry_discount(self):
        """Apply a clearance discount to selected expiry batches."""
        location_ids = self._get_selected_location_ids()
        if not location_ids:
            QMessageBox.warning(self, "No Selection", "Please select one or more batches.")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Apply Expiry Discount")
        dialog.setModal(True)
        dialog.resize(760, 460)
        layout = QVBoxLayout(dialog)
        body = QHBoxLayout()
        body.setSpacing(14)
        layout.addLayout(body, 1)

        preview_rows = self._get_expiry_discount_dialog_rows(location_ids)
        first = preview_rows[0] if preview_rows else None

        left_panel = QFrame()
        left_layout = QVBoxLayout(left_panel)
        form = QFormLayout()

        selected_label = QLabel(f"Selected batches: {len(location_ids)}")
        selected_label.setStyleSheet("font-weight: 600;")
        percent_input = QDoubleSpinBox()
        percent_input.setRange(0.01, 100.0)
        percent_input.setDecimals(2)
        percent_input.setSuffix(" %")
        existing_percent = float(first[10] or 0) if first else 0
        percent_input.setValue(existing_percent if existing_percent > 0 else 20.0)
        start_date = QDateEdit()
        start_date.setCalendarPopup(True)
        start_date.setDisplayFormat("yyyy-MM-dd")
        start_date.setDate(QDate.currentDate())
        end_date = QDateEdit()
        end_date.setCalendarPopup(True)
        end_date.setDisplayFormat("yyyy-MM-dd")
        default_end = QDate.fromString(first[9], "yyyy-MM-dd") if first and first[9] else QDate.currentDate().addDays(7)
        end_date.setDate(default_end if default_end.isValid() else QDate.currentDate().addDays(7))
        if first and first[15]:
            existing_start = QDate.fromString(first[15], "yyyy-MM-dd")
            if existing_start.isValid():
                start_date.setDate(existing_start)
        if first and first[16]:
            existing_end = QDate.fromString(first[16], "yyyy-MM-dd")
            if existing_end.isValid():
                end_date.setDate(existing_end)
        note_input = QLineEdit()
        note_input.setPlaceholderText("Clearance sale / expiring soon")
        if first and first[11]:
            note_input.setText(first[11])

        form.addRow("", selected_label)
        form.addRow("Discount:", percent_input)
        form.addRow("Start Date:", start_date)
        form.addRow("End Date:", end_date)
        form.addRow("Note:", note_input)
        left_layout.addLayout(form)
        left_layout.addStretch()
        body.addWidget(left_panel, 2)

        right_panel = QFrame()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(8)
        image_title = QLabel("Product Image")
        image_title.setStyleSheet("font-weight: 600;")
        right_layout.addWidget(image_title)
        image_preview = QLabel()
        image_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        image_preview.setMinimumSize(240, 240)
        image_preview.setWordWrap(True)
        right_layout.addWidget(image_preview, 1)
        info_title = QLabel("Product Information")
        info_title.setStyleSheet("font-weight: 600;")
        right_layout.addWidget(info_title)
        product_details = QLabel()
        product_details.setWordWrap(True)
        right_layout.addWidget(product_details)
        body.addWidget(right_panel, 1)

        self._populate_expiry_discount_preview(first, image_preview, product_details)
        self._style_expiry_discount_dialog(dialog, image_preview, product_details)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        if end_date.date() < start_date.date():
            QMessageBox.warning(self, "Invalid Date", "End date must be after start date.")
            return

        if self._set_expiry_discount(
            location_ids,
            percent_input.value(),
            note_input.text().strip(),
            start_date.date().toString("yyyy-MM-dd"),
            end_date.date().toString("yyyy-MM-dd"),
        ):
            QMessageBox.information(self, "Discount Applied", f"Applied expiry discount to {len(location_ids)} batch(es).")
            self.refresh()
            self._refresh_discount_page()
        else:
            QMessageBox.critical(self, "Update Failed", "Could not apply expiry discount.")

    def _get_expiry_discount_dialog_rows(self, location_ids):
        if not location_ids:
            return []
        conn = connect_db()
        cursor = conn.cursor()
        placeholders = ",".join("?" for _ in location_ids)
        cursor.execute(f"""
            SELECT
                pl.id,
                pl.product_id,
                p.name,
                COALESCE(p.sku, ''),
                COALESCE(p.barcode, ''),
                COALESCE(p.category, ''),
                COALESCE(p.price, 0),
                COALESCE(p.stock, 0),
                COALESCE(p.image, ''),
                COALESCE(pl.expire_date, ''),
                COALESCE(pl.expiry_discount_percent, 0),
                COALESCE(pl.clearance_note, ''),
                COALESCE(pl.batch_no, ''),
                COALESCE(pl.location, ''),
                COALESCE(pl.quantity, 0),
                COALESCE(pl.expiry_discount_start_date, ''),
                COALESCE(pl.expiry_discount_end_date, '')
            FROM product_locations pl
            JOIN products p ON p.id = pl.product_id
            WHERE pl.id IN ({placeholders})
            ORDER BY p.name, pl.expire_date
        """, location_ids)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def _populate_expiry_discount_preview(self, row, image_preview, product_details):
        if not row:
            image_preview.setText("No Image\n\nSelect a batch to preview")
            product_details.setText("No selected batch")
            return
        (
            _loc_id, _product_id, name, sku, barcode, category, price, stock,
            image_path, expire_date, discount_percent, note, batch_no, location, quantity,
            discount_start, discount_end
        ) = row
        if image_path:
            pixmap = QPixmap(image_path)
            if not pixmap.isNull():
                image_preview.setPixmap(pixmap.scaled(
                    image_preview.width() - 16,
                    image_preview.height() - 16,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                ))
            else:
                image_preview.setText("Image not available")
                image_preview.setPixmap(QPixmap())
        else:
            image_preview.setText("No Image Available")
            image_preview.setPixmap(QPixmap())
        product_details.setText(
            f"Name: {name or '-'}\n"
            f"SKU: {sku or '-'}\n"
            f"Barcode: {barcode or '-'}\n"
            f"Category: {category or '-'}\n"
            f"Price: {float(price or 0):g}\n"
            f"Total Stock: {int(stock or 0)}\n"
            f"Batch: {batch_no or '-'}\n"
            f"Location: {location or '-'}\n"
            f"Batch Qty: {int(quantity or 0)}\n"
            f"Expiry: {expire_date or '-'}\n"
            f"Discount Start: {discount_start or '-'}\n"
            f"Discount End: {discount_end or '-'}\n"
            f"Current Discount: {float(discount_percent or 0):g}%\n"
            f"Note: {note or '-'}"
        )

    def _style_expiry_discount_dialog(self, dialog, image_preview, product_details):
        colors = get_theme_colors()
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {colors.get('bg', '#ffffff')};
                color: {colors.get('text', '#212529')};
            }}
            QFrame {{
                background-color: {colors.get('card_bg', '#ffffff')};
                border: 1px solid {colors.get('border', '#dee2e6')};
                border-radius: 6px;
            }}
            QLabel {{
                color: {colors.get('text', '#212529')};
                background: transparent;
                border: none;
            }}
            QLineEdit, QDoubleSpinBox {{
                background-color: {colors.get('card_bg', '#ffffff')};
                color: {colors.get('text', '#212529')};
                border: 1px solid {colors.get('border', '#dee2e6')};
                border-radius: 5px;
                padding: 6px 8px;
            }}
        """)
        image_preview.setStyleSheet(f"""
            QLabel {{
                background-color: {colors.get('bg_hover', '#f8f9fa')};
                color: {colors.get('text_secondary', '#6c757d')};
                border: 1px dashed {colors.get('border', '#dee2e6')};
                border-radius: 6px;
                padding: 10px;
            }}
        """)
        product_details.setStyleSheet(f"""
            QLabel {{
                color: {colors.get('text_secondary', '#6c757d')};
                padding: 2px;
            }}
        """)

    def clear_expiry_discount(self):
        """Clear clearance discount from selected batches."""
        location_ids = self._get_selected_location_ids()
        if not location_ids:
            QMessageBox.warning(self, "No Selection", "Please select one or more batches.")
            return

        reply = QMessageBox.question(
            self,
            "Clear Expiry Discount",
            f"Clear expiry discount from {len(location_ids)} selected batch(es)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        if self._set_expiry_discount(location_ids, 0.0, "", "", "", enabled=False):
            QMessageBox.information(self, "Discount Cleared", "Expiry discount cleared.")
            self.refresh()
            self._refresh_discount_page()
        else:
            QMessageBox.critical(self, "Update Failed", "Could not clear expiry discount.")

    def _refresh_discount_page(self):
        main_window = self.window()
        discount_page = getattr(main_window, "discount_page", None)
        if discount_page and hasattr(discount_page, "load_discounts"):
            discount_page.load_discounts()

    def print_clearance_label(self):
        """Print internal clearance labels for selected batches."""
        location_ids = self._get_selected_location_ids()
        if not location_ids:
            QMessageBox.warning(self, "No Selection", "Please select one or more batches.")
            return

        rows = self._get_clearance_label_rows(location_ids)
        if not rows:
            QMessageBox.warning(self, "Not Found", "Selected batch data was not found.")
            self.refresh()
            return

        for row in rows:
            loc_id, product_name, price, location, batch_no, expire_date, discount_enabled, discount_percent = row
            price = float(price or 0)
            discount_percent = float(discount_percent or 0)
            final_price = price
            if discount_enabled and discount_percent > 0:
                final_price = max(0.0, price * (1 - min(discount_percent, 100) / 100.0))
            label_name = (
                f"{product_name} | EXP {expire_date or 'N/A'} | "
                f"{discount_percent:g}% OFF | {format_money(final_price, get_currency_symbol())}"
            )
            if batch_no:
                label_name = f"{label_name} | Batch {batch_no}"
            if location:
                label_name = f"{label_name} | {location}"
            dialog = PrintBarcodeDialog(loc_id, label_name, f"ZAYBATCH:{loc_id}", self)
            dialog.exec()

    def _get_clearance_label_rows(self, location_ids):
        conn = connect_db()
        cursor = conn.cursor()
        placeholders = ",".join("?" for _ in location_ids)
        cursor.execute(f"""
            SELECT
                pl.id,
                p.name,
                p.price,
                pl.location,
                pl.batch_no,
                pl.expire_date,
                COALESCE(pl.expiry_discount_enabled, 0),
                COALESCE(pl.expiry_discount_percent, 0)
            FROM product_locations pl
            JOIN products p ON p.id = pl.product_id
            WHERE pl.id IN ({placeholders})
            ORDER BY p.name, pl.expire_date
        """, location_ids)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def _set_expiry_discount(
        self,
        location_ids,
        percent: float,
        note: str,
        start_date: str = "",
        end_date: str = "",
        enabled: bool = True,
    ) -> bool:
        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("BEGIN IMMEDIATE")
            enabled_value = 1 if enabled and percent > 0 else 0
            percent_value = float(percent) if enabled_value else 0.0
            note_value = note if enabled_value else ""
            start_value = start_date if enabled_value else ""
            end_value = end_date if enabled_value else ""
            placeholders = ",".join("?" for _ in location_ids)
            cursor.execute(f"""
                UPDATE product_locations
                SET expiry_discount_enabled = ?,
                    expiry_discount_percent = ?,
                    expiry_discount_start_date = ?,
                    expiry_discount_end_date = ?,
                    clearance_note = ?,
                    last_updated = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders})
            """, [enabled_value, percent_value, start_value, end_value, note_value, *location_ids])

            cursor.execute(f"""
                SELECT DISTINCT product_id
                FROM product_locations
                WHERE id IN ({placeholders})
            """, location_ids)
            product_ids = [row[0] for row in cursor.fetchall()]
            for product_id in product_ids:
                cursor.execute("""
                    INSERT INTO stock_movements
                    (product_id, type, quantity, reason, reference, notes)
                    VALUES (?, 'adjustment', 0, 'Expiry Discount', ?, ?)
                """, (
                    product_id,
                    f"EXPIRY_DISCOUNT_{datetime.now().strftime('%Y%m%d%H%M%S')}",
                    f"{'Applied' if enabled_value else 'Cleared'} expiry discount: {percent_value:g}% ({start_value or '-'} to {end_value or '-'})"
                ))

            conn.commit()
            return True
        except Exception as e:
            logger.error(f"Failed to update expiry discount: {e}", exc_info=True)
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            return False
        finally:
            if conn:
                conn.close()

    def _get_expired_location_ids(self):
        """Get all expired location IDs"""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id FROM product_locations
            WHERE quantity > 0
              AND expire_date IS NOT NULL 
              AND expire_date != ''
              AND date(expire_date) < date('now')
        """)
        rows = cursor.fetchall()
        conn.close()
        return [row[0] for row in rows]

    def write_off_selected(self):
        """Write off selected expired items"""
        location_ids = self._get_selected_location_ids()
        
        if not location_ids:
            QMessageBox.warning(self, "No Selection", "Please select expired items to write off.")
            return
        
        # Check if all selected are expired
        lang = self.get_lang()
        expired_ids = self._get_expired_location_ids()
        non_expired = [lid for lid in location_ids if lid not in expired_ids]
        
        if non_expired:
            if lang == "my":
                msg = f"ရွေးချယ်ထားသော ပစ္စည်း {len(non_expired)} ခုသည် သက်တမ်းမကုန်သေးပါ။\nဆက်လုပ်မည်လား?"
            else:
                msg = f"{len(non_expired)} selected item(s) are not expired.\nContinue anyway?"
            reply = QMessageBox.question(self, "Warning", msg, QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply != QMessageBox.StandardButton.Yes:
                return
        
        # Confirm
        if lang == "my":
            msg = f"ရွေးချယ်ထားသော ပစ္စည်း {len(location_ids)} ခုကို စာရင်းမှထုတ်မည်။\nဆက်လုပ်မည်လား?"
        else:
            msg = f"Write off {len(location_ids)} selected item(s)?\nThis action will remove them from inventory."
        
        reply = QMessageBox.question(
            self, "Write Off Confirmation", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Execute write off
        success_count, failed_count = self._execute_write_off(location_ids, "Manual Write Off")
        
        # Show result
        if lang == "my":
            msg = f"စာရင်းမှထုတ်ပြီးပါပြီ။\nအောင်မြင်: {success_count}\nမအောင်မြင်: {failed_count}"
        else:
            msg = f"Write off completed.\nSuccess: {success_count}\nFailed: {failed_count}"
        
        QMessageBox.information(self, "Write Off Complete", msg)
        self.load_data()

    def write_off_all_expired(self):
        """Write off all expired items"""
        location_ids = self._get_expired_location_ids()
        
        if not location_ids:
            lang = self.get_lang()
            msg = "No expired items found." if lang != "my" else "သက်တမ်းကုန်ပြီးပစ္စည်းမရှိပါ။"
            QMessageBox.information(self, "No Expired Items", msg)
            return
        
        lang = self.get_lang()
        if lang == "my":
            msg = f"သက်တမ်းကုန်ပြီးပစ္စည်း {len(location_ids)} ခုကို စာရင်းမှထုတ်မည်။\nဆက်လုပ်မည်လား?"
        else:
            msg = f"Write off all {len(location_ids)} expired items?\nThis action cannot be undone."
        
        reply = QMessageBox.question(
            self, "Write Off All Expired", msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Execute with progress dialog
        progress = QProgressDialog(
            "Writing off expired items..." if lang != "my" else "သက်တမ်းကုန်ပစ္စည်းများကို စာရင်းမှထုတ်နေသည်...",
            "Cancel", 0, len(location_ids), self
        )
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.show()
        
        success_count = 0
        failed_count = 0
        
        for i, loc_id in enumerate(location_ids):
            progress.setValue(i)
            if progress.wasCanceled():
                break
            
            result = self._write_off_single_location(loc_id, "Bulk Write Off - All Expired")
            if result:
                success_count += 1
            else:
                failed_count += 1
        
        progress.setValue(len(location_ids))
        
        # Show result
        if lang == "my":
            msg = f"စာရင်းမှထုတ်ပြီးပါပြီ။\nအောင်မြင်: {success_count}\nမအောင်မြင်: {failed_count}"
        else:
            msg = f"Write off completed.\nSuccess: {success_count}\nFailed: {failed_count}"
        
        QMessageBox.information(self, "Write Off Complete", msg)
        self.load_data()

    def _write_off_single_location(self, location_id: int, reason: str = "Write Off") -> bool:
        """
        Write off a single location by setting quantity to 0.
        Creates stock movement log.
        
        Returns:
            bool: True if successful
        """
        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("BEGIN IMMEDIATE")
            
            # Get location info
            cursor.execute("""
                SELECT pl.product_id, pl.location, pl.quantity, pl.batch_no, pl.expire_date,
                       p.name, p.stock
                FROM product_locations pl
                JOIN products p ON pl.product_id = p.id
                WHERE pl.id = ?
            """, (location_id,))
            row = cursor.fetchone()
            
            if not row:
                if conn:
                    conn.rollback()
                return False
            
            product_id, location, qty, batch_no, expire_date, product_name, current_stock = row
            
            if qty <= 0:
                if conn:
                    conn.rollback()
                return True  # Already zero
            
            # Update product stock
            new_stock = current_stock - qty
            cursor.execute("""
                UPDATE products 
                SET stock = ?, last_updated = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (new_stock, product_id))
            
            # Delete location entry (or set to 0)
            cursor.execute("DELETE FROM product_locations WHERE id = ?", (location_id,))
            
            # Log stock movement
            cursor.execute("""
                INSERT INTO stock_movements 
                (product_id, type, quantity, old_stock, new_stock, reason, reference, notes, location)
                VALUES (?, 'write_off', ?, ?, ?, ?, 'EXPIRED_WRITE_OFF', ?, ?)
            """, (
                product_id,
                qty,
                current_stock,
                new_stock,
                reason,
                f"Batch: {batch_no or 'N/A'}, Expiry: {expire_date or 'N/A'}",
                location
            ))
            
            # Also log to expiry_alerts_log if table exists
            try:
                cursor.execute("""
                    INSERT INTO expiry_alerts_log 
                    (product_id, product_name, location, batch_no, expire_date, quantity, alert_type, message, is_resolved)
                    VALUES (?, ?, ?, ?, ?, ?, 'write_off', ?, 1)
                """, (
                    product_id,
                    product_name,
                    location,
                    batch_no or '',
                    expire_date or '',
                    qty,
                    f"Written off {qty} units due to expiry"
                ))
            except:
                pass  # Table might not exist
            
            if conn:
                conn.commit()
            logger.info(f"Write off: {product_name} - {qty} units from {location} (reason: {reason})")
            return True
            
        except Exception as e:
            logger.error(f"Write off failed for location {location_id}: {e}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            return False
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

    def _execute_write_off(self, location_ids: list, reason: str = "Write Off"):
        """Execute write off for multiple locations"""
        success_count = 0
        failed_count = 0
        
        for loc_id in location_ids:
            if self._write_off_single_location(loc_id, reason):
                success_count += 1
            else:
                failed_count += 1
        
        return success_count, failed_count

    # =========================================================================
    # CONTEXT MENU
    # =========================================================================

    def _show_context_menu(self, position):
        """Show context menu for table"""
        index = self.table.indexAt(position)
        if not index.isValid():
            return
        if not self.table.selectionModel().isRowSelected(index.row(), index.parent()):
            self.table.selectRow(index.row())
        
        menu = QMenu(self)
        lang = self.get_lang()
        
        # Edit expiry action
        edit_expiry_action = QAction("Edit Expiry Date" if lang != "my" else "သက်တမ်းကုန်ရက် ပြင်မည်", self)
        edit_expiry_action.triggered.connect(self.edit_selected_expiry_date)
        edit_expiry_action.setEnabled(len(self._get_selected_location_ids()) == 1)
        menu.addAction(edit_expiry_action)

        apply_discount_action = QAction("Apply Expiry Discount", self)
        apply_discount_action.triggered.connect(self.apply_expiry_discount)
        menu.addAction(apply_discount_action)

        clear_discount_action = QAction("Clear Expiry Discount", self)
        clear_discount_action.triggered.connect(self.clear_expiry_discount)
        menu.addAction(clear_discount_action)

        print_label_action = QAction("Print Clearance Label", self)
        print_label_action.triggered.connect(self.print_clearance_label)
        menu.addAction(print_label_action)
        menu.addSeparator()

        # Write Off action
        write_off_action = QAction("Write Off Selected" if lang != "my" else "ရွေးချယ်ထားသော ပစ္စည်းများကို စာရင်းမှထုတ်မည်", self)
        write_off_action.triggered.connect(self.write_off_selected)
        menu.addAction(write_off_action)
        
        # Refresh action
        refresh_action = QAction("Refresh" if lang != "my" else "ပြန်လည်ဆန်းသစ်မည်", self)
        refresh_action.triggered.connect(self.refresh)
        menu.addAction(refresh_action)
        
        viewport = self.table.viewport()
        if viewport is not None:
            menu.exec(viewport.mapToGlobal(position))

    # =========================================================================
    # RESTORE WRITTEN OFF ITEMS
    # =========================================================================

    def restore_written_off_item(self, movement_id: int):
        """
        Restore a previously written off item.
        This reverses the write-off stock movement.
        """
        conn = None
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Get movement details
            cursor.execute("""
                SELECT product_id, quantity, location, notes
                FROM stock_movements
                WHERE id = ? AND type = 'write_off'
            """, (movement_id,))
            row = cursor.fetchone()
            
            if not row:
                return False, "Movement not found"
            
            product_id, qty, location, notes = row
            
            cursor.execute("BEGIN IMMEDIATE")
            
            # Restore to product_locations
            cursor.execute("""
                INSERT INTO product_locations (product_id, location, quantity)
                VALUES (?, ?, ?)
                ON CONFLICT(product_id, location, batch_no, expire_date)
                DO UPDATE SET quantity = quantity + excluded.quantity,
                              last_updated = CURRENT_TIMESTAMP
            """, (product_id, location, qty))
            
            # Update product stock
            cursor.execute("""
                UPDATE products 
                SET stock = stock + ?, last_updated = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (qty, product_id))
            
            # Log restoration
            cursor.execute("""
                INSERT INTO stock_movements 
                (product_id, type, quantity, reason, reference, notes, location)
                VALUES (?, 'restore', ?, 'Restore from write-off', ?, ?, ?)
            """, (product_id, qty, f"RESTORE_{movement_id}", notes, location))
            
            if conn:
                conn.commit()
            return True, "Restored successfully"
            
        except Exception as e:
            logger.error(f"Restore failed: {e}")
            if conn:
                try:
                    conn.rollback()
                except:
                    pass
            return False, str(e)
        finally:
            if conn:
                try:
                    conn.close()
                except:
                    pass

    # =========================================================================
    # EXPORT
    # =========================================================================

    def get_all_data(self):
        """Get all data for export."""
        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                p.name as product_name,
                pl.batch_no,
                pl.location,
                pl.expire_date,
                pl.quantity,
                pl.last_updated,
                p.sku,
                p.category,
                p.price,
                pl.id as location_id
            FROM product_locations pl
            JOIN products p ON pl.product_id = p.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND pl.quantity > 0
            ORDER BY 
                CASE 
                    WHEN pl.expire_date IS NULL OR pl.expire_date = '' THEN 1 
                    ELSE 0 
                END,
                pl.expire_date ASC,
                pl.last_updated ASC
        """)

        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        """Export expiry report to Excel."""
        lang = self.get_lang()

        file_path = ExcelExporter.save_file_dialog(
            self,
            f"expiry_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Expiry Report" if lang != "my" else "သက်တမ်းကုန်ရက်အစီရင်ခံစာ ထုတ်ရန်"
        )
        if not file_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            rows = self.get_all_data()
            symbol = get_currency_symbol()
            today = date.today()

            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise RuntimeError("Unable to create Excel worksheet")
            ws.title = "Expiry Report"

            # Title
            ws.merge_cells('A1:L1')
            cell_a1 = ws['A1']
            if cell_a1 is not None:
                cell_a1.value = "EXPIRY REPORT (BATCH LEVEL)" if lang != "my" else "သက်တမ်းကုန်ရက်အစီရင်ခံစာ (အသုတ်အလိုက်)"
                cell_a1.font = Font(bold=True, size=14)
                cell_a1.alignment = Alignment(horizontal="center")

            cell_a2 = ws['A2']
            if cell_a2 is not None:
                cell_a2.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                cell_a2.font = Font(size=10, color="7f8c8d")
            cell_a3 = ws['A3']
            if cell_a3 is not None:
                cell_a3.value = f"Total Batches: {len(rows)}"
                cell_a3.font = Font(size=10, color="7f8c8d")

            # Headers
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "အမျိုးအစား", "အသုတ်အမှတ်", 
                          "နေရာ", "သက်တမ်းကုန်ရက်", "ကျန်ရက်များ", "အရေအတွက်", 
                          "အခြေအနေ", "စျေးနှုန်း", "စုစုပေါင်းတန်ဖိုး", "Location ID"]
            else:
                headers = ["Product Name", "SKU", "Category", "Batch No", 
                          "Location", "Expiry Date", "Days Left", "Quantity", 
                          "Status", "Price", "Total Value", "Location ID"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                if cell is not None:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

            total_qty = 0
            total_value = 0
            expired_count = 0
            expiring_count = 0

            for row_idx, row_data in enumerate(rows, start=6):
                name, batch_no, location, expire_date, qty, last_updated, sku, category, price, location_id = row_data

                # Calculate days left and status
                if expire_date and expire_date.strip():
                    try:
                        exp_d = date.fromisoformat(expire_date)
                        days_left = (exp_d - today).days
                        if days_left < 0:
                            status = "Expired"
                            expired_count += 1
                        elif days_left <= 7:
                            status = "Expiring Soon"
                            expiring_count += 1
                        else:
                            status = "OK"
                    except:
                        days_left = "Invalid"
                        status = "Invalid Date"
                else:
                    days_left = "N/A"
                    status = "No Expiry"

                total_value_row = (price or 0) * qty

                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=category or "")
                ws.cell(row=row_idx, column=4, value=batch_no or "-")
                ws.cell(row=row_idx, column=5, value=location or "-")
                ws.cell(row=row_idx, column=6, value=expire_date or "-")
                ws.cell(row=row_idx, column=7, value=days_left if isinstance(days_left, int) else days_left)
                ws.cell(row=row_idx, column=8, value=qty)
                ws.cell(row=row_idx, column=9, value=status)
                ws.cell(row=row_idx, column=10, value=float(price or 0))
                ws.cell(row=row_idx, column=11, value=float(total_value_row))
                ws.cell(row=row_idx, column=12, value=location_id or "")

                total_qty += qty
                total_value += total_value_row

                # Color code status
                status_cell = ws.cell(row=row_idx, column=9)
                if status_cell is not None:
                    if status == "Expired":
                        status_cell.font = Font(color="FF0000", bold=True)
                    elif status == "Expiring Soon":
                        status_cell.font = Font(color="FF8C00", bold=True)

            # Summary rows
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=7, value="SUMMARY").font = Font(bold=True, size=12)
            ws.cell(row=summary_row + 1, column=7, value=f"Total Batches: {len(rows)}")
            ws.cell(row=summary_row + 2, column=7, value=f"Expired: {expired_count}")
            ws.cell(row=summary_row + 3, column=7, value=f"Expiring Soon: {expiring_count}")
            ws.cell(row=summary_row + 4, column=7, value=f"Total Quantity: {total_qty}")
            ws.cell(row=summary_row + 5, column=7, value=f"Total Value: {format_money(total_value, symbol)}")

            # Auto adjust columns
            for col in range(1, 13):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['A'].width = 30

            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)

        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def retranslateUi(self):
        """Retranslate UI."""
        lang = self.get_lang()
        
        if lang == "my":
            self.search_widget.set_placeholder_text("ပစ္စည်း / အသုတ် / နေရာဖြင့် ရှာရန်...")
            self.btn_write_off.setText(" ရွေးချယ်ထားသော ပစ္စည်းများကို စာရင်းမှထုတ်မည်")
            self.btn_write_off_all.setText(" သက်တမ်းကုန်ပြီးအားလုံးကို စာရင်းမှထုတ်မည်")
            self.btn_export.setText(" Excel ထုတ်မည်")
        else:
            self.search_widget.set_placeholder_text("Search by product, batch, or location...")
            self.btn_write_off.setText(" Write Off Selected")
            self.btn_write_off_all.setText(" Write Off All Expired")
            self.btn_export.setText(" Export Excel")
        
        # Update card titles
        if lang == "my":
            self.total_card.set_title("စုစုပေါင်းအသုတ်များ")
            self.expired_card.set_title("သက်တမ်းကုန်ပြီးအသုတ်များ")
            self.expiring_card.set_title("သက်တမ်းနီးပြီး (ရက်ပေါင်း ၇)")
            self.total_qty_card.set_title("စုစုပေါင်းပမာဏ")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "သက်တမ်းကုန်ပြီ")
            self.status_filter.setItemText(2, "သက်တမ်းနီးပြီ")
            self.status_filter.setItemText(3, "ကောင်းသည်")
        else:
            self.total_card.set_title("Total Batches")
            self.expired_card.set_title("Expired Batches")
            self.expiring_card.set_title("Expiring Soon (7 days)")
            self.total_qty_card.set_title("Total Quantity")
            self.status_filter.setItemText(0, "All")
            self.status_filter.setItemText(1, "Expired")
            self.status_filter.setItemText(2, "Expiring Soon")
            self.status_filter.setItemText(3, "OK")

        self._update_button_icons()
        self._update_card_icons()
        self._set_headers()
        self.load_data()

    def showEvent(self, event):
        """Handle show event."""
        self.load_locations()
        self.load_data()
        super().showEvent(event)

    # =========================================================================
    # ADDITIONAL UTILITY METHODS
    # =========================================================================

    def load_expired_summary(self):
        """Load summary of expired stock for dashboard/alerts."""
        conn = connect_db()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT COUNT(*), COALESCE(SUM(quantity), 0)
            FROM product_locations
            WHERE quantity > 0
              AND expire_date IS NOT NULL 
              AND expire_date != ''
              AND date(expire_date) < date('now')
        """)
        expired_count, expired_qty = cursor.fetchone()
        
        cursor.execute("""
            SELECT COUNT(*), COALESCE(SUM(quantity), 0)
            FROM product_locations
            WHERE quantity > 0
              AND expire_date IS NOT NULL 
              AND expire_date != ''
              AND date(expire_date) >= date('now')
              AND date(expire_date) <= date('now', '+7 days')
        """)
        expiring_count, expiring_qty = cursor.fetchone()
        
        conn.close()
        
        return {
            'expired_count': expired_count or 0,
            'expired_qty': expired_qty or 0,
            'expiring_count': expiring_count or 0,
            'expiring_qty': expiring_qty or 0
        }

    def get_expired_stock_detail(self):
        """Get detailed list of expired stock."""
        return get_all_expired_stock()

    def get_expiring_soon_detail(self, days=7):
        """Get detailed list of expiring soon stock."""
        return get_expiring_soon_stock(days)
    
    def get_write_off_history(self, limit=100):
        """Get write-off history from stock movements."""
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                sm.id,
                p.name as product_name,
                sm.quantity,
                sm.reason,
                sm.notes,
                sm.location,
                sm.created_at,
                sm.created_by
            FROM stock_movements sm
            JOIN products p ON sm.product_id = p.id
            WHERE sm.type = 'write_off'
            ORDER BY sm.created_at DESC
            LIMIT ?
        """, (limit,))
        rows = cursor.fetchall()
        conn.close()
        return rows
