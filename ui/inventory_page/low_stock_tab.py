# ui/inventory_page/low_stock_tab.py
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QFileDialog, QHBoxLayout, QMessageBox, QLabel, QComboBox
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QColor, QIcon, QPixmap, QPainter
from models.database import connect_db
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.modern_button import ModernButton
from ui.widgets.search_widget import SearchWidget
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime
import os


class LowStockTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        self.current_page = 1
        self.page_size = 25
        self._is_dark = is_dark_theme()
        
        layout = QVBoxLayout()
        layout.setSpacing(12)

        # Top button layout
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)
        
        # ✅ Export button with SVG icon
        self.btn_export = ModernButton(" Export Low Stock Report", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(False)
        self.btn_export.clicked.connect(self.export_to_excel)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_export)
        layout.addLayout(btn_layout)

        # ✅ Filter section with SearchWidget
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(0, 8, 0, 8)
        
        # ✅ SearchWidget with SVG icon
        self.search_widget = SearchWidget(
            placeholder="Search product name or SKU...",
            show_label=False
        )
        self.search_widget.search_changed.connect(self.on_filter_changed)
        self.search_widget.search_cleared.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_widget, 2)
        
        # Status filter
        filter_layout.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Critical", "Warning"])
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.status_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        
        # ✅ NO custom table style - use PyQt6 default
        # self._apply_table_theme()  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.table)

        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.refresh()
        self.retranslateUi()

    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        # ✅ Only update button icons and reload data - no table style update
        self._update_button_icons()
        self.load_data()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_export.set_icon("file_export", size=(16, 16))

    def on_page_changed(self, page: int, page_size: int):
        self.current_page = page
        self.page_size = page_size
        self.load_data()

    def on_filter_changed(self):
        self.current_page = 1
        self.load_data()

    def refresh(self):
        self.load_data()
        self.retranslateUi()

    def load_data(self, page=None, page_size=None):
        if page is None:
            page = self.current_page
        if page_size is None:
            page_size = self.page_size
            
        lang = self.get_lang()
        search_text = self.search_widget.get_text().lower()
        status_filter = self.status_filter.currentText()
        
        if lang == "my":
            headers = [
                "ပစ္စည်းအမည်", "SKU", "လက်ကျန်", "အနည်းဆုံးပမာဏ", "ပြန်မှာသင့်ပမာဏ",
                "ပေးသွင်းသူ", "နောက်ဆုံးဝယ်ယူရက်", "အခြေအနေ"
            ]
        else:
            headers = [
                "Product Name", "SKU", "Current Qty", "Minimum Qty", "Suggested Reorder Qty",
                "Supplier", "Last Purchase Date", "Status"
            ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        conn = connect_db()
        cursor = conn.cursor()
        
        # Build query with search and status filters
        base_query = """
            FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)
        """
        params = []
        
        if search_text:
            like = f'%{search_text}%'
            base_query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ?)"
            params.extend([like, like])
        
        if status_filter == "Critical":
            base_query += " AND COALESCE(p.stock, 0) = 0"
        elif status_filter == "Warning":
            base_query += " AND COALESCE(p.stock, 0) > 0"
        
        # Count total
        count_query = f"SELECT COUNT(*) {base_query}"
        cursor.execute(count_query, params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        data_query = f"""
            SELECT p.name, p.sku, p.stock, p.low_stock, 
                   (COALESCE(p.low_stock, 0) * 2) as suggested,
                   s.name as supplier,
                   (SELECT MAX(created_at) FROM stock_movements 
                    WHERE product_id=p.id AND type='in') as last_purchase,
                   CASE 
                       WHEN COALESCE(p.stock, 0) = 0 THEN 'Critical' 
                       ELSE 'Warning' 
                   END as status
            {base_query}
            ORDER BY p.stock ASC
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_query, params + [page_size, offset])
        rows = cursor.fetchall()
        conn.close()

        self.table.setRowCount(0)
        for row in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)
            
            for col, val in enumerate(row):
                item = QTableWidgetItem(str(val) if val else "")
                
                # ✅ Color code status column with hardcoded colors
                if col == 7:  # Status column
                    if val == "Critical":
                        item.setForeground(QColor("#dc3545"))  # Red
                    elif val == "Warning":
                        item.setForeground(QColor("#f39c12"))  # Orange
                
                self.table.setItem(r, col, item)

    def get_all_low_stock_data(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT p.name, p.sku, p.stock, p.low_stock, 
                   (COALESCE(p.low_stock, 0) * 2) as suggested,
                   s.name as supplier,
                   (SELECT MAX(created_at) FROM stock_movements 
                    WHERE product_id=p.id AND type='in') as last_purchase,
                   CASE 
                       WHEN COALESCE(p.stock, 0) = 0 THEN 'Critical' 
                       ELSE 'Warning' 
                   END as status
            FROM products p
            LEFT JOIN suppliers s ON p.supplier_id = s.id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
              AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)
            ORDER BY p.stock ASC
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"low_stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Low Stock Report" if lang != "my" else "စတော့နည်းနေသောစာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_low_stock_data()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Low Stock Alert"
            
            ws.merge_cells('A1:G1')
            ws['A1'] = "LOW STOCK ALERT REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Products with Low Stock: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "လက်ကျန်", "အနည်းဆုံးပမာဏ", 
                          "ပြန်မှာသင့်ပမာဏ", "ပေးသွင်းသူ", "အခြေအနေ"]
            else:
                headers = ["Product Name", "SKU", "Current Stock", "Min Stock", 
                          "Suggested Order", "Supplier", "Status"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            critical_count = 0
            warning_count = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                name, sku, stock, low_stock, suggested, supplier, last_purchase, status = row_data
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=stock)
                ws.cell(row=row_idx, column=4, value=low_stock)
                ws.cell(row=row_idx, column=5, value=suggested)
                ws.cell(row=row_idx, column=6, value=supplier or "No Supplier")
                
                status_cell = ws.cell(row=row_idx, column=7, value=status)
                if status == "Critical":
                    status_cell.font = Font(color="FF0000", bold=True)
                    critical_count += 1
                else:
                    status_cell.font = Font(color="FF8C00", bold=True)
                    warning_count += 1
            
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=5, value="SUMMARY").font = Font(bold=True, size=12)
            ws.cell(row=summary_row + 1, column=5, value=f"Critical (Out of Stock): {critical_count}")
            ws.cell(row=summary_row + 2, column=5, value=f"Warning (Low Stock): {warning_count}")
            ws.cell(row=summary_row + 3, column=5, value=f"Total: {len(rows)}")
            
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].width = 18
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def retranslateUi(self):
        lang = self.get_lang()
        
        # Update SearchWidget placeholder
        if lang == "my":
            self.search_widget.set_placeholder_text("ပစ္စည်းအမည် သို့မဟုတ် SKU ဖြင့် ရှာရန်...")
            self.btn_export.setText(" စတော့နည်းနေသောစာရင်းထုတ်မည်")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "ကုန်သွားပြီ")
            self.status_filter.setItemText(2, "စတော့နည်းနေပြီ")
        else:
            self.search_widget.set_placeholder_text("Search product name or SKU...")
            self.btn_export.setText(" Export Low Stock Report")
            self.status_filter.setItemText(0, "All")
            self.status_filter.setItemText(1, "Critical")
            self.status_filter.setItemText(2, "Warning")
        
        # Update button icons
        self._update_button_icons()
        
        self.load_data()
    
    def showEvent(self, event):
        """Handle show event - refresh data"""
        self.load_data()
        super().showEvent(event)