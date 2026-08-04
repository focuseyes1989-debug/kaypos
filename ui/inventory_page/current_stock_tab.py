# ui/inventory_page/current_stock_tab.py
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QMessageBox, QFileDialog, QLabel, QComboBox
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QColor, QIcon, QPixmap, QPainter
from models.database import connect_db
from models.database.queries import reverse_stock_movement
from utils.currency import format_money, get_currency_symbol
from utils.translations import tr
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.modern_button import ModernButton
from ui.widgets.search_widget import SearchWidget
from ui.inventory_page.product_transaction_history_dialog import ProductTransactionHistoryDialog
from ui.product_detail_dialog import ProductDetailDialog
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime
import os


class CurrentStockTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_page = parent
        self.current_page = 1
        self.page_size = 50
        self.selected_product_id = None
        self.selected_product_name = None
        self._is_dark = is_dark_theme()
        
        layout = QVBoxLayout()
        layout.setSpacing(12)

        # Top button layout
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)
        
        # ✅ Stock In button with SVG icon
        self.btn_stock_in = ModernButton(" " + tr("stock_in"), ModernButton.PRIMARY)
        self.btn_stock_in.set_icon("add", size=(16, 16))
        self.btn_stock_in.set_compact(False)
        self.btn_stock_in.clicked.connect(self.open_stock_in)
        btn_layout.addWidget(self.btn_stock_in)
        
        # ✅ Stock Out button with SVG icon
        self.btn_stock_out = ModernButton(" " + tr("stock_out"), ModernButton.PRIMARY)
        self.btn_stock_out.set_icon("remove", size=(16, 16))
        self.btn_stock_out.set_compact(False)
        self.btn_stock_out.clicked.connect(self.open_stock_out)
        btn_layout.addWidget(self.btn_stock_out)
        
        # ✅ Adjustment button with SVG icon
        self.btn_adjustment = ModernButton(" " + tr("adjustment"), ModernButton.PRIMARY)
        self.btn_adjustment.set_icon("edit", size=(16, 16))
        self.btn_adjustment.set_compact(False)
        self.btn_adjustment.clicked.connect(self.open_adjustment)
        btn_layout.addWidget(self.btn_adjustment)
        
        # ✅ Transfer button with SVG icon
        self.btn_transfer = ModernButton(" Transfer", ModernButton.PRIMARY)
        self.btn_transfer.set_icon("swap_horiz", size=(16, 16))
        self.btn_transfer.set_compact(False)
        self.btn_transfer.clicked.connect(self.open_transfer)
        btn_layout.addWidget(self.btn_transfer)
        
        # ✅ View Movements button with SVG icon
        self.btn_view_movements = ModernButton(" View Movements", ModernButton.SECONDARY)
        self.btn_view_movements.set_icon("history", size=(16, 16))
        self.btn_view_movements.set_compact(False)
        self.btn_view_movements.clicked.connect(self.show_stock_movements)
        btn_layout.addWidget(self.btn_view_movements)
        
        btn_layout.addStretch()
        
        # ✅ Export button with SVG icon
        self.btn_export = ModernButton(" Export Current Stock", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(False)
        self.btn_export.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(self.btn_export)
        
        layout.addLayout(btn_layout)

        # Filter section with SearchWidget
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(0, 8, 0, 8)
        
        # ✅ Use SearchWidget with SVG icon
        self.search_widget = SearchWidget(
            placeholder="Search by name, SKU or barcode...",
            show_label=False
        )
        self.search_widget.search_changed.connect(self.on_filter_changed)
        self.search_widget.search_cleared.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_widget, 2)
        
        # Category filter
        filter_layout.addWidget(QLabel("Category:"))
        self.category_filter = QComboBox()
        self.category_filter.addItem("All Categories")
        self.category_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.category_filter, 1)
        
        # Status filter
        filter_layout.addWidget(QLabel("Status:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All Status", "In Stock", "Low Stock", "Out of Stock"])
        self.status_filter.currentTextChanged.connect(self.on_filter_changed)
        filter_layout.addWidget(self.status_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.stock_table = QTableWidget()
        self.stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.stock_table.verticalHeader().setDefaultSectionSize(52)
        self.stock_table.verticalHeader().setMinimumSectionSize(48)
        self.stock_table.verticalHeader().setVisible(True)
        self.stock_table.setAlternatingRowColors(True)
        
        self.stock_table.cellClicked.connect(self.on_cell_clicked)
        self.stock_table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        
        # ✅ NO custom table style - use PyQt6 default
        # self._apply_table_theme()  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.stock_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        layout.addWidget(self.stock_table)

        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)

        self.setLayout(layout)
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.load_categories()
        self.refresh()
        self.retranslateUi()

    def _centered_cell_widget(self, widget):
        container = QWidget()
        container.setStyleSheet("background: transparent; border: none;")
        container.setMinimumHeight(48)
        layout = QGridLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(widget, 0, 0, alignment=Qt.AlignmentFlag.AlignCenter)
        return container

    def _get_themed_icon(self, icon_name, size=(16, 16)):
        """Get themed SVG icon"""
        try:
            from ui.themes.theme_manager import get_themed_icon
            return get_themed_icon(icon_name, size=size)
        except:
            return QIcon()
    
    def _load_colored_icon(self, icon_name, size=(16, 16)):
        """Load SVG icon with color based on theme"""
        is_dark = is_dark_theme()
        color_hex = "#ffffff" if is_dark else "#495057"
        
        # Try SVG first
        paths = [
            f"assets/icons/{icon_name}.svg",
            f"assets/icons/{icon_name}.png",
        ]
        
        for path in paths:
            if os.path.exists(path):
                try:
                    pixmap = QPixmap(path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(
                            size[0], size[1],
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        
                        # Color the icon
                        colored = scaled.copy()
                        painter = QPainter(colored)
                        painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
                        painter.fillRect(colored.rect(), QColor(color_hex))
                        painter.end()
                        
                        return QIcon(colored)
                except Exception as e:
                    print(f"Could not load icon {path}: {e}")
        
        return QIcon()

    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        # ✅ Only update button icons and reload data - no table style update
        self._update_button_icons()
        self.load_data()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_stock_in.set_icon("add", size=(16, 16))
        self.btn_stock_out.set_icon("remove", size=(16, 16))
        self.btn_adjustment.set_icon("edit", size=(16, 16))
        self.btn_transfer.set_icon("swap_horiz", size=(16, 16))
        self.btn_view_movements.set_icon("history", size=(16, 16))
        self.btn_export.set_icon("file_export", size=(16, 16))

    def on_cell_clicked(self, row, column):
        id_item = self.stock_table.item(row, 0)
        name_item = self.stock_table.item(row, 1)
        if id_item:
            try:
                self.selected_product_id = int(id_item.text())
                if name_item:
                    self.selected_product_name = name_item.text()
                else:
                    self.selected_product_name = None
            except ValueError:
                self.selected_product_id = None
                self.selected_product_name = None

    def get_selected_product(self):
        return self.selected_product_id, self.selected_product_name

    def load_categories(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM categories ORDER BY name")
        rows = cursor.fetchall()
        self.category_filter.blockSignals(True)
        current = self.category_filter.currentText()
        self.category_filter.clear()
        self.category_filter.addItem("All Categories")
        for (name,) in rows:
            self.category_filter.addItem(name)
        idx = self.category_filter.findText(current)
        if idx >= 0:
            self.category_filter.setCurrentIndex(idx)
        else:
            self.category_filter.setCurrentIndex(0)
        self.category_filter.blockSignals(False)
        conn.close()

    def on_filter_changed(self):
        self.current_page = 1
        self.load_data()

    def open_stock_in(self):
        from ui.inventory_page.stock_in_dialog import StockInDialog
        dialog = StockInDialog(self)
        product_id, product_name = self.get_selected_product()
        if product_id:
            dialog.set_product(product_id, product_name)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_stock_out(self):
        from ui.inventory_page.stock_out_dialog import StockOutDialog
        dialog = StockOutDialog(self)
        product_id, product_name = self.get_selected_product()
        if product_id:
            dialog.set_product(product_id, product_name)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_adjustment(self):
        from ui.inventory_page.adjustment_dialog import AdjustmentDialog
        dialog = AdjustmentDialog(self)
        product_id, product_name = self.get_selected_product()
        if product_id:
            dialog.set_product(product_id, product_name)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def open_transfer(self):
        from ui.inventory_page.stock_transfer_dialog import StockTransferDialog
        dialog = StockTransferDialog(self)
        product_id, product_name = self.get_selected_product()
        if product_id:
            dialog.set_product(product_id, product_name)
        if dialog.exec():
            self.refresh()
            self.refresh_stock_alerts()

    def show_stock_movements(self):
        product_id, product_name = self.get_selected_product()
        if not product_id:
            lang = self.get_lang()
            msg = "Please select a product first." if lang != "my" else "ကျေးဇူးပြု၍ ပစ္စည်းတစ်ခုကို ဦးစွာရွေးချယ်ပါ။"
            QMessageBox.warning(self, "No Selection" if lang != "my" else "မရွေးရသေး", msg)
            return
        
        from ui.inventory_page.stock_movement_dialog import StockMovementDialog
        dialog = StockMovementDialog(product_id, self)
        dialog.movement_reversed.connect(self.on_movement_reversed)
        dialog.exec()

    def on_movement_reversed(self):
        self.refresh()
        self.refresh_stock_alerts()
        lang = self.get_lang()
        msg = "Stock movement reversed successfully." if lang != "my" else "စတော့လှုပ်ရှားမှုကို အောင်မြင်စွာ ပြန်ဖျက်ပြီးပါပြီ။"
        QMessageBox.information(self, "Updated" if lang != "my" else "ပြင်ဆင်ပြီး", msg)

    def reverse_last_stock_in(self):
        product_id, product_name = self.get_selected_product()
        if not product_id:
            lang = self.get_lang()
            msg = "Please select a product first." if lang != "my" else "ကျေးဇူးပြု၍ ပစ္စည်းတစ်ခုကို ဦးစွာရွေးချယ်ပါ။"
            QMessageBox.warning(self, "No Selection" if lang != "my" else "မရွေးရသေး", msg)
            return
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, created_at, quantity, location, created_by
            FROM stock_movements
            WHERE product_id = ? AND type = 'in'
            ORDER BY created_at DESC
            LIMIT 1
        """, (product_id,))
        last_in = cursor.fetchone()
        conn.close()
        
        if not last_in:
            lang = self.get_lang()
            msg = "No stock in record found for this product." if lang != "my" else "ဤပစ္စည်းအတွက် စတော့ဝင်မှတ်တမ်း မတွေ့ပါ။"
            QMessageBox.warning(self, "No Stock In" if lang != "my" else "စတော့ဝင်မှတ်တမ်းမရှိ", msg)
            return
        
        mov_id, created_at, qty, location, created_by = last_in
        lang = self.get_lang()
        
        if lang == "my":
            msg = (f"'{product_name}' အတွက် နောက်ဆုံး စတော့ဝင်မှုကို ပြန်ဖျက်မည်လား?\n\n"
                   f"ရက်စွဲ: {created_at}\n"
                   f"ပမာဏ: {qty}\n"
                   f"နေရာ: {location or 'N/A'}\n"
                   f"ဖန်တီးသူ: {created_by or 'System'}\n\n"
                   f"ဤသည်မှာ စတော့မှ {qty} ကို ဖယ်ရှားမည်ဖြစ်သည်။")
        else:
            msg = (f"Reverse the last Stock In for '{product_name}'?\n\n"
                   f"Date: {created_at}\n"
                   f"Quantity: {qty}\n"
                   f"Location: {location or 'N/A'}\n"
                   f"Created by: {created_by or 'System'}\n\n"
                   f"This will remove {qty} from stock.")
        
        reply = QMessageBox.question(
            self,
            "Reverse Stock In" if lang != "my" else "စတော့ဝင်မှုကို ပြန်ဖျက်ရန်",
            msg,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            main_window = self.window()
            created_by = main_window.current_user['username'] if hasattr(main_window, 'current_user') else 'System'
            
            result = reverse_stock_movement(mov_id, reason="User requested reversal", created_by=created_by)
            if result['success']:
                self.on_movement_reversed()
            else:
                QMessageBox.critical(self, "Error", result['message'])

    def refresh_stock_alerts(self):
        main_window = self.window()
        if hasattr(main_window, 'check_stock_alerts'):
            main_window.check_stock_alerts()

    def on_page_changed(self, page: int, page_size: int):
        self.current_page = page
        self.page_size = page_size
        self.load_data()

    def refresh(self):
        self.load_data()
        self.retranslateUi()

    def on_cell_double_clicked(self, row, column):
        id_item = self.stock_table.item(row, 0)
        if id_item:
            try:
                product_id = int(id_item.text())
                dialog = ProductDetailDialog(product_id)
                dialog.exec()
            except ValueError:
                pass

    def load_data(self, page=None, page_size=None):
        if page is None:
            page = self.current_page
        if page_size is None:
            page_size = self.page_size
            
        lang = self.get_lang()
        search_text = self.search_widget.get_text().lower()
        category = self.category_filter.currentText()
        status_filter = self.status_filter.currentText()
        use_category = category != "All Categories"
        
        if lang == "my":
            main_headers = [
                "ID", "ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", "လက်ကျန်",
                "ကုန်ကျစရိတ်", "ရောင်းဈေး", "စုစုပေါင်းတန်ဖိုး", "သတိပေးပမာဏ",
                "အခြေအနေ", "နောက်ဆုံးပြင်ဆင်ချိန်", "နေရာ"
            ]
        else:
            main_headers = [
                "ID", "Product Name", "SKU", "Barcode", "Category", "Current Qty",
                "Cost Price", "Selling Price", "Stock Value", "Low Stock Level",
                "Status", "Last Updated", "Location"
            ]
        
        # ✅ Add History header
        headers = main_headers + (["မှတ်တမ်း"] if lang == "my" else ["History"])
        self.stock_table.setColumnCount(len(headers))
        self.stock_table.setHorizontalHeaderLabels(headers)
        
        self.stock_table.setColumnHidden(0, True)
        
        header = self.stock_table.horizontalHeader()
        for col in range(1, len(headers) - 1):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        
        history_col = len(headers) - 1
        header.setSectionResizeMode(history_col, QHeaderView.ResizeMode.Fixed)
        self.stock_table.setColumnWidth(history_col, 110)

        conn = connect_db()
        cursor = conn.cursor()
        
        base_query = """
            FROM products p
            LEFT JOIN product_locations pl ON p.id = pl.product_id
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
        """
        params = []
        
        if use_category:
            base_query += " AND p.category = ?"
            params.append(category)
        
        if search_text:
            like = f'%{search_text}%'
            base_query += " AND (LOWER(p.name) LIKE ? OR LOWER(p.sku) LIKE ? OR LOWER(p.barcode) LIKE ?)"
            params.extend([like, like, like])
        
        if status_filter == "In Stock" or status_filter == "စတော့ရှိပါ":
            base_query += " AND COALESCE(p.stock, 0) > COALESCE(p.low_stock, 0)"
        elif status_filter == "Low Stock" or status_filter == "စတော့နည်းနေပြီ":
            base_query += " AND COALESCE(p.stock, 0) > 0 AND COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0)"
        elif status_filter == "Out of Stock" or status_filter == "ကုန်သွားပြီ":
            base_query += " AND COALESCE(p.stock, 0) = 0"
        
        count_query = f"SELECT COUNT(DISTINCT p.id) {base_query}"
        cursor.execute(count_query, params)
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)

        offset = (page - 1) * page_size
        
        cursor.execute(f"""
            SELECT 
                p.id, 
                p.name, 
                p.sku, 
                p.barcode, 
                p.category,
                COALESCE(p.stock, 0) as total_stock,
                COALESCE(p.cost, 0) as cost,
                COALESCE(p.price, 0) as price,
                (COALESCE(p.cost, 0) * COALESCE(p.stock, 0)) as stock_value,
                COALESCE(p.low_stock, 0) as low_stock,
                p.sold_by,
                CASE 
                    WHEN p.sold_by = 'Service' THEN 'Service'
                    WHEN COALESCE(p.stock, 0) = 0 THEN 'Out of Stock'
                    WHEN COALESCE(p.stock, 0) <= COALESCE(p.low_stock, 0) THEN 'Low Stock'
                    ELSE 'In Stock'
                END as status,
                strftime('%Y-%m-%d %H:%M', p.last_updated) as last_upd,
                GROUP_CONCAT(pl.location || ' (' || pl.quantity || ')', ', ') as locations
            {base_query}
            GROUP BY p.id
            ORDER BY p.name
            LIMIT ? OFFSET ?
        """, params + [page_size, offset])
        rows = cursor.fetchall()
        conn.close()

        # ✅ Load history icon for buttons
        history_icon = self._load_colored_icon("history", size=(14, 14))
        
        self.stock_table.setRowCount(0)
        for row in rows:
            prod_id = row[0]
            name = row[1]
            sku = row[2]
            barcode = row[3]
            category = row[4]
            stock = row[5]
            cost = row[6]
            price = row[7]
            stock_value = row[8]
            low_stock = row[9]
            sold_by = row[10]
            status = row[11]
            last_upd = row[12]
            locations = row[13]
            
            r = self.stock_table.rowCount()
            self.stock_table.insertRow(r)
            self.stock_table.setRowHeight(r, 52)
            
            # ✅ Use PyQt6 default colors
            id_item = QTableWidgetItem(str(prod_id))
            self.stock_table.setItem(r, 0, id_item)
            
            name_item = QTableWidgetItem(str(name) if name else "")
            self.stock_table.setItem(r, 1, name_item)
            
            self.stock_table.setItem(r, 2, QTableWidgetItem(str(sku) if sku else ""))
            self.stock_table.setItem(r, 3, QTableWidgetItem(str(barcode) if barcode else ""))
            self.stock_table.setItem(r, 4, QTableWidgetItem(str(category) if category else ""))
            
            stock_item = QTableWidgetItem(str(stock))
            if stock == 0:
                stock_item.setForeground(QColor("#dc3545"))  # Red
            elif stock <= low_stock:
                stock_item.setForeground(QColor("#f39c12"))  # Orange
            self.stock_table.setItem(r, 5, stock_item)
            
            self.stock_table.setItem(r, 6, QTableWidgetItem(format_money(cost)))
            self.stock_table.setItem(r, 7, QTableWidgetItem(format_money(price)))
            self.stock_table.setItem(r, 8, QTableWidgetItem(format_money(stock_value)))
            self.stock_table.setItem(r, 9, QTableWidgetItem(str(low_stock)))
            
            status_item = QTableWidgetItem(str(status))
            if status == "Out of Stock" or status == "ကုန်သွားပြီ":
                status_item.setForeground(QColor("#dc3545"))
            elif status == "Low Stock" or status == "စတော့နည်းနေပြီ":
                status_item.setForeground(QColor("#f39c12"))
            else:
                status_item.setForeground(QColor("#28a745"))
            self.stock_table.setItem(r, 10, status_item)
            
            self.stock_table.setItem(r, 11, QTableWidgetItem(str(last_upd) if last_upd else ""))
            
            locations_str = str(locations) if locations else ""
            if locations_str:
                loc_list = [loc.strip() for loc in locations_str.split(',') if loc.strip()]
                seen = set()
                unique_locs = []
                for loc in loc_list:
                    if loc not in seen:
                        seen.add(loc)
                        unique_locs.append(loc)
                locations_str = ', '.join(unique_locs)
            
            self.stock_table.setItem(r, 12, QTableWidgetItem(locations_str))
            
            # ✅ History button (keep styled for functionality)
            btn_history = QPushButton()
            btn_history.setIcon(history_icon)
            btn_history.setText(" " + ("ကြည့်ရန်" if lang == "my" else "View"))
            btn_history.setFixedSize(95, 32)
            btn_history.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_history.setStyleSheet("""
                QPushButton {
                    background-color: #5865f2;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-size: 9pt;
                    font-weight: 500;
                    padding: 2px 8px;
                }
                QPushButton:hover {
                    background-color: #4752c4;
                }
                QPushButton:pressed {
                    background-color: #3c45a3;
                }
            """)
            btn_history.clicked.connect(lambda checked, pid=prod_id, pname=name: self.show_transaction_history(pid, pname))
            self.stock_table.setCellWidget(r, 13, self._centered_cell_widget(btn_history))

    def show_transaction_history(self, product_id, product_name):
        dialog = ProductTransactionHistoryDialog(product_id, product_name, self)
        dialog.exec()

    def get_all_stock_data(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                p.id, 
                p.name, 
                p.sku, 
                p.barcode, 
                p.category,
                COALESCE(p.stock, 0) as total_stock,
                p.cost, 
                p.price,
                (COALESCE(p.cost, 0) * COALESCE(p.stock, 0)) as stock_value,
                p.low_stock,
                (SELECT GROUP_CONCAT(pl.location || ' (' || pl.quantity || ')', ', ') 
                 FROM product_locations pl
                 WHERE pl.product_id = p.id AND pl.quantity > 0) as locations
            FROM products p
            WHERE (p.sold_by IS NULL OR p.sold_by != 'Service')
            ORDER BY p.name
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_to_excel(self):
        from utils.excel_exporter import ExcelExporter
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        lang = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"current_stock_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Current Stock" if lang != "my" else "လက်ရှိစတော့စာရင်း ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            rows = self.get_all_stock_data()
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Current Stock"
            
            ws.merge_cells('A1:J1')
            ws['A1'] = "CURRENT STOCK REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            
            if lang == "my":
                headers = ["ပစ္စည်းအမည်", "SKU", "ဘားကုဒ်", "အမျိုးအစား", 
                          "လက်ကျန်", "ကုန်ကျစရိတ်", "ရောင်းဈေး", "စတော့တန်ဖိုး", 
                          "သတိပေးပမာဏ", "နေရာ"]
            else:
                headers = ["Product Name", "SKU", "Barcode", "Category", 
                          "Current Stock", "Cost Price", "Selling Price", "Stock Value", 
                          "Low Stock Alert", "Location"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            total_stock = 0
            total_stock_value = 0
            total_cost_value = 0
            
            for row_idx, row_data in enumerate(rows, start=6):
                name, sku, barcode, category, stock, cost, price, stock_value, low_stock, location = row_data[1:]
                
                ws.cell(row=row_idx, column=1, value=name)
                ws.cell(row=row_idx, column=2, value=sku or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=category or "")
                ws.cell(row=row_idx, column=5, value=stock)
                ws.cell(row=row_idx, column=6, value=format_money(cost, symbol))
                ws.cell(row=row_idx, column=7, value=format_money(price, symbol))
                ws.cell(row=row_idx, column=8, value=format_money(stock_value, symbol))
                ws.cell(row=row_idx, column=9, value=low_stock)
                ws.cell(row=row_idx, column=10, value=location or "")
                
                total_stock += stock
                total_stock_value += stock_value
                total_cost_value += cost * stock
            
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=4, value="TOTAL").font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=total_stock)
            ws.cell(row=summary_row, column=7, value=format_money(total_stock_value, symbol))
            ws.cell(row=summary_row + 1, column=4, value="TOTAL COST VALUE").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=5, value=format_money(total_cost_value, symbol))
            ws.cell(row=summary_row + 2, column=4, value="POTENTIAL PROFIT").font = Font(bold=True)
            ws.cell(row=summary_row + 2, column=5, value=format_money(total_stock_value - total_cost_value, symbol))
            
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 20
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def get_lang(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key='language'")
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else "en"
        except:
            return "en"

    def retranslateUi(self):
        lang = self.get_lang()
        
        # Update SearchWidget placeholder
        if lang == "my":
            self.search_widget.set_placeholder_text("ပစ္စည်းအမည် / SKU / ဘားကုဒ်ဖြင့် ရှာရန်...")
        else:
            self.search_widget.set_placeholder_text("Search by name, SKU or barcode...")
        
        if lang == "my":
            self.btn_stock_in.setText(" " + tr("stock_in"))
            self.btn_stock_out.setText(" " + tr("stock_out"))
            self.btn_adjustment.setText(" " + tr("adjustment"))
            self.btn_transfer.setText(" လွှဲပြောင်းမည်")
            self.btn_export.setText(" စတော့စာရင်းထုတ်မည်")
            self.btn_view_movements.setText(" လှုပ်ရှားမှုများကြည့်ရန်")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "စတော့ရှိပါ")
            self.status_filter.setItemText(2, "စတော့နည်းနေပြီ")
            self.status_filter.setItemText(3, "ကုန်သွားပြီ")
        else:
            self.btn_stock_in.setText(" " + tr("stock_in"))
            self.btn_stock_out.setText(" " + tr("stock_out"))
            self.btn_adjustment.setText(" " + tr("adjustment"))
            self.btn_transfer.setText(" Transfer")
            self.btn_export.setText(" Export Current Stock")
            self.btn_view_movements.setText(" View Movements")
            self.status_filter.setItemText(0, "All Status")
            self.status_filter.setItemText(1, "In Stock")
            self.status_filter.setItemText(2, "Low Stock")
            self.status_filter.setItemText(3, "Out of Stock")
        
        # ✅ Update button icons
        self._update_button_icons()
        
        self.load_data()
    
    def showEvent(self, event):
        """Handle show event - refresh data"""
        self.load_data()
        super().showEvent(event)
