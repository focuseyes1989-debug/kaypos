# ui/reports/financial_summary.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget, QGroupBox
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from utils.translations import tr
from utils.language import lang
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from loguru import logger
from datetime import datetime


class FinancialSummaryWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Sales data
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0), COUNT(*), COALESCE(AVG(total), 0)
                FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales, transaction_count, avg_sale = cursor.fetchone()
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0), COUNT(*)
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses, expense_count = cursor.fetchone()
            
            # Sales by category
            cursor.execute("""
                SELECT COALESCE(p.category, 'Uncategorized') as category,
                       COALESCE(SUM(si.total), 0) as sales,
                       COALESCE(SUM(p.cost * si.qty), 0) as cogs
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                LEFT JOIN products p ON si.product_name = p.name
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY p.category
                ORDER BY sales DESC
            """, (self.from_date, self.to_date))
            sales_categories = cursor.fetchall()
            
            # Expenses by category
            cursor.execute("""
                SELECT category, COALESCE(SUM(amount), 0) as total
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY category
                ORDER BY total DESC
            """, (self.from_date, self.to_date))
            expense_categories = cursor.fetchall()
            
            conn.close()
            
            gross_profit = (total_sales or 0) - (total_cogs or 0)
            net_profit = gross_profit - (total_expenses or 0)
            net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            self.result.emit({
                'total_sales': total_sales or 0,
                'total_cogs': total_cogs or 0,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses or 0,
                'net_profit': net_profit,
                'net_margin': net_margin,
                'transaction_count': transaction_count,
                'expense_count': expense_count,
                'sales_categories': sales_categories,
                'expense_categories': expense_categories
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class FinancialSummaryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.setup_ui()
        self.retranslateUi()
        lang.language_changed.connect(self.retranslateUi)
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        is_dark = is_dark_theme()
        
        # Table styles
        for table in [self.sales_category_table, self.expense_category_table]:
            if is_dark:
                table_style = """
                    QTableWidget {
                        background-color: #2f3136;
                        alternate-background-color: #36393f;
                        selection-background-color: #40444b;
                        selection-color: #dcddde;
                        gridline-color: #40444b;
                        border: 1px solid #40444b;
                        border-radius: 6px;
                        color: #dcddde;
                    }
                    QTableWidget::item {
                        padding: 8px 12px;
                        color: #dcddde;
                    }
                    QTableWidget::item:selected {
                        background-color: #40444b;
                        color: #dcddde;
                    }
                    QHeaderView::section {
                        background-color: #202225;
                        padding: 8px 12px;
                        border: none;
                        border-bottom: 2px solid #40444b;
                        font-weight: 600;
                        font-size: 10pt;
                        color: #b9bbbe;
                    }
                    QTableWidget::item:hover {
                        background-color: #40444b;
                    }
                """
            else:
                table_style = """
                    QTableWidget {
                        background-color: white;
                        alternate-background-color: #f8f9fa;
                        selection-background-color: #e9ecef;
                        selection-color: #212529;
                        gridline-color: #dee2e6;
                        border: 1px solid #dee2e6;
                        border-radius: 6px;
                        color: #212529;
                    }
                    QTableWidget::item {
                        padding: 8px 12px;
                        color: #212529;
                    }
                    QTableWidget::item:selected {
                        background-color: #e9ecef;
                        color: #212529;
                    }
                    QHeaderView::section {
                        background-color: #f8f9fa;
                        padding: 8px 12px;
                        border: none;
                        border-bottom: 2px solid #dee2e6;
                        font-weight: 600;
                        font-size: 10pt;
                        color: #2c3e50;
                    }
                    QTableWidget::item:hover {
                        background-color: #f1f3f5;
                    }
                """
            table.setStyleSheet(table_style)
        
        # Groupbox styles
        for group in [self.sales_cat_group, self.exp_cat_group]:
            group.setStyleSheet(f"""
                QGroupBox {{
                    font-weight: 600;
                    font-size: 10pt;
                    border: 1px solid {colors['border']};
                    border-radius: 8px;
                    padding-top: 10px;
                    margin-top: 5px;
                    color: {colors['text']};
                    background-color: {colors['card_bg']};
                }}
                QGroupBox::title {{
                    subcontrol-origin: margin;
                    left: 10px;
                    padding: 0 5px 0 5px;
                    color: {colors['text']};
                }}
            """)
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Row 1 - Using create_summary_card with SVG icons
        card_layout1 = QHBoxLayout()
        card_layout1.setSpacing(10)
        
        # Total Sales Card
        self.sales_card = self.parent_dialog.create_summary_card(
            title=tr("total_sales"),
            value="0",
            icon="attach_money",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.sales_card, 1)
        
        # COGS Card
        self.cogs_card = self.parent_dialog.create_summary_card(
            title=tr("cogs"),
            value="0",
            icon="package",
            color="#e74c3c",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.cogs_card, 1)
        
        # Gross Profit Card
        self.gross_card = self.parent_dialog.create_summary_card(
            title=tr("gross_profit"),
            value="0",
            icon="trending_up",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.gross_card, 1)
        
        layout.addLayout(card_layout1)
        
        # Row 2
        card_layout2 = QHBoxLayout()
        card_layout2.setSpacing(10)
        
        # Expenses Card
        self.expense_card = self.parent_dialog.create_summary_card(
            title=tr("expenses"),
            value="0",
            icon="money_off",
            color="#e67e22",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.expense_card, 1)
        
        # Net Profit Card
        self.net_card = self.parent_dialog.create_summary_card(
            title=tr("net_profit"),
            value="0",
            icon="bar_chart",
            color="#9b59b6",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.net_card, 1)
        
        # Net Margin Card
        self.margin_card = self.parent_dialog.create_summary_card(
            title=tr("net_margin"),
            value="0%",
            icon="analytics",
            color="#1abc9c",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.margin_card, 1)
        
        layout.addLayout(card_layout2)
        
        # Category breakdown tables
        split_layout = QHBoxLayout()
        split_layout.setSpacing(10)
        
        # Sales by category
        self.sales_cat_group = QGroupBox(tr("sales_by_category"))
        sales_cat_layout = QVBoxLayout()
        self.sales_category_table = QTableWidget()
        self.sales_category_table.setColumnCount(3)
        self.sales_category_table.setHorizontalHeaderLabels([tr("category"), tr("sales_header"), tr("cogs")])
        self.sales_category_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.sales_category_table.setAlternatingRowColors(True)
        sales_cat_layout.addWidget(self.sales_category_table)
        self.sales_cat_group.setLayout(sales_cat_layout)
        split_layout.addWidget(self.sales_cat_group, 1)
        
        # Expenses by category
        self.exp_cat_group = QGroupBox(tr("expenses_by_category"))
        exp_cat_layout = QVBoxLayout()
        self.expense_category_table = QTableWidget()
        self.expense_category_table.setColumnCount(2)
        self.expense_category_table.setHorizontalHeaderLabels([tr("category"), tr("amount")])
        self.expense_category_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.expense_category_table.setAlternatingRowColors(True)
        exp_cat_layout.addWidget(self.expense_category_table)
        self.exp_cat_group.setLayout(exp_cat_layout)
        split_layout.addWidget(self.exp_cat_group, 1)
        
        layout.addLayout(split_layout)
        
        # Apply initial theme
        self._apply_theme()
        
        self.setLayout(layout)
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        loading_text = tr("loading")
        self.sales_card.set_value(loading_text)
        self.cogs_card.set_value(loading_text)
        self.gross_card.set_value(loading_text)
        self.expense_card.set_value(loading_text)
        self.net_card.set_value(loading_text)
        self.margin_card.set_value(loading_text)
        self.sales_category_table.setRowCount(0)
        self.expense_category_table.setRowCount(0)
        
        worker = FinancialSummaryWorker(from_date, to_date)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        
        # Update cards using update_summary_card
        self.parent_dialog.update_summary_card(self.sales_card, result['total_sales'], symbol)
        self.parent_dialog.update_summary_card(self.cogs_card, result['total_cogs'], symbol)
        self.parent_dialog.update_summary_card(self.gross_card, result['gross_profit'], symbol)
        self.parent_dialog.update_summary_card(self.expense_card, result['total_expenses'], symbol)
        self.parent_dialog.update_summary_card(self.net_card, result['net_profit'], symbol)
        self.margin_card.set_value(f"{result['net_margin']:.1f}%")
        
        # Color coding
        if result['net_profit'] >= 0:
            self.set_card_color(self.net_card, "#2ecc71")
            self.set_card_color(self.margin_card, "#2ecc71")
        else:
            self.set_card_color(self.net_card, "#e74c3c")
            self.set_card_color(self.margin_card, "#e74c3c")
        
        # Sales by category with color coding
        is_dark = is_dark_theme()
        text_color = "#dcddde" if is_dark else "#212529"
        green_color = "#3ba55d" if is_dark else "#28a745"
        red_color = "#ed4245" if is_dark else "#dc3545"
        
        self.sales_category_table.setRowCount(len(result['sales_categories']))
        for i, (cat, sales, cogs) in enumerate(result['sales_categories']):
            cat_item = QTableWidgetItem(cat or "Uncategorized")
            cat_item.setForeground(QColor(text_color))
            self.sales_category_table.setItem(i, 0, cat_item)
            
            sales_item = QTableWidgetItem(format_money(sales, symbol))
            sales_item.setForeground(QColor(green_color))
            self.sales_category_table.setItem(i, 1, sales_item)
            
            cogs_item = QTableWidgetItem(format_money(cogs, symbol))
            cogs_item.setForeground(QColor(red_color))
            self.sales_category_table.setItem(i, 2, cogs_item)
        
        # Expenses by category
        self.expense_category_table.setRowCount(len(result['expense_categories']))
        for i, (cat, amount) in enumerate(result['expense_categories']):
            cat_item = QTableWidgetItem(cat)
            cat_item.setForeground(QColor(text_color))
            self.expense_category_table.setItem(i, 0, cat_item)
            
            amount_item = QTableWidgetItem(format_money(amount, symbol))
            amount_item.setForeground(QColor(red_color))
            self.expense_category_table.setItem(i, 1, amount_item)
        
        self.parent_dialog.on_refresh_complete()

    def set_card_color(self, card, color):
        """Update card color"""
        card.set_color(color)
    
    def retranslateUi(self):
        """Update UI text based on language"""
        if hasattr(self, 'sales_card'):
            self.sales_card.set_title(tr("total_sales"))
            self.cogs_card.set_title(tr("cogs"))
            self.gross_card.set_title(tr("gross_profit"))
            self.expense_card.set_title(tr("expenses"))
            self.net_card.set_title(tr("net_profit"))
            self.margin_card.set_title(tr("net_margin"))
        
        if hasattr(self, 'sales_cat_group'):
            self.sales_cat_group.setTitle(tr("sales_by_category"))
            self.sales_category_table.setHorizontalHeaderLabels([tr("category"), tr("sales_header"), tr("cogs")])
            self.exp_cat_group.setTitle(tr("expenses_by_category"))
            self.expense_category_table.setHorizontalHeaderLabels([tr("category"), tr("amount")])
        
        # Update card icons after language change
        self.sales_card.set_icon("attach_money", is_svg=True, size=(24, 24))
        self.cogs_card.set_icon("package", is_svg=True, size=(24, 24))
        self.gross_card.set_icon("trending_up", is_svg=True, size=(24, 24))
        self.expense_card.set_icon("money_off", is_svg=True, size=(24, 24))
        self.net_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        self.margin_card.set_icon("analytics", is_svg=True, size=(24, 24))
        
        if self._current_data:
            symbol = get_currency_symbol()
            self.parent_dialog.update_summary_card(self.sales_card, self._current_data['total_sales'], symbol)
            self.parent_dialog.update_summary_card(self.cogs_card, self._current_data['total_cogs'], symbol)
            self.parent_dialog.update_summary_card(self.gross_card, self._current_data['gross_profit'], symbol)
            self.parent_dialog.update_summary_card(self.expense_card, self._current_data['total_expenses'], symbol)
            self.parent_dialog.update_summary_card(self.net_card, self._current_data['net_profit'], symbol)
            self.margin_card.set_value(f"{self._current_data['net_margin']:.1f}%")
            
            profit_color = "#2ecc71" if self._current_data['net_profit'] >= 0 else "#e74c3c"
            self.set_card_color(self.net_card, profit_color)
            self.set_card_color(self.margin_card, profit_color)
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"financial_summary_{from_date}_to_{to_date}.xlsx",
            "Export Financial Summary"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            # Summary data
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Sales by category
            cursor.execute("""
                SELECT COALESCE(p.category, 'Uncategorized') as category,
                       COALESCE(SUM(si.total), 0) as sales,
                       COALESCE(SUM(p.cost * si.qty), 0) as cogs
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                LEFT JOIN products p ON si.product_name = p.name
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY p.category
                ORDER BY sales DESC
            """, (from_date, to_date))
            sales_categories = cursor.fetchall()
            
            # Expenses by category
            cursor.execute("""
                SELECT category, COALESCE(SUM(amount), 0) as total
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                GROUP BY category
                ORDER BY total DESC
            """, (from_date, to_date))
            expense_categories = cursor.fetchall()
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Summary"
            
            # Title
            ws.merge_cells('A1:C1')
            ws['A1'] = "FINANCIAL SUMMARY"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            
            summary_data = [
                ("Total Sales", format_money(total_sales, symbol)),
                ("COGS", format_money(total_cogs, symbol)),
                ("Gross Profit", format_money(gross_profit, symbol)),
                ("Total Expenses", format_money(total_expenses, symbol)),
                ("Net Profit", format_money(net_profit, symbol)),
                ("Net Margin", f"{margin:.1f}%"),
            ]
            
            for i, (label, value) in enumerate(summary_data, start=6):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # Sales by category
            row = 14
            ws.cell(row=row, column=1, value="Sales by Category")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Category")
            ws.cell(row=row, column=2, value="Sales")
            ws.cell(row=row, column=3, value="COGS")
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            for cat, sales, cogs in sales_categories:
                ws.cell(row=row, column=1, value=cat or "Uncategorized")
                ws.cell(row=row, column=2, value=float(sales))
                ws.cell(row=row, column=3, value=float(cogs))
                row += 1
            
            # Expenses by category
            row += 2
            ws.cell(row=row, column=1, value="Expenses by Category")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            ws.cell(row=row, column=1, value="Category")
            ws.cell(row=row, column=2, value="Amount")
            for col in range(1, 3):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            for cat, amount in expense_categories:
                ws.cell(row=row, column=1, value=cat)
                ws.cell(row=row, column=2, value=float(amount))
                row += 1
            
            # Auto adjust columns
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)