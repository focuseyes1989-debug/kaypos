# ui/reports/profit_loss_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget, QSizePolicy
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from loguru import logger
from datetime import datetime


class PLReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total sales
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales = cursor.fetchone()[0]
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(
                    COALESCE(
                        NULLIF(si.cost, 0),
                        (SELECT p.cost FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                        0
                    ) * si.qty
                ), 0)
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                WHERE s.status = 'completed'
                  AND date(s.created_at) BETWEEN ? AND ?
                  AND COALESCE(
                        (SELECT p.sold_by FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                        ''
                      ) != 'Service'
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses = cursor.fetchone()[0]
            
            # Monthly breakdown
            cursor.execute("""
                WITH months AS (
                    SELECT strftime('%Y-%m', created_at) AS month
                    FROM sales
                    WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
                    UNION
                    SELECT strftime('%Y-%m', expense_date) AS month
                    FROM expenses
                    WHERE expense_date BETWEEN ? AND ?
                ),
                sales_by_month AS (
                    SELECT strftime('%Y-%m', created_at) AS month,
                           COALESCE(SUM(total), 0) AS sales
                    FROM sales
                    WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', created_at)
                ),
                cogs_by_month AS (
                    SELECT strftime('%Y-%m', s.created_at) AS month,
                           COALESCE(SUM(
                               COALESCE(
                                   NULLIF(si.cost, 0),
                                   (SELECT p.cost FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                                   0
                               ) * si.qty
                           ), 0) AS cogs
                    FROM sale_items si
                    JOIN sales s ON si.sale_id = s.id
                    WHERE s.status = 'completed'
                      AND date(s.created_at) BETWEEN ? AND ?
                      AND COALESCE(
                            (SELECT p.sold_by FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                            ''
                          ) != 'Service'
                    GROUP BY strftime('%Y-%m', s.created_at)
                ),
                expense_by_month AS (
                    SELECT strftime('%Y-%m', expense_date) AS month,
                           COALESCE(SUM(amount), 0) AS expenses
                    FROM expenses
                    WHERE expense_date BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', expense_date)
                )
                SELECT m.month,
                       COALESCE(s.sales, 0) AS sales,
                       COALESCE(c.cogs, 0) AS cogs,
                       COALESCE(e.expenses, 0) AS expenses
                FROM months m
                LEFT JOIN sales_by_month s ON s.month = m.month
                LEFT JOIN cogs_by_month c ON c.month = m.month
                LEFT JOIN expense_by_month e ON e.month = m.month
                ORDER BY m.month
            """, (
                self.from_date, self.to_date,
                self.from_date, self.to_date,
                self.from_date, self.to_date,
                self.from_date, self.to_date,
                self.from_date, self.to_date,
            ))
            monthly_rows = cursor.fetchall()
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            self.result.emit({
                'total_sales': total_sales,
                'total_cogs': total_cogs,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'net_margin': net_margin,
                'monthly_rows': monthly_rows
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ProfitLossReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.setup_ui()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        is_dark = is_dark_theme()
        
        if is_dark:
            table_style = """
                QTableWidget {
                    background-color: #2f3136;
                    alternate-background-color: #36393f;
                    selection-background-color: #40444b;
                    selection-color: #dcddde;
                    gridline-color: #40444b;
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    color: #dcddde;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #dcddde;
                }
                QTableWidget::item:selected {
                    background-color: #40444b;
                    color: #dcddde;
                }
                QHeaderView::section {
                    background-color: #202225;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #40444b;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #b9bbbe;
                }
                QTableWidget::item:hover {
                    background-color: #40444b;
                }
            """
        else:
            table_style = """
                QTableWidget {
                    background-color: white;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e9ecef;
                    selection-color: #212529;
                    gridline-color: #dee2e6;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    color: #212529;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #212529;
                }
                QTableWidget::item:selected {
                    background-color: #e9ecef;
                    color: #212529;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #dee2e6;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #2c3e50;
                }
                QTableWidget::item:hover {
                    background-color: #f1f3f5;
                }
            """
        
        self.table.setStyleSheet(table_style)
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Row 1 - Using SummaryCardWidget with SVG icons
        card_layout1 = QHBoxLayout()
        card_layout1.setSpacing(10)
        
        # Total Sales Card
        self.sales_card = self.parent_dialog.create_summary_card(
            title="Total Sales",
            value="0",
            icon="attach_money",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.sales_card, 1)
        
        # COGS Card
        self.cogs_card = self.parent_dialog.create_summary_card(
            title="COGS",
            value="0",
            icon="package",
            color="#e74c3c",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.cogs_card, 1)
        
        # Gross Profit Card
        self.gross_card = self.parent_dialog.create_summary_card(
            title="Gross Profit",
            value="0",
            icon="trending_up",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout1.addWidget(self.gross_card, 1)
        
        layout.addLayout(card_layout1)
        
        # Row 2
        card_layout2 = QHBoxLayout()
        card_layout2.setSpacing(10)
        
        # Operating Expenses Card
        self.expense_card = self.parent_dialog.create_summary_card(
            title="Operating Expenses",
            value="0",
            icon="money_off",
            color="#e67e22",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.expense_card, 1)
        
        # Net Profit Card
        self.net_card = self.parent_dialog.create_summary_card(
            title="Net Profit",
            value="0",
            icon="bar_chart",
            color="#9b59b6",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.net_card, 1)
        
        # Net Margin Card
        self.margin_card = self.parent_dialog.create_summary_card(
            title="Net Margin",
            value="0%",
            icon="analytics",
            color="#1abc9c",
            icon_is_svg=True
        )
        card_layout2.addWidget(self.margin_card, 1)
        
        layout.addLayout(card_layout2)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Month", "Sales", "COGS", "Gross Profit", "Expenses", "Net Profit", "Margin %"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Apply initial theme
        self._apply_theme()
        
        header = self.table.horizontalHeader()
        for i in range(7):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        header.setStretchLastSection(True)
        layout.addWidget(self.table, 1)
        
        self.setLayout(layout)
    
    def refresh(self, from_date, to_date):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.table.setRowCount(0)
        
        # Use set_value() instead of amount_label.setText()
        self.sales_card.set_value("Loading...")
        self.cogs_card.set_value("Loading...")
        self.gross_card.set_value("Loading...")
        self.expense_card.set_value("Loading...")
        self.net_card.set_value("Loading...")
        self.margin_card.set_value("Loading...")
        
        worker = PLReportWorker(from_date, to_date)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        is_dark = is_dark_theme()
        
        self._current_data = result
        self._is_loading = False
        
        # Color definitions
        green_color = "#3ba55d" if is_dark else "#28a745"
        red_color = "#ed4245" if is_dark else "#dc3545"
        orange_color = "#faa81a" if is_dark else "#f39c12"
        text_color = "#dcddde" if is_dark else "#212529"
        
        # Update cards using update_summary_card
        self.parent_dialog.update_summary_card(self.sales_card, result['total_sales'], symbol)
        self.parent_dialog.update_summary_card(self.cogs_card, result['total_cogs'], symbol)
        self.parent_dialog.update_summary_card(self.gross_card, result['gross_profit'], symbol)
        self.parent_dialog.update_summary_card(self.expense_card, result['total_expenses'], symbol)
        self.parent_dialog.update_summary_card(self.net_card, result['net_profit'], symbol)
        self.margin_card.set_value(f"{result['net_margin']:.1f}%")
        
        # Color coding
        if result['net_profit'] >= 0:
            self.net_card.set_color(green_color)
            self.margin_card.set_color(green_color)
        else:
            self.net_card.set_color(red_color)
            self.margin_card.set_color(red_color)
        
        # Monthly breakdown
        rows = result['monthly_rows']
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            month, sales, cogs, expenses = row
            gross = sales - cogs
            net = gross - expenses
            margin = (net / sales * 100) if sales > 0 else 0
            
            # Month
            month_item = QTableWidgetItem(month)
            month_item.setForeground(QColor(text_color))
            self.table.setItem(i, 0, month_item)
            
            # Sales
            sales_item = QTableWidgetItem(format_money(sales, symbol))
            sales_item.setForeground(QColor(green_color))
            self.table.setItem(i, 1, sales_item)
            
            # COGS
            cogs_item = QTableWidgetItem(format_money(cogs, symbol))
            cogs_item.setForeground(QColor(red_color))
            self.table.setItem(i, 2, cogs_item)
            
            # Gross Profit
            gross_item = QTableWidgetItem(format_money(gross, symbol))
            gross_item.setForeground(QColor(green_color) if gross >= 0 else QColor(red_color))
            self.table.setItem(i, 3, gross_item)
            
            # Expenses
            expenses_item = QTableWidgetItem(format_money(expenses, symbol))
            expenses_item.setForeground(QColor(red_color))
            self.table.setItem(i, 4, expenses_item)
            
            # Net Profit
            net_item = QTableWidgetItem(format_money(net, symbol))
            net_item.setForeground(QColor(green_color) if net >= 0 else QColor(red_color))
            self.table.setItem(i, 5, net_item)
            
            # Margin %
            margin_item = QTableWidgetItem(f"{margin:.1f}%")
            if margin >= 10:
                margin_item.setForeground(QColor(green_color))
            elif margin >= 0:
                margin_item.setForeground(QColor(orange_color))
            else:
                margin_item.setForeground(QColor(red_color))
            self.table.setItem(i, 6, margin_item)
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"profit_loss_report_{from_date}_to_{to_date}.xlsx",
            "Export Profit & Loss Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                WITH months AS (
                    SELECT strftime('%Y-%m', created_at) AS month
                    FROM sales
                    WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
                    UNION
                    SELECT strftime('%Y-%m', expense_date) AS month
                    FROM expenses
                    WHERE expense_date BETWEEN ? AND ?
                ),
                sales_by_month AS (
                    SELECT strftime('%Y-%m', created_at) AS month,
                           COALESCE(SUM(total), 0) AS sales
                    FROM sales
                    WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', created_at)
                ),
                cogs_by_month AS (
                    SELECT strftime('%Y-%m', s.created_at) AS month,
                           COALESCE(SUM(
                               COALESCE(
                                   NULLIF(si.cost, 0),
                                   (SELECT p.cost FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                                   0
                               ) * si.qty
                           ), 0) AS cogs
                    FROM sale_items si
                    JOIN sales s ON si.sale_id = s.id
                    WHERE s.status = 'completed'
                      AND date(s.created_at) BETWEEN ? AND ?
                      AND COALESCE(
                            (SELECT p.sold_by FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                            ''
                          ) != 'Service'
                    GROUP BY strftime('%Y-%m', s.created_at)
                ),
                expense_by_month AS (
                    SELECT strftime('%Y-%m', expense_date) AS month,
                           COALESCE(SUM(amount), 0) AS expenses
                    FROM expenses
                    WHERE expense_date BETWEEN ? AND ?
                    GROUP BY strftime('%Y-%m', expense_date)
                )
                SELECT m.month,
                       COALESCE(s.sales, 0) AS sales,
                       COALESCE(c.cogs, 0) AS cogs,
                       COALESCE(e.expenses, 0) AS expenses
                FROM months m
                LEFT JOIN sales_by_month s ON s.month = m.month
                LEFT JOIN cogs_by_month c ON c.month = m.month
                LEFT JOIN expense_by_month e ON e.month = m.month
                ORDER BY m.month
            """, (
                from_date, to_date,
                from_date, to_date,
                from_date, to_date,
                from_date, to_date,
                from_date, to_date,
            ))
            rows = cursor.fetchall()
            
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(
                    COALESCE(
                        NULLIF(si.cost, 0),
                        (SELECT p.cost FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                        0
                    ) * si.qty
                ), 0)
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                WHERE s.status = 'completed'
                  AND date(s.created_at) BETWEEN ? AND ?
                  AND COALESCE(
                        (SELECT p.sold_by FROM products p WHERE p.name = si.product_name ORDER BY p.id DESC LIMIT 1),
                        ''
                      ) != 'Service'
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"
            
            # Title
            ws.merge_cells('A1:G1')
            ws['A1'] = "PROFIT & LOSS REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Sales: {format_money(total_sales, symbol)}"
            ws['A7'] = f"COGS: {format_money(total_cogs, symbol)}"
            ws['A8'] = f"Gross Profit: {format_money(gross_profit, symbol)}"
            ws['A9'] = f"Total Expenses: {format_money(total_expenses, symbol)}"
            ws['A10'] = f"Net Profit: {format_money(net_profit, symbol)}"
            ws['A11'] = f"Net Margin: {margin:.1f}%"
            
            # Headers
            headers = ["Month", "Sales", "COGS", "Gross Profit", "Expenses", "Net Profit", "Margin %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=13, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row in enumerate(rows, start=14):
                month, sales, cogs, expenses = row
                gross = sales - cogs
                net = gross - expenses
                margin_pct = (net / sales * 100) if sales > 0 else 0
                
                ws.cell(row=row_idx, column=1, value=month)
                ws.cell(row=row_idx, column=2, value=float(sales))
                ws.cell(row=row_idx, column=3, value=float(cogs))
                ws.cell(row=row_idx, column=4, value=float(gross))
                ws.cell(row=row_idx, column=5, value=float(expenses))
                ws.cell(row=row_idx, column=6, value=float(net))
                ws.cell(row=row_idx, column=7, value=float(margin_pct))
            
            # Auto adjust columns
            for col in range(1, 8):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def retranslateUi(self):
        """Retranslate UI for language change"""
        lang = self.parent_dialog.get_lang() if hasattr(self.parent_dialog, 'get_lang') else "en"
        
        if lang == "my":
            self.sales_card.set_title("စုစုပေါင်းရောင်းအား")
            self.cogs_card.set_title("ကုန်ပစ္စည်းကုန်ကျစရိတ်")
            self.gross_card.set_title("အသားတင်အမြတ်")
            self.expense_card.set_title("လည်ပတ်စရိတ်")
            self.net_card.set_title("အသားတင်အမြတ်")
            self.margin_card.set_title("အသားတင်အမြတ်ရာခိုင်နှုန်း")
            self.table.setHorizontalHeaderLabels([
                "လ", "ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်", 
                "အသုံးစရိတ်", "အသားတင်အမြတ်", "ရာခိုင်နှုန်း"
            ])
        else:
            self.sales_card.set_title("Total Sales")
            self.cogs_card.set_title("COGS")
            self.gross_card.set_title("Gross Profit")
            self.expense_card.set_title("Operating Expenses")
            self.net_card.set_title("Net Profit")
            self.margin_card.set_title("Net Margin")
            self.table.setHorizontalHeaderLabels([
                "Month", "Sales", "COGS", "Gross Profit", "Expenses", "Net Profit", "Margin %"
            ])
        
        # Update card icons after language change
        self.sales_card.set_icon("attach_money", is_svg=True, size=(24, 24))
        self.cogs_card.set_icon("package", is_svg=True, size=(24, 24))
        self.gross_card.set_icon("trending_up", is_svg=True, size=(24, 24))
        self.expense_card.set_icon("money_off", is_svg=True, size=(24, 24))
        self.net_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        self.margin_card.set_icon("analytics", is_svg=True, size=(24, 24))
        
        self._apply_theme()