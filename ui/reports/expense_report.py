# ui/reports/expense_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox, QWidget, QComboBox, QLabel
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QIcon, QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from ui.widgets import PaginationWidget
from ui.widgets.modern_button import ModernButton
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from loguru import logger
from datetime import datetime
import os


class ExpenseReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date, category=None, page=1, page_size=25):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
        self.category = category
        self.page = page
        self.page_size = page_size
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Build query for counting
            count_query = """
                SELECT COUNT(*) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """
            count_params = [self.from_date, self.to_date]
            
            if self.category and self.category != "All Categories":
                count_query += " AND category = ?"
                count_params.append(self.category)
            
            cursor.execute(count_query, count_params)
            total_count = cursor.fetchone()[0]
            
            # Calculate offset
            offset = (self.page - 1) * self.page_size
            
            # Build main query
            query = """
                SELECT expense_date, expense_no, category, description, amount, payment_method
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """
            params = [self.from_date, self.to_date]
            
            if self.category and self.category != "All Categories":
                query += " AND category = ?"
                params.append(self.category)
            
            query += " ORDER BY expense_date DESC LIMIT ? OFFSET ?"
            params.extend([self.page_size, offset])
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            
            # Get totals for current page
            total_expenses = sum(row[4] for row in rows) if rows else 0
            expense_count = len(rows)
            avg_expense = total_expenses / expense_count if expense_count > 0 else 0
            
            conn.close()
            
            self.result.emit({
                'rows': rows,
                'total_expenses': total_expenses,
                'expense_count': expense_count,
                'avg_expense': avg_expense,
                'total_count': total_count,
                'current_page': self.page,
                'page_size': self.page_size,
                'total_pages': (total_count + self.page_size - 1) // self.page_size if total_count > 0 else 1
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ExpenseReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.current_page = 1
        self.page_size = 25
        self.total_items = 0
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.setup_ui()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        
        # Table style
        is_dark = is_dark_theme()
        if is_dark:
            table_style = """
                QTableWidget {
                    background-color: #2f3136;
                    alternate-background-color: #36393f;
                    selection-background-color: #40444b;
                    selection-color: #dcddde;
                    gridline-color: #40444b;
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    color: #dcddde;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #dcddde;
                }
                QTableWidget::item:selected {
                    background-color: #40444b;
                    color: #dcddde;
                }
                QHeaderView::section {
                    background-color: #202225;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #40444b;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #b9bbbe;
                }
                QTableWidget::item:hover {
                    background-color: #40444b;
                }
            """
        else:
            table_style = """
                QTableWidget {
                    background-color: white;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e9ecef;
                    selection-color: #212529;
                    gridline-color: #dee2e6;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    color: #212529;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #212529;
                }
                QTableWidget::item:selected {
                    background-color: #e9ecef;
                    color: #212529;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #dee2e6;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #2c3e50;
                }
                QTableWidget::item:hover {
                    background-color: #f1f3f5;
                }
            """
        
        self.table.setStyleSheet(table_style)
        
        # Category filter style
        self.category_filter.setStyleSheet(self._get_combobox_style(colors))
    
    def _get_combobox_style(self, colors):
        return f"""
            QComboBox {{
                padding: 6px 12px;
                border: 1px solid {colors['border']};
                border-radius: 6px;
                background: {colors['card_bg']};
                color: {colors['text']};
                font-size: 10pt;
                min-width: 120px;
            }}
            QComboBox:focus {{
                border-color: #5865f2;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox QAbstractItemView {{
                background-color: {colors['card_bg']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                color: {colors['text']};
                selection-background-color: #5865f2;
                selection-color: white;
                padding: 4px;
            }}
            QComboBox QAbstractItemView::item {{
                padding: 4px 8px;
                border-radius: 2px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: {colors['bg_hover']};
            }}
            QComboBox QAbstractItemView::item:selected {{
                background-color: #5865f2;
                color: white;
            }}
        """
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Summary cards using SummaryCardWidget with SVG icons
        card_layout = QHBoxLayout()
        card_layout.setSpacing(10)
        
        # Total Expenses Card
        self.total_card = self.parent_dialog.create_summary_card(
            title="Total Expenses",
            value="0",
            icon="money_off",
            color="#e74c3c",
            icon_is_svg=True
        )
        card_layout.addWidget(self.total_card, 1)
        
        # Number of Expenses Card
        self.count_card = self.parent_dialog.create_summary_card(
            title="Number of Expenses",
            value="0",
            icon="receipt_long",
            color="#3498db",
            icon_is_svg=True
        )
        card_layout.addWidget(self.count_card, 1)
        
        # Average Expense Card
        self.avg_card = self.parent_dialog.create_summary_card(
            title="Average Expense",
            value="0",
            icon="analytics",
            color="#f39c12",
            icon_is_svg=True
        )
        card_layout.addWidget(self.avg_card, 1)
        
        layout.addLayout(card_layout)
        
        # Category filter with SVG icon concept
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        filter_layout.setContentsMargins(0, 8, 0, 8)
        
        filter_layout.addWidget(QLabel("Category:"))
        self.category_filter = QComboBox()
        self.category_filter.addItem("All Categories")
        self.category_filter.currentTextChanged.connect(self.on_category_changed)
        self.category_filter.setStyleSheet("""
            QComboBox {
                padding: 6px 12px;
                border: 1px solid #dfe6e9;
                border-radius: 6px;
                background: white;
                font-size: 10pt;
                min-width: 120px;
            }
            QComboBox:focus {
                border-color: #5865f2;
            }
        """)
        filter_layout.addWidget(self.category_filter)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Date", "Expense No", "Category", "Description", "Amount", "Payment Method"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        
        # Apply initial theme
        self._apply_theme()
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        self.load_categories()
        self.setLayout(layout)
    
    def load_categories(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM expense_categories ORDER BY name")
            rows = cursor.fetchall()
            for row in rows:
                self.category_filter.addItem(row[0])
            conn.close()
        except Exception as e:
            logger.error(f"Failed to load categories: {e}")
    
    def on_category_changed(self):
        if not self._is_loading:
            self.current_page = 1
            self.parent_dialog.refresh_current_tab()
    
    def on_page_changed(self, page, page_size):
        """Handle page change from pagination widget"""
        self.current_page = page
        self.page_size = page_size
        self.refresh_current_page()
    
    def refresh_current_page(self):
        """Refresh current page only"""
        if self._is_loading:
            return
        
        from_date, to_date = self.parent_dialog.get_date_range()
        category = self.category_filter.currentText()
        self.refresh(from_date, to_date, category, self.current_page, self.page_size)
    
    def refresh(self, from_date, to_date, category=None, page=1, page_size=25):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.current_page = page
        self.page_size = page_size
        
        if category is None:
            category = self.category_filter.currentText()
        
        self.table.setRowCount(0)
        
        # Use set_value() instead of amount_label.setText()
        self.total_card.set_value("Loading...")
        self.count_card.set_value("Loading...")
        self.avg_card.set_value("Loading...")
        
        worker = ExpenseReportWorker(from_date, to_date, category, page, page_size)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        
        self._current_data = result
        self._is_loading = False
        self.total_items = result.get('total_count', 0)
        
        # Update pagination
        self.pagination.set_total_items(self.total_items, emit_signal=False)
        
        # Update cards using update_summary_card
        self.parent_dialog.update_summary_card(self.total_card, result['total_expenses'], symbol)
        self.count_card.set_value(str(result['expense_count']))
        self.parent_dialog.update_summary_card(self.avg_card, result['avg_expense'], symbol)
        
        # Update table
        is_dark = is_dark_theme()
        text_color = "#dcddde" if is_dark else "#212529"
        
        self.table.setRowCount(len(result['rows']))
        for i, row in enumerate(result['rows']):
            # Date
            date_item = QTableWidgetItem(row[0] or "")
            date_item.setForeground(QColor(text_color))
            self.table.setItem(i, 0, date_item)
            
            # Expense No
            no_item = QTableWidgetItem(row[1] or "")
            no_item.setForeground(QColor(text_color))
            self.table.setItem(i, 1, no_item)
            
            # Category
            cat_item = QTableWidgetItem(row[2] or "")
            cat_item.setForeground(QColor(text_color))
            self.table.setItem(i, 2, cat_item)
            
            # Description
            desc_item = QTableWidgetItem(row[3] or "")
            desc_item.setForeground(QColor(text_color))
            self.table.setItem(i, 3, desc_item)
            
            # Amount
            amount_item = QTableWidgetItem(format_money(row[4], symbol))
            if row[4] > 0:
                amount_item.setForeground(QColor("#ed4245" if is_dark else "#dc3545"))
            else:
                amount_item.setForeground(QColor(text_color))
            self.table.setItem(i, 4, amount_item)
            
            # Payment Method
            method_item = QTableWidgetItem(row[5] or "")
            method_item.setForeground(QColor(text_color))
            self.table.setItem(i, 5, method_item)
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel (exports ALL data, not just current page)"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"expense_report_{from_date}_to_{to_date}.xlsx",
            "Export Expense Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT expense_date, expense_no, category, description, amount, payment_method
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
                ORDER BY expense_date DESC
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0), COUNT(*), COALESCE(AVG(amount), 0)
                FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses, count, avg_expense = cursor.fetchone()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Expense Report"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "EXPENSE REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Expenses: {format_money(total_expenses, symbol)}"
            ws['A7'] = f"Transactions: {count}"
            ws['A8'] = f"Average Expense: {format_money(avg_expense, symbol)}"
            
            # Headers
            headers = ["Date", "Expense No", "Category", "Description", "Amount", "Payment Method"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row_idx, row in enumerate(rows, start=11):
                ws.cell(row=row_idx, column=1, value=row[0] or "")
                ws.cell(row=row_idx, column=2, value=row[1] or "")
                ws.cell(row=row_idx, column=3, value=row[2] or "")
                ws.cell(row=row_idx, column=4, value=row[3] or "")
                ws.cell(row=row_idx, column=5, value=float(row[4]) if row[4] else 0)
                ws.cell(row=row_idx, column=6, value=row[5] or "")
            
            # Auto adjust columns
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def refresh_from_parent(self, from_date, to_date):
        """Called from parent dialog when date range changes"""
        self.current_page = 1
        category = self.category_filter.currentText()
        self.refresh(from_date, to_date, category, 1, self.page_size)
    
    def retranslateUi(self):
        """Retranslate UI for language change"""
        lang = self.parent_dialog.get_lang() if hasattr(self.parent_dialog, 'get_lang') else "en"
        
        if lang == "my":
            self.total_card.set_title("စုစုပေါင်းအသုံးစရိတ်")
            self.count_card.set_title("အသုံးစရိတ်အရေအတွက်")
            self.avg_card.set_title("ပျမ်းမျှအသုံးစရိတ်")
            self.table.setHorizontalHeaderLabels(["ရက်စွဲ", "အမှတ်", "အမျိုးအစား", "ဖော်ပြချက်", "ပမာဏ", "ငွေပေးချေမှုနည်းလမ်း"])
            self.category_filter.setItemText(0, "အားလုံး")
        else:
            self.total_card.set_title("Total Expenses")
            self.count_card.set_title("Number of Expenses")
            self.avg_card.set_title("Average Expense")
            self.table.setHorizontalHeaderLabels(["Date", "Expense No", "Category", "Description", "Amount", "Payment Method"])
            self.category_filter.setItemText(0, "All Categories")
        
        # Update card icons
        self.total_card.set_icon("money_off", is_svg=True, size=(24, 24))
        self.count_card.set_icon("receipt_long", is_svg=True, size=(24, 24))
        self.avg_card.set_icon("analytics", is_svg=True, size=(24, 24))
        
        self._apply_theme()