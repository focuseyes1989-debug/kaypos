# ui/reports/profit_loss_report_dialog.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QTableWidget, QTableWidgetItem, QHeaderView, QFrame
from PyQt6.QtCore import Qt, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import format_money
from ui.reports.base_report_dialog import BaseReportDialog
from ui.widgets.summary_card_widget import SummaryCardWidget
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime
import csv
import os


class ProfitLossWorker(QObject):
    """Worker for calculating profit/loss in background thread"""
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
    
    def run(self):
        """Run the calculation in background thread"""
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            # Total Sales
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_sales = cursor.fetchone()[0]
            
            # COGS
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' 
                  AND date(sales.created_at) BETWEEN ? AND ?
                  AND (products.sold_by IS NULL OR products.sold_by != 'Service')
            """, (self.from_date, self.to_date))
            total_cogs = cursor.fetchone()[0]
            
            # Gross Profit
            gross_profit = total_sales - total_cogs
            
            # Total Expenses
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_expenses = cursor.fetchone()[0]
            
            # Net Profit
            net_profit = gross_profit - total_expenses
            
            # Profit Margin
            profit_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            conn.close()
            
            result = {
                'total_sales': total_sales,
                'total_cogs': total_cogs,
                'gross_profit': gross_profit,
                'total_expenses': total_expenses,
                'net_profit': net_profit,
                'profit_margin': profit_margin,
            }
            
            self.result.emit(result)
            
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class ProfitLossReportDialog(BaseReportDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_dark = is_dark_theme()
        self.create_content_area()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        # Initial refresh
        self.refresh_report()
        self.retranslateUi()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        is_dark = is_dark_theme()
        
        # Update table style
        if is_dark:
            table_style = """
                QTableWidget {
                    background-color: #2f3136;
                    alternate-background-color: #36393f;
                    selection-background-color: #40444b;
                    selection-color: #dcddde;
                    gridline-color: #40444b;
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    color: #dcddde;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #dcddde;
                }
                QTableWidget::item:selected {
                    background-color: #40444b;
                    color: #dcddde;
                }
                QHeaderView::section {
                    background-color: #202225;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #40444b;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #b9bbbe;
                }
                QTableWidget::item:hover {
                    background-color: #40444b;
                }
            """
        else:
            table_style = """
                QTableWidget {
                    background-color: white;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e9ecef;
                    selection-color: #212529;
                    gridline-color: #dee2e6;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    color: #212529;
                }
                QTableWidget::item {
                    padding: 8px 12px;
                    color: #212529;
                }
                QTableWidget::item:selected {
                    background-color: #e9ecef;
                    color: #212529;
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    padding: 8px 12px;
                    border: none;
                    border-bottom: 2px solid #dee2e6;
                    font-weight: 600;
                    font-size: 10pt;
                    color: #2c3e50;
                }
                QTableWidget::item:hover {
                    background-color: #f1f3f5;
                }
            """
        self.table.setStyleSheet(table_style)
        
        # Update labels
        for child in self.findChildren(QLabel):
            child.setStyleSheet(f"color: {colors['text']};")
    
    def create_content_area(self):
        """Create main content area with cards and table"""
        # Summary cards layout with SVG icons
        card_layout = QHBoxLayout()
        card_layout.setSpacing(12)
        
        # Sales Card
        self.sales_card = SummaryCardWidget(
            title="Total Sales",
            value="0",
            icon="attach_money",
            color="#3498db",
            icon_is_svg=True
        )
        self.sales_card.set_icon("attach_money", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.sales_card, 1)
        
        # COGS Card
        self.cogs_card = SummaryCardWidget(
            title="COGS",
            value="0",
            icon="package",
            color="#e74c3c",
            icon_is_svg=True
        )
        self.cogs_card.set_icon("package", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.cogs_card, 1)
        
        # Gross Profit Card
        self.gross_card = SummaryCardWidget(
            title="Gross Profit",
            value="0",
            icon="trending_up",
            color="#2ecc71",
            icon_is_svg=True
        )
        self.gross_card.set_icon("trending_up", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.gross_card, 1)
        
        # Expenses Card
        self.expenses_card = SummaryCardWidget(
            title="Operating Expenses",
            value="0",
            icon="money_off",
            color="#e67e22",
            icon_is_svg=True
        )
        self.expenses_card.set_icon("money_off", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.expenses_card, 1)
        
        # Net Profit Card
        self.net_card = SummaryCardWidget(
            title="Net Profit",
            value="0",
            icon="bar_chart",
            color="#9b59b6",
            icon_is_svg=True
        )
        self.net_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.net_card, 1)
        
        # Margin Card
        self.margin_card = SummaryCardWidget(
            title="Net Margin",
            value="0%",
            icon="analytics",
            color="#1abc9c",
            icon_is_svg=True
        )
        self.margin_card.set_icon("analytics", is_svg=True, size=(24, 24))
        card_layout.addWidget(self.margin_card, 1)
        
        self.main_layout.insertLayout(2, card_layout)
        
        # Summary table with theme support
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["Metric", "Amount", "% of Sales", "Status", "Trend"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Apply initial theme
        self._apply_theme()
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        
        self.main_layout.insertWidget(3, self.table)
    
    def create_worker(self):
        """Create worker for background refresh"""
        from_date, to_date = self.get_date_range()
        return ProfitLossWorker(from_date, to_date)
    
    def update_ui_with_result(self, result):
        """Update UI with calculation result"""
        symbol = self.get_currency_symbol()
        is_dark = is_dark_theme()
        
        total_sales = result['total_sales']
        total_cogs = result['total_cogs']
        gross_profit = result['gross_profit']
        total_expenses = result['total_expenses']
        net_profit = result['net_profit']
        profit_margin = result['profit_margin']
        
        # Color definitions
        green_color = "#3ba55d" if is_dark else "#28a745"
        red_color = "#ed4245" if is_dark else "#dc3545"
        text_color = "#dcddde" if is_dark else "#212529"
        
        # Update cards
        self.sales_card.set_value(format_money(total_sales, symbol))
        self.cogs_card.set_value(format_money(total_cogs, symbol))
        self.gross_card.set_value(format_money(gross_profit, symbol))
        self.expenses_card.set_value(format_money(total_expenses, symbol))
        self.net_card.set_value(format_money(net_profit, symbol))
        self.margin_card.set_value(f"{profit_margin:.1f}%")
        
        # Color coding for net profit
        if net_profit >= 0:
            self.net_card.set_color(green_color)
            self.margin_card.set_color(green_color)
        else:
            self.net_card.set_color(red_color)
            self.margin_card.set_color(red_color)
        
        # Update table
        data = [
            ("Total Sales", total_sales, 100, "Income", "↑"),
            ("COGS", total_cogs, (total_cogs/total_sales*100) if total_sales > 0 else 0, "Expense", "↓"),
            ("Gross Profit", gross_profit, (gross_profit/total_sales*100) if total_sales > 0 else 0, 
             "Profit" if gross_profit >= 0 else "Loss", "↑" if gross_profit >= 0 else "↓"),
            ("Operating Expenses", total_expenses, (total_expenses/total_sales*100) if total_sales > 0 else 0, "Expense", "↓"),
            ("Net Profit", net_profit, (net_profit/total_sales*100) if total_sales > 0 else 0,
             "Profit" if net_profit >= 0 else "Loss", "↑" if net_profit >= 0 else "↓"),
        ]
        
        self.table.setRowCount(len(data))
        for i, (metric, amount, percentage, status, trend) in enumerate(data):
            # Metric
            metric_item = QTableWidgetItem(metric)
            metric_item.setForeground(QColor(text_color))
            self.table.setItem(i, 0, metric_item)
            
            # Amount
            amount_item = QTableWidgetItem(format_money(amount, symbol))
            amount_item.setForeground(QColor(text_color))
            self.table.setItem(i, 1, amount_item)
            
            # Percentage
            percent_item = QTableWidgetItem(f"{percentage:.1f}%")
            percent_item.setForeground(QColor(text_color))
            self.table.setItem(i, 2, percent_item)
            
            # Status
            status_item = QTableWidgetItem(status)
            if status == "Profit" or status == "Income":
                status_item.setForeground(QColor(green_color))
            elif status == "Loss":
                status_item.setForeground(QColor(red_color))
            else:
                status_item.setForeground(QColor(text_color))
            self.table.setItem(i, 3, status_item)
            
            # Trend
            trend_item = QTableWidgetItem(trend)
            if trend == "↑":
                trend_item.setForeground(QColor(green_color))
            elif trend == "↓":
                trend_item.setForeground(QColor(red_color))
            else:
                trend_item.setForeground(QColor(text_color))
            self.table.setItem(i, 4, trend_item)
    
    def export_to_excel(self):
        """Export to Excel"""
        from_date, to_date = self.get_date_range()
        
        from utils.excel_exporter import ExcelExporter
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"profit_loss_{from_date}_to_{to_date}.xlsx",
            "Export Profit & Loss Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = self.get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(products.cost * sale_items.qty), 0)
                FROM sale_items
                JOIN sales ON sale_items.sale_id = sales.id
                JOIN products ON sale_items.product_name = products.name
                WHERE sales.status = 'completed' AND date(sales.created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_cogs = cursor.fetchone()[0]
            
            cursor.execute("""
                SELECT COALESCE(SUM(amount), 0) FROM expenses
                WHERE expense_date BETWEEN ? AND ?
            """, (from_date, to_date))
            total_expenses = cursor.fetchone()[0]
            
            conn.close()
            
            gross_profit = total_sales - total_cogs
            net_profit = gross_profit - total_expenses
            profit_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Profit & Loss"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "PROFIT & LOSS REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Headers
            headers = ["Metric", "Amount", "% of Sales"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            data = [
                ("Total Sales", total_sales, "100%"),
                ("COGS", total_cogs, f"{(total_cogs/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Gross Profit", gross_profit, f"{(gross_profit/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Operating Expenses", total_expenses, f"{(total_expenses/total_sales*100):.1f}%" if total_sales > 0 else "0%"),
                ("Net Profit", net_profit, f"{profit_margin:.1f}%"),
            ]
            
            for row_idx, (metric, amount, percentage) in enumerate(data, start=6):
                ws.cell(row=row_idx, column=1, value=metric)
                ws.cell(row=row_idx, column=2, value=format_money(amount, symbol))
                ws.cell(row=row_idx, column=3, value=percentage)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def refresh_report(self):
        """Refresh the report with current date range"""
        from_date, to_date = self.get_date_range()
        
        # Show loading state
        loading_text = "Loading..."
        self.sales_card.set_value(loading_text)
        self.cogs_card.set_value(loading_text)
        self.gross_card.set_value(loading_text)
        self.expenses_card.set_value(loading_text)
        self.net_card.set_value(loading_text)
        self.margin_card.set_value(loading_text)
        
        worker = self.create_worker()
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.update_ui_with_result)
        worker.error.connect(self.on_refresh_error)
        
        self.threads.append(thread)
        self.workers.append(worker)
        thread.start()
    
    def retranslateUi(self):
        """Update UI text based on language"""
        lang = self.get_lang()
        
        if lang == "my":
            self.setWindowTitle("အမြတ်အစွန်း အစီရင်ခံစာ")
            self.btn_export.setText(" Excel ထုတ်မည်")
            self.btn_close.setText(" ပိတ်မည်")
            self.btn_refresh.setText(" ပြန်လည်")
            
            self.sales_card.set_title("စုစုပေါင်းရောင်းအား")
            self.cogs_card.set_title("ကုန်ပစ္စည်းကုန်ကျစရိတ်")
            self.gross_card.set_title("အသားတင်အမြတ်")
            self.expenses_card.set_title("လည်ပတ်စရိတ်")
            self.net_card.set_title("အသားတင်အမြတ်")
            self.margin_card.set_title("အသားတင်အမြတ်ရာခိုင်နှုန်း")
            
            self.table.setHorizontalHeaderLabels([
                "အမျိုးအစား", "ပမာဏ", "ရောင်းအား၏ ရာခိုင်နှုန်း", "အခြေအနေ", "လမ်းကြောင်း"
            ])
        else:
            self.setWindowTitle("Profit & Loss Report")
            self.btn_export.setText(" Export Excel")
            self.btn_close.setText(" Close")
            self.btn_refresh.setText(" Refresh")
            
            self.sales_card.set_title("Total Sales")
            self.cogs_card.set_title("COGS")
            self.gross_card.set_title("Gross Profit")
            self.expenses_card.set_title("Operating Expenses")
            self.net_card.set_title("Net Profit")
            self.margin_card.set_title("Net Margin")
            
            self.table.setHorizontalHeaderLabels([
                "Metric", "Amount", "% of Sales", "Status", "Trend"
            ])
        
        # Update button icons
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_close.set_icon("close", size=(16, 16))
        
        # Update card icons
        self.sales_card.set_icon("attach_money", is_svg=True, size=(24, 24))
        self.cogs_card.set_icon("package", is_svg=True, size=(24, 24))
        self.gross_card.set_icon("trending_up", is_svg=True, size=(24, 24))
        self.expenses_card.set_icon("money_off", is_svg=True, size=(24, 24))
        self.net_card.set_icon("bar_chart", is_svg=True, size=(24, 24))
        self.margin_card.set_icon("analytics", is_svg=True, size=(24, 24))
        
        self._apply_theme()