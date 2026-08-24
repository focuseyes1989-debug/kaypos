# ui/reports/sales_report.py
from PyQt6.QtWidgets import QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QWidget
from PyQt6.QtCore import QThread, QObject, pyqtSignal
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from ui.widgets import PaginationWidget
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from loguru import logger
from datetime import datetime


class SalesReportWorker(QObject):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    result = pyqtSignal(dict)
    
    def __init__(self, from_date, to_date, page=1, page_size=25):
        super().__init__()
        self.from_date = from_date
        self.to_date = to_date
        self.page = page
        self.page_size = page_size
    
    def run(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT COUNT(*) FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
            """, (self.from_date, self.to_date))
            total_count = cursor.fetchone()[0]
            
            offset = (self.page - 1) * self.page_size
            
            cursor.execute("""
                SELECT s.created_at, s.invoice_no, c.name, s.total, s.payment, s.change_amount
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                ORDER BY s.created_at DESC
                LIMIT ? OFFSET ?
            """, (self.from_date, self.to_date, self.page_size, offset))
            rows = cursor.fetchall()
            
            total_sales = sum(row[3] for row in rows) if rows else 0
            transaction_count = len(rows)
            avg_sale = total_sales / transaction_count if transaction_count > 0 else 0
            
            conn.close()
            
            self.result.emit({
                'rows': rows,
                'total_sales': total_sales,
                'transaction_count': transaction_count,
                'avg_sale': avg_sale,
                'total_count': total_count,
                'current_page': self.page,
                'page_size': self.page_size,
                'total_pages': (total_count + self.page_size - 1) // self.page_size if total_count > 0 else 1
            })
        except Exception as e:
            self.error.emit(str(e))
        finally:
            self.finished.emit()


class SalesReportTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self._is_loading = False
        self._current_data = None
        self.current_page = 1
        self.page_size = 25
        self.total_items = 0
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        self.setup_ui()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        colors = get_theme_colors()
        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {colors['card_bg']};
                alternate-background-color: {colors['table_alt']};
                selection-background-color: {colors['bg_hover']};
                selection-color: {colors['text']};
                gridline-color: transparent;
                border: 1px solid {colors['border']};
                border-radius: 12px;
                color: {colors['text']};
            }}
            QTableWidget::item {{ padding: 8px 12px; }}
            QHeaderView::section {{
                background-color: {colors['bg_hover']};
                padding: 9px 12px;
                border: none;
                border-bottom: 1px solid {colors['border']};
                font-weight: 600;
                color: {colors['text_secondary']};
            }}
        """)
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Summary cards using SummaryCardWidget with SVG icons
        card_layout = QHBoxLayout()
        card_layout.setSpacing(10)
        
        # Total Sales Card
        self.total_card = self.parent_dialog.create_summary_card(
            title="Total Sales",
            value="0",
            icon="attach_money",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout.addWidget(self.total_card, 1)
        
        # Transactions Card
        self.count_card = self.parent_dialog.create_summary_card(
            title="Transactions",
            value="0",
            icon="receipt_long",
            color="#3498db",
            icon_is_svg=True
        )
        card_layout.addWidget(self.count_card, 1)
        
        # Average Sale Card
        self.avg_card = self.parent_dialog.create_summary_card(
            title="Average Sale",
            value="0",
            icon="analytics",
            color="#f39c12",
            icon_is_svg=True
        )
        card_layout.addWidget(self.avg_card, 1)
        
        layout.addLayout(card_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Date", "Invoice No", "Customer", "Total", "Payment", "Change"])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        self.table.verticalHeader().setVisible(False)
        
        # Apply initial theme
        self._apply_theme()
        
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.table)
        
        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        self.setLayout(layout)
    
    def on_page_changed(self, page, page_size):
        self.current_page = page
        self.page_size = page_size
        self.refresh_current_page()
    
    def refresh_current_page(self):
        if self._is_loading:
            return
        from_date, to_date = self.parent_dialog.get_date_range()
        self.refresh(from_date, to_date, self.current_page, self.page_size)
    
    def refresh(self, from_date, to_date, page=1, page_size=25):
        if self._is_loading:
            return
        
        self._is_loading = True
        self.current_page = page
        self.page_size = page_size
        
        self.table.setRowCount(0)
        self.total_card.set_value("Loading...")
        self.count_card.set_value("Loading...")
        self.avg_card.set_value("Loading...")
        
        worker = SalesReportWorker(from_date, to_date, page, page_size)
        thread = QThread()
        worker.moveToThread(thread)
        
        thread.started.connect(worker.run)
        worker.finished.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        worker.result.connect(self.on_refresh_complete)
        worker.error.connect(self.on_refresh_error)
        
        self.parent_dialog.threads.append(thread)
        self.parent_dialog.workers.append(worker)
        thread.start()
    
    def on_refresh_complete(self, result):
        symbol = get_currency_symbol()
        is_dark = is_dark_theme()
        text_color = "#dcddde" if is_dark else "#212529"
        green_color = "#3ba55d" if is_dark else "#28a745"
        
        self._current_data = result
        self._is_loading = False
        self.total_items = result.get('total_count', 0)
        
        self.pagination.set_total_items(self.total_items, emit_signal=False)
        
        # Update cards using summary card methods
        self.parent_dialog.update_summary_card(self.total_card, result['total_sales'], symbol)
        self.count_card.set_value(str(result['transaction_count']))
        self.parent_dialog.update_summary_card(self.avg_card, result['avg_sale'], symbol)
        
        self.table.setRowCount(len(result['rows']))
        for i, row in enumerate(result['rows']):
            # Date
            date_item = QTableWidgetItem(str(row[0])[:16] if row[0] else "")
            date_item.setForeground(QColor(text_color))
            self.table.setItem(i, 0, date_item)
            
            # Invoice No
            inv_item = QTableWidgetItem(row[1] or "")
            inv_item.setForeground(QColor(text_color))
            self.table.setItem(i, 1, inv_item)
            
            # Customer
            cust_item = QTableWidgetItem(row[2] if row[2] else "Walk-in")
            cust_item.setForeground(QColor(text_color))
            self.table.setItem(i, 2, cust_item)
            
            # Total
            total_item = QTableWidgetItem(format_money(row[3], symbol))
            total_item.setForeground(QColor(green_color))
            self.table.setItem(i, 3, total_item)
            
            # Payment
            payment_item = QTableWidgetItem(format_money(row[4], symbol))
            payment_item.setForeground(QColor(text_color))
            self.table.setItem(i, 4, payment_item)
            
            # Change
            change_item = QTableWidgetItem(format_money(row[5], symbol))
            change_item.setForeground(QColor(text_color))
            self.table.setItem(i, 5, change_item)
        
        self.parent_dialog.on_refresh_complete()
    
    def on_refresh_error(self, error_msg):
        self._is_loading = False
        self.parent_dialog.on_refresh_error(error_msg)
    
    def export(self, from_date, to_date):
        """Export to Excel"""
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"sales_report_{from_date}_to_{to_date}.xlsx",
            "Export Sales Report"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT s.created_at, s.invoice_no, c.name, s.total, s.payment, s.change_amount
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
                ORDER BY s.created_at DESC
            """, (from_date, to_date))
            rows = cursor.fetchall()
            
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0), COUNT(*), COALESCE(AVG(total), 0)
                FROM sales
                WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_sales, count, avg_sale = cursor.fetchone()
            conn.close()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "SALES REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f"Period: {from_date} to {to_date}"
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # Summary
            ws['A5'] = "Summary"
            ws['A5'].font = Font(bold=True)
            ws['A6'] = f"Total Sales: {format_money(total_sales, symbol)}"
            ws['A7'] = f"Transactions: {count}"
            ws['A8'] = f"Average Sale: {format_money(avg_sale, symbol)}"
            
            # Headers
            headers = ["Date", "Invoice No", "Customer", "Total", "Payment", "Change"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=10, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_idx, row in enumerate(rows, start=11):
                ws.cell(row=row_idx, column=1, value=str(row[0]) if row[0] else "")
                ws.cell(row=row_idx, column=2, value=row[1] or "")
                ws.cell(row=row_idx, column=3, value=row[2] if row[2] else "Walk-in")
                ws.cell(row=row_idx, column=4, value=float(row[3]) if row[3] else 0)
                ws.cell(row=row_idx, column=5, value=float(row[4]) if row[4] else 0)
                ws.cell(row=row_idx, column=6, value=float(row[5]) if row[5] else 0)
            
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].auto_size = True
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def refresh_from_parent(self, from_date, to_date):
        self.current_page = 1
        self.refresh(from_date, to_date, 1, self.page_size)
    
    def retranslateUi(self):
        """Retranslate UI for language change"""
        lang = self.parent_dialog.get_lang() if hasattr(self.parent_dialog, 'get_lang') else "en"
        
        if lang == "my":
            self.total_card.set_title("စုစုပေါင်းရောင်းအား")
            self.count_card.set_title("ငွေပေးချေမှုအရေအတွက်")
            self.avg_card.set_title("ပျမ်းမျှရောင်းအား")
            self.table.setHorizontalHeaderLabels([
                "ရက်စွဲ", "ပြေစာအမှတ်", "ဝယ်ယူသူ", "စုစုပေါင်း", "ငွေပေးချေမှု", "ပြန်အမ်းငွေ"
            ])
        else:
            self.total_card.set_title("Total Sales")
            self.count_card.set_title("Transactions")
            self.avg_card.set_title("Average Sale")
            self.table.setHorizontalHeaderLabels([
                "Date", "Invoice No", "Customer", "Total", "Payment", "Change"
            ])
        
        # Update card icons
        self.total_card.set_icon("attach_money", is_svg=True, size=(24, 24))
        self.count_card.set_icon("receipt_long", is_svg=True, size=(24, 24))
        self.avg_card.set_icon("analytics", is_svg=True, size=(24, 24))
        
        self._apply_theme()
