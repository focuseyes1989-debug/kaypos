# ui/sales_summary/sales_summary_page.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTabWidget, QFrame, QApplication
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QIcon, QColor, QPixmap
from ui.sales_summary.base_sales_summary import BaseSalesSummary
from ui.sales_summary.top_items_tab import TopItemsTab
from ui.sales_summary.items_tab import ItemsTab
from ui.sales_summary.categories_tab import CategoriesTab
from ui.sales_summary.category_parents_tab import CategoryParentsTab
from ui.sales_summary.category_groups_tab import CategoryGroupsTab
from ui.sales_summary.payment_tab import PaymentTab
from utils.language import lang
from utils.excel_exporter import ExcelExporter
from models.database import connect_db
from datetime import datetime

# Import widgets
from ui.widgets import (
    DateRangeWidget,
    ActionButtonWidget,
    ToastNotificationWidget,
    LoadingSpinnerWidget,
    SummaryCardWidget,
    ModernButton  # ✅ Added ModernButton import
)
from ui.themes.theme_manager import theme_manager, is_dark_theme
from utils.currency import get_currency_symbol, format_money
import os


class SalesSummaryPage(BaseSalesSummary):
    def __init__(self):
        super().__init__()
        main_layout = QVBoxLayout()
        main_layout.setSpacing(15)

        # ========== Toast Notification ==========
        self.toast = ToastNotificationWidget(self)

        # ========== Filter Row with DateRangeWidget ==========
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        self.date_range = DateRangeWidget()
        self.date_range.date_range_changed.connect(self.on_date_range_changed)
        filter_layout.addWidget(self.date_range)

        filter_layout.addStretch()
        
        # ✅ Export button - Using ModernButton with icon
        self.btn_export_excel = ModernButton("Export Excel", ModernButton.PRIMARY)
        self.btn_export_excel.set_icon("file_export", size=(18, 18))
        self.btn_export_excel.set_compact(True)
        self.btn_export_excel.clicked.connect(self.export_to_excel)
        
        filter_layout.addWidget(self.btn_export_excel)

        main_layout.addLayout(filter_layout)

        # ========== Summary Cards with SVG Icons ==========
        card_layout = QHBoxLayout()
        card_layout.setSpacing(12)

        # Total Sales
        self.total_sales_card = SummaryCardWidget(
            title="Total Sales", 
            value="0", 
            icon="attach_money",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout.addWidget(self.total_sales_card)

        # Total Orders
        self.total_orders_card = SummaryCardWidget(
            title="Total Orders", 
            value="0", 
            icon="orders",
            color="#3498db",
            icon_is_svg=True
        )
        card_layout.addWidget(self.total_orders_card)

        # Average Order
        self.avg_order_card = SummaryCardWidget(
            title="Average Order", 
            value="0", 
            icon="analytics",
            color="#f39c12",
            icon_is_svg=True
        )
        card_layout.addWidget(self.avg_order_card)

        # Top Category
        self.top_category_card = SummaryCardWidget(
            title="Top Category", 
            value="—", 
            icon="category",
            color="#9b59b6",
            icon_is_svg=True
        )
        card_layout.addWidget(self.top_category_card)

        # Total Discount Card
        self.total_discount_card = SummaryCardWidget(
            title="Total Discount", 
            value="0", 
            icon="percent_discount",
            color="#e74c3c",
            icon_is_svg=True
        )
        card_layout.addWidget(self.total_discount_card)

        card_layout.addStretch()
        main_layout.addLayout(card_layout)

        # ========== Tabs ==========
        self.tabs = QTabWidget()
        self.tab_names = {
            0: "Top 20 Sales by Item",
            1: "Sales by Item",
            2: "Sales by Category",
            3: "Sales by Parent Category",
            4: "Sales by Category Group",
            5: "Sales by Payment Type"
        }
        
        self.tab_icons = {
            0: "leaderboard",
            1: "list_alt",
            2: "grid_view",
            3: "folder_open",
            4: "group_work",
            5: "payments"
        }
        
        self.top_items_tab = TopItemsTab(self)
        self.tabs.addTab(self.top_items_tab, self._load_colored_tab_icon(0), self.tab_names[0])
        
        self.items_tab = ItemsTab(self)
        self.tabs.addTab(self.items_tab, self._load_colored_tab_icon(1), self.tab_names[1])
        
        self.categories_tab = CategoriesTab(self)
        self.tabs.addTab(self.categories_tab, self._load_colored_tab_icon(2), self.tab_names[2])
        
        self.category_parents_tab = CategoryParentsTab(self)
        self.tabs.addTab(self.category_parents_tab, self._load_colored_tab_icon(3), self.tab_names[3])
        
        self.category_groups_tab = CategoryGroupsTab(self)
        self.tabs.addTab(self.category_groups_tab, self._load_colored_tab_icon(4), self.tab_names[4])
        
        self.payment_tab = PaymentTab(self)
        self.tabs.addTab(self.payment_tab, self._load_colored_tab_icon(5), self.tab_names[5])

        self._apply_tab_bar_style()

        main_layout.addWidget(self.tabs)

        # ========== Loading Spinner ==========
        self.spinner = LoadingSpinnerWidget("Loading sales data...")
        self.spinner.hide()
        main_layout.addWidget(self.spinner)

        self.setLayout(main_layout)

        # Connect signals
        lang.language_changed.connect(self.retranslateUi)
        theme_manager.theme_changed.connect(self.on_theme_changed)
        
        # Load initial data
        self.load_all_tabs()
        self.retranslateUi()

    def _apply_tab_bar_style(self):
        """Apply tab bar style based on theme"""
        is_dark = is_dark_theme()
        
        if is_dark:
            self.tabs.setStyleSheet("""
                QTabWidget::pane {
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    background-color: #2f3136;
                }
                QTabBar::tab {
                    background-color: #2f3136;
                    color: #b9bbbe;
                    padding: 8px 16px;
                    margin-right: 2px;
                    border-top-left-radius: 4px;
                    border-top-right-radius: 4px;
                    border: none;
                }
                QTabBar::tab:selected {
                    background-color: #40444b;
                    color: #ffffff;
                }
                QTabBar::tab:hover {
                    background-color: #36393f;
                    color: #ffffff;
                }
                QTabBar::tab:!selected {
                    background-color: #202225;
                    color: #72767d;
                }
            """)
        else:
            self.tabs.setStyleSheet("""
                QTabWidget::pane {
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    background-color: #ffffff;
                }
                QTabBar::tab {
                    background-color: #f8f9fa;
                    color: #495057;
                    padding: 8px 16px;
                    margin-right: 2px;
                    border-top-left-radius: 4px;
                    border-top-right-radius: 4px;
                    border: 1px solid #dee2e6;
                    border-bottom: none;
                }
                QTabBar::tab:selected {
                    background-color: #ffffff;
                    color: #212529;
                    border-bottom: 2px solid #5865f2;
                }
                QTabBar::tab:hover {
                    background-color: #e9ecef;
                    color: #212529;
                }
            """)
        
        self._update_tab_icons_color()

    def _update_tab_icons_color(self):
        """Update all tab icons color based on theme"""
        for index in range(self.tabs.count()):
            icon = self._load_colored_tab_icon(index)
            self.tabs.setTabIcon(index, icon)

    def _load_colored_tab_icon(self, index):
        """Load SVG icon with color based on theme for tabs"""
        from PyQt6.QtGui import QIcon, QPixmap, QPainter, QColor
        import os
        
        icon_name = self.tab_icons.get(index, "")
        if not icon_name:
            return QIcon()
        
        paths = [
            f"assets/icons/{icon_name}.svg",
            f"assets/icons/{icon_name}.png",
        ]
        
        for path in paths:
            if os.path.exists(path):
                try:
                    pixmap = QPixmap(path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(
                            20, 20,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        
                        is_dark = is_dark_theme()
                        color_hex = "#ffffff" if is_dark else "#495057"
                        
                        colored = scaled.copy()
                        painter = QPainter(colored)
                        painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
                        painter.fillRect(colored.rect(), QColor(color_hex))
                        painter.end()
                        
                        return QIcon(colored)
                except Exception as e:
                    print(f"Could not load icon {path}: {e}")
        
        return QIcon()

    def on_theme_changed(self, theme_name):
        """Handle theme change - update all cards and tab icons"""
        self.update_card_theme()
        self._apply_tab_bar_style()
        self._update_tab_icons_color()
        # ✅ Update export button icon when theme changes
        self.btn_export_excel._on_theme_changed(theme_name)

    def update_card_theme(self):
        """Update all cards when theme changes"""
        for card in [self.total_sales_card, self.total_orders_card, 
                     self.avg_order_card, self.top_category_card,
                     self.total_discount_card]:
            if hasattr(card, 'update_theme'):
                card.update_theme()

    def on_date_range_changed(self, from_date, to_date):
        """Handle date range change - auto refresh"""
        self.load_all_tabs()

    def showEvent(self, event):
        self.load_all_tabs()
        self._apply_tab_bar_style()
        super().showEvent(event)

    def get_date_range(self):
        """Override to use DateRangeWidget"""
        return self.date_range.get_from_date(), self.date_range.get_to_date()

    def load_all_tabs(self):
        """Load all tabs with spinner"""
        self.spinner.start()
        
        from_date, to_date = self.date_range.get_from_date(), self.date_range.get_to_date()
        lang_code = self.get_lang()
        
        try:
            self.top_items_tab.load(from_date, to_date, lang_code)
            self.items_tab.load(from_date, to_date)
            self.categories_tab.load(from_date, to_date)
            self.category_parents_tab.load(from_date, to_date)
            self.category_groups_tab.load(from_date, to_date)
            self.payment_tab.load(from_date, to_date, lang_code)
            
            # Update summary cards
            self.update_summary_cards(from_date, to_date)
            
        except Exception as e:
            self.toast.show_toast(f"Error loading data: {str(e)}", "error")
        finally:
            self.spinner.stop()

    # ============================================================
    # ✅ FIXED: update_summary_cards() - Using sale_items for totals
    # ============================================================
    def update_summary_cards(self, from_date, to_date):
        """Update summary cards with data - ✅ FIXED: Show Net Sales (after discount) with K/M/B formatting"""
        try:
            symbol = get_currency_symbol()
            conn = connect_db()
            cursor = conn.cursor()
            
            # ✅ Total Net Sales = SUM(si.total) - SUM(COALESCE(s.discount_amount, 0))
            cursor.execute("""
                SELECT 
                    COALESCE(SUM(si.total), 0) as total_sales_before_discount,
                    COALESCE(SUM(COALESCE(s.discount_amount, 0)), 0) as total_discount
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                WHERE s.status = 'completed' 
                AND date(s.created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            row = cursor.fetchone()
            total_before_discount = row[0] if row else 0
            total_discount = row[1] if row else 0
            total_sales = total_before_discount - total_discount  # ✅ Net Sales
            
            # Total Orders
            cursor.execute("""
                SELECT COUNT(*) 
                FROM sales 
                WHERE status = 'completed' 
                AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_orders = cursor.fetchone()[0]
            
            # Total Discount
            cursor.execute("""
                SELECT COALESCE(SUM(discount_amount), 0) 
                FROM sales 
                WHERE status = 'completed' 
                AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
            total_discount_amount = cursor.fetchone()[0]
            
            # Top Category - using Net Sales
            cursor.execute("""
                SELECT 
                    COALESCE(p.category, 'Uncategorized') as category,
                    COALESCE(SUM(si.total) - SUM(COALESCE(s.discount_amount, 0)), 0) as net_sales
                FROM sale_items si
                JOIN sales s ON si.sale_id = s.id
                LEFT JOIN products p ON si.product_name = p.name
                WHERE s.status = 'completed' 
                AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY p.category
                ORDER BY net_sales DESC
                LIMIT 1
            """, (from_date, to_date))
            top_category_row = cursor.fetchone()
            
            conn.close()
            
            # ✅ Update ALL cards with formatted values
            # 1. Total Sales - with currency symbol and K/M/B formatting
            self.total_sales_card.set_value(
                total_sales, 
                currency_symbol=symbol, 
                is_currency=True
            )
            
            # 2. Total Orders - just number with K/M/B formatting
            self.total_orders_card.set_value(
                total_orders, 
                currency_symbol=None, 
                is_currency=False
            )
            
            # 3. Average Order - with currency symbol and K/M/B formatting
            avg_order = total_sales / total_orders if total_orders > 0 else 0
            self.avg_order_card.set_value(
                avg_order, 
                currency_symbol=symbol, 
                is_currency=True
            )
            
            # 4. Total Discount - with currency symbol and K/M/B formatting
            self.total_discount_card.set_value(
                total_discount_amount, 
                currency_symbol=symbol, 
                is_currency=True
            )
            
            # 5. Top Category - no formatting needed (string value)
            if top_category_row:
                self.top_category_card.set_value_raw(top_category_row[0])
            else:
                self.top_category_card.set_value_raw("—")
                
        except Exception as e:
            print(f"Error updating summary cards: {e}")

    def retranslateUi(self):
        lang_code = self.get_lang()
        
        # Retranslate DateRangeWidget
        self.date_range.retranslateUi(lang_code)
        
        # ✅ Retranslate Export Button
        if lang_code == "my":
            self.btn_export_excel.setText("Excel ထုတ်မည်")
            self.btn_export_excel.set_icon("file_export", size=(18, 18))
        else:
            self.btn_export_excel.setText("Export Excel")
            self.btn_export_excel.set_icon("file_export", size=(18, 18))
        
        # Retranslate Summary Cards
        if lang_code == "my":
            self.total_sales_card.set_title("စုစုပေါင်းရောင်းအား")
            self.total_orders_card.set_title("စုစုပေါင်းအမှာစာ")
            self.avg_order_card.set_title("ပျမ်းမျှအမှာစာ")
            self.top_category_card.set_title("ထိပ်ဆုံးအမျိုးအစား")
            self.total_discount_card.set_title("စုစုပေါင်းလျှော့စျေး")
            
            tab_titles_my = {
                0: "ထိပ်ဆုံးရောင်းအားရှိပစ္စည်း ၂၀",
                1: "ပစ္စည်းအလိုက်ရောင်းအား",
                2: "အမျိုးအစားအလိုက်ရောင်းအား",
                3: "မိဘအမျိုးအစားအလိုက်ရောင်းအား",
                4: "အမျိုးအစားအုပ်စုအလိုက်ရောင်းအား",
                5: "ငွေပေးချေမှုအလိုက်ရောင်းအား"
            }
            for idx, title in tab_titles_my.items():
                self.tabs.setTabText(idx, title)
        else:
            self.total_sales_card.set_title("Total Sales")
            self.total_orders_card.set_title("Total Orders")
            self.avg_order_card.set_title("Average Order")
            self.top_category_card.set_title("Top Category")
            self.total_discount_card.set_title("Total Discount")
            
            for idx, title in self.tab_names.items():
                self.tabs.setTabText(idx, title)
        
        # Retranslate tabs
        self.top_items_tab.retranslateUi()
        self.items_tab.retranslateUi()
        self.categories_tab.retranslateUi()
        self.category_parents_tab.retranslateUi()
        self.category_groups_tab.retranslateUi()
        self.payment_tab.retranslateUi()
        
        self.load_all_tabs()

    # ============================================================
    # ✅ FIXED: Export methods - Using sale_items for totals
    # ============================================================

    def export_to_excel(self):
        """Export all sales summary tabs to Excel (.xlsx)"""
        from_date, to_date = self.date_range.get_from_date(), self.date_range.get_to_date()
        symbol = get_currency_symbol()
        lang_code = self.get_lang()
        
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"sales_summary_{from_date}_to_{to_date}.xlsx",
            "Export Sales Summary" if lang_code != "my" else "ရောင်းအားအကျဉ်းချုပ် ထုတ်ရန်"
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            
            # Tab 1: Top Products
            ws1 = wb.active
            ws1.title = "Top Products"
            self._export_top_products(ws1, from_date, to_date, symbol, lang_code)
            
            # Tab 2: Products Detail
            ws2 = wb.create_sheet("Products Detail")
            self._export_products_detail(ws2, from_date, to_date, symbol, lang_code)
            
            # Tab 3: Categories
            ws3 = wb.create_sheet("Categories")
            self._export_categories(ws3, from_date, to_date, symbol, lang_code)
            
            # Tab 4: Parent Categories
            ws4 = wb.create_sheet("Parent Categories")
            self._export_parent_categories(ws4, from_date, to_date, symbol, lang_code)
            
            # Tab 5: Category Groups
            ws5 = wb.create_sheet("Category Groups")
            self._export_category_groups(ws5, from_date, to_date, symbol, lang_code)
            
            # Tab 6: Payment Types
            ws6 = wb.create_sheet("Payment Types")
            self._export_payment(ws6, from_date, to_date, symbol, lang_code)
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def _export_top_products(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        # ✅ FIX: Use sale_items for accurate totals
        cursor.execute("""
            SELECT 
                si.product_name, 
                COALESCE(SUM(si.qty * si.price), 0) as total_sales
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
            GROUP BY si.product_name
            ORDER BY total_sales DESC
            LIMIT 20
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "TOP 20 SALES BY ITEM"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        headers = ["ပစ္စည်းအမည်", "စုစုပေါင်းရောင်းအား"] if lang_code == "my" else ["Product Name", "Total Sales"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, (name, total) in enumerate(rows, start=6):
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=format_money(total, symbol))
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _export_products_detail(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        # ✅ FIX: Use sale_items for accurate totals
        cursor.execute("""
            SELECT 
                si.product_name,
                COALESCE(p.category, 'Uncategorized') as category,
                COALESCE(SUM(si.qty), 0) as total_qty,
                COALESCE(SUM(si.price * si.qty), 0) as gross_sales,
                COALESCE(SUM(s.discount_amount), 0) as total_discount,
                COALESCE(SUM(si.qty * si.price) - SUM(s.discount_amount), 0) as net_sales,
                COALESCE(SUM(p.cost * si.qty), 0) as cogs,
                COALESCE(SUM(si.qty * si.price) - SUM(s.discount_amount) - SUM(p.cost * si.qty), 0) as gross_profit
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_name = p.name
            WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
            GROUP BY si.product_name
            ORDER BY si.product_name
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:H1')
        ws['A1'] = "SALES BY PRODUCT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["ပစ္စည်းအမည်", "အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", 
                      "စုစုပေါင်းရောင်းအား (အကြမ်း)", "လျှော့စျေး", "အသားတင်ရောင်းအား",
                      "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Product Name", "Category", "Items Sold", "Gross Sales", 
                      "Discount", "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_qty = total_gross = total_discount = total_net = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            name, category, qty, gross, discount, net, cogs, profit = row_data
            ws.cell(row=row_idx, column=1, value=name)
            ws.cell(row=row_idx, column=2, value=category)
            ws.cell(row=row_idx, column=3, value=qty)
            ws.cell(row=row_idx, column=4, value=format_money(gross, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(discount, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(net, symbol))
            ws.cell(row=row_idx, column=7, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=8, value=format_money(profit, symbol))
            total_qty += qty
            total_gross += gross
            total_discount += discount
            total_net += net
            total_cogs += cogs
            total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=total_qty)
        ws.cell(row=summary_row, column=4, value=format_money(total_gross, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_discount, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_net, symbol))
        ws.cell(row=summary_row, column=7, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=8, value=format_money(total_profit, symbol))
        
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_categories(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        # ✅ FIX: Use sale_items for accurate totals
        cursor.execute("""
            SELECT 
                COALESCE(p.category, 'Uncategorized') as category,
                COALESCE(SUM(si.qty), 0) as items_sold,
                COALESCE(SUM(si.price * si.qty), 0) as gross_sales,
                COALESCE(SUM(s.discount_amount), 0) as total_discount,
                COALESCE(SUM(si.qty * si.price) - SUM(s.discount_amount), 0) as net_sales,
                COALESCE(SUM(p.cost * si.qty), 0) as cogs
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_name = p.name
            WHERE s.status = 'completed' AND date(s.created_at) BETWEEN ? AND ?
            GROUP BY p.category
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:G1')
        ws['A1'] = "SALES BY CATEGORY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["အမျိုးအစား", "ရောင်းရသည့်အရေအတွက်", 
                      "စုစုပေါင်းရောင်းအား (အကြမ်း)", "လျှော့စျေး", 
                      "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Category", "Items Sold", "Gross Sales", "Discount", 
                      "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_items = total_gross = total_discount = total_net = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            category, items, gross, discount, net, cogs = row_data
            profit = net - cogs
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=items)
            ws.cell(row=row_idx, column=3, value=format_money(gross, symbol))
            ws.cell(row=row_idx, column=4, value=format_money(discount, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(net, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=7, value=format_money(profit, symbol))
            total_items += items
            total_gross += gross
            total_discount += discount
            total_net += net
            total_cogs += cogs
            total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_items)
        ws.cell(row=summary_row, column=3, value=format_money(total_gross, symbol))
        ws.cell(row=summary_row, column=4, value=format_money(total_discount, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_net, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=7, value=format_money(total_profit, symbol))
        
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_parent_categories(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                COALESCE(pc.name, 'No Parent') as parent_name,
                COALESCE(SUM(si.qty), 0) as items_sold,
                COALESCE(SUM(si.price * si.qty), 0) as gross_sales,
                COALESCE(SUM(s.discount_amount), 0) as total_discount,
                COALESCE(SUM(si.qty * si.price) - SUM(s.discount_amount), 0) as net_sales,
                COALESCE(SUM(p.cost * si.qty), 0) as cogs
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_name = p.name
            LEFT JOIN categories c ON p.category_id = c.id
            LEFT JOIN categories pc ON c.parent_id = pc.id
            WHERE s.status = 'completed' 
              AND date(s.created_at) BETWEEN ? AND ?
            GROUP BY pc.id, pc.name
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:G1')
        ws['A1'] = "SALES BY PARENT CATEGORY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["မိဘအမျိုးအစား", "ရောင်းရသည့်အရေအတွက်",
                      "စုစုပေါင်းရောင်းအား (အကြမ်း)", "လျှော့စျေး",
                      "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Parent Category", "Items Sold", "Gross Sales", "Discount",
                      "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_items = total_gross = total_discount = total_net = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            parent_name, items, gross, discount, net, cogs = row_data
            profit = net - cogs
            ws.cell(row=row_idx, column=1, value=parent_name)
            ws.cell(row=row_idx, column=2, value=items)
            ws.cell(row=row_idx, column=3, value=format_money(gross, symbol))
            ws.cell(row=row_idx, column=4, value=format_money(discount, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(net, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=7, value=format_money(profit, symbol))
            total_items += items
            total_gross += gross
            total_discount += discount
            total_net += net
            total_cogs += cogs
            total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_items)
        ws.cell(row=summary_row, column=3, value=format_money(total_gross, symbol))
        ws.cell(row=summary_row, column=4, value=format_money(total_discount, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_net, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=7, value=format_money(total_profit, symbol))
        
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_category_groups(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                COALESCE(cg.name, 'Uncategorized') as group_name,
                COALESCE(SUM(si.qty), 0) as items_sold,
                COALESCE(SUM(si.price * si.qty), 0) as gross_sales,
                COALESCE(SUM(s.discount_amount), 0) as total_discount,
                COALESCE(SUM(si.qty * si.price) - SUM(s.discount_amount), 0) as net_sales,
                COALESCE(SUM(p.cost * si.qty), 0) as cogs
            FROM sale_items si
            JOIN sales s ON si.sale_id = s.id
            LEFT JOIN products p ON si.product_name = p.name
            LEFT JOIN categories c ON p.category = c.name
            LEFT JOIN category_groups cg ON c.group_id = cg.id
            WHERE s.status = 'completed' 
              AND date(s.created_at) BETWEEN ? AND ?
            GROUP BY cg.id, cg.name
            ORDER BY net_sales DESC
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:G1')
        ws['A1'] = "SALES BY CATEGORY GROUP"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["အုပ်စု", "ရောင်းရသည့်အရေအတွက်",
                      "စုစုပေါင်းရောင်းအား (အကြမ်း)", "လျှော့စျေး",
                      "အသားတင်ရောင်းအား", "ကုန်ကျစရိတ်", "အသားတင်အမြတ်"]
        else:
            headers = ["Category Group", "Items Sold", "Gross Sales", "Discount",
                      "Net Sales", "Cost of Goods", "Gross Profit"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_items = total_gross = total_discount = total_net = total_cogs = total_profit = 0
        for row_idx, row_data in enumerate(rows, start=6):
            group_name, items, gross, discount, net, cogs = row_data
            profit = net - cogs
            ws.cell(row=row_idx, column=1, value=group_name)
            ws.cell(row=row_idx, column=2, value=items)
            ws.cell(row=row_idx, column=3, value=format_money(gross, symbol))
            ws.cell(row=row_idx, column=4, value=format_money(discount, symbol))
            ws.cell(row=row_idx, column=5, value=format_money(net, symbol))
            ws.cell(row=row_idx, column=6, value=format_money(cogs, symbol))
            ws.cell(row=row_idx, column=7, value=format_money(profit, symbol))
            total_items += items
            total_gross += gross
            total_discount += discount
            total_net += net
            total_cogs += cogs
            total_profit += profit
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_items)
        ws.cell(row=summary_row, column=3, value=format_money(total_gross, symbol))
        ws.cell(row=summary_row, column=4, value=format_money(total_discount, symbol))
        ws.cell(row=summary_row, column=5, value=format_money(total_net, symbol))
        ws.cell(row=summary_row, column=6, value=format_money(total_cogs, symbol))
        ws.cell(row=summary_row, column=7, value=format_money(total_profit, symbol))
        
        for col in range(1, 8):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _export_payment(self, ws, from_date, to_date, symbol, lang_code):
        from openpyxl.styles import Font, PatternFill, Alignment
        
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COALESCE(payment_type, 'Other') as payment_type,
                   COUNT(*) as transaction_count,
                   COALESCE(SUM(total), 0) as total_amount
            FROM sales
            WHERE status = 'completed' AND date(created_at) BETWEEN ? AND ?
            GROUP BY payment_type
            ORDER BY payment_type
        """, (from_date, to_date))
        rows = cursor.fetchall()
        conn.close()
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "SALES BY PAYMENT TYPE"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A2'] = f"Period: {from_date} to {to_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        if lang_code == "my":
            headers = ["ငွေပေးချေမှုအမျိုးအစား", "ငွေပေးချေမှုအရေအတွက်", "ငွေပေးချေမှုပမာဏ"]
        else:
            headers = ["Payment Type", "Transaction Count", "Amount"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_count = total_amount = 0
        for row_idx, row_data in enumerate(rows, start=6):
            ptype, count, amount = row_data
            ws.cell(row=row_idx, column=1, value=ptype)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=format_money(amount, symbol))
            total_count += count
            total_amount += amount
        
        summary_row = len(rows) + 7
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_count)
        ws.cell(row=summary_row, column=3, value=format_money(total_amount, symbol))
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20