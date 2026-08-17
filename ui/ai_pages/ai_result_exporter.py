"""Export and print already-authorized AI result rows."""

import csv
import html
from pathlib import Path

from PyQt6.QtGui import QPageLayout, QTextDocument
from PyQt6.QtPrintSupport import QPrintDialog, QPrinter


class AIResultExporter:
    @staticmethod
    def export(rows,file_path,title="AI Result"):
        path=Path(file_path);suffix=path.suffix.lower()
        if suffix==".csv":AIResultExporter._csv(rows,path)
        elif suffix==".xlsx":AIResultExporter._xlsx(rows,path,title)
        elif suffix==".pdf":AIResultExporter._pdf(rows,path,title)
        else:raise ValueError("Supported formats are CSV, Excel (.xlsx), and PDF")
        return str(path)

    @staticmethod
    def print_rows(rows,parent=None,title="AI Result"):
        printer=QPrinter(QPrinter.PrinterMode.HighResolution)
        dialog=QPrintDialog(printer,parent)
        if dialog.exec()!=QPrintDialog.DialogCode.Accepted:return False
        AIResultExporter._document(rows,title).print(printer);return True

    @staticmethod
    def _headers(rows):
        headers=[]
        for row in rows:
            for key in row:
                if key not in headers:headers.append(key)
        return headers

    @staticmethod
    def _value(value):
        if value is None:return ""
        if isinstance(value,(bytes,bytearray,memoryview)):return "[binary data omitted]"
        return str(value)

    @classmethod
    def _csv(cls,rows,path):
        headers=cls._headers(rows)
        with path.open("w",newline="",encoding="utf-8-sig") as stream:
            writer=csv.DictWriter(stream,fieldnames=headers,extrasaction="ignore");writer.writeheader()
            writer.writerows({key:cls._value(row.get(key)) for key in headers} for row in rows)

    @classmethod
    def _xlsx(cls,rows,path,title):
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill
        workbook=Workbook();sheet=workbook.active;sheet.title="AI Result"
        headers=cls._headers(rows);sheet.append(headers)
        for cell in sheet[1]:cell.font=Font(bold=True,color="FFFFFF");cell.fill=PatternFill("solid",fgColor="5865F2")
        for row in rows:sheet.append([cls._value(row.get(key)) for key in headers])
        sheet.freeze_panes="A2";sheet.auto_filter.ref=sheet.dimensions
        for column in sheet.columns:
            width=min(45,max(10,max(len(str(cell.value or "")) for cell in column)+2));sheet.column_dimensions[column[0].column_letter].width=width
        sheet.sheet_properties.pageSetUpPr.fitToPage=True;sheet.page_setup.fitToWidth=1
        workbook.save(path)

    @classmethod
    def _pdf(cls,rows,path,title):
        printer=QPrinter(QPrinter.PrinterMode.HighResolution);printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat);printer.setOutputFileName(str(path))
        printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        cls._document(rows,title).print(printer)

    @classmethod
    def _document(cls,rows,title):
        headers=cls._headers(rows)
        head="".join(f"<th>{html.escape(str(key))}</th>" for key in headers)
        body="".join("<tr>"+"".join(f"<td>{html.escape(cls._value(row.get(key)))}</td>" for key in headers)+"</tr>" for row in rows)
        document=QTextDocument();document.setHtml(f"""<html><head><style>body{{font-family:Arial,sans-serif;font-size:9pt}}h2{{color:#2d3436}}table{{border-collapse:collapse;width:100%}}th{{background:#5865f2;color:white}}th,td{{border:1px solid #ccd1d9;padding:5px;text-align:left}}</style></head><body><h2>{html.escape(title)}</h2><p>Records: {len(rows)}</p><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></body></html>""")
        return document
