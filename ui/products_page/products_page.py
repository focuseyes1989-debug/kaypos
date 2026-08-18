# ui/products_page/products_page.py
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QMessageBox, QLabel, QSizePolicy
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QPixmap, QIcon, QPainter, QColor
from models.database import connect_db
from ui.products_page.product_filters import ProductFilters
from ui.products_page.product_card import ProductCards
from ui.products_page.product_table import ProductTable
from ui.products_page.product_service import ProductService
from ui.products_page.product_form_dialog import ProductFormDialog
from ui.products_page.product_ai_chat_panel import ProductAIChatDialog
from ui.products_page.manage_category_groups_dialog import ManageCategoryGroupsDialog
from ui.print_barcode_dialog import PrintBarcodeDialog
from utils.language import lang
from utils.currency import get_currency_symbol, format_money
from utils.excel_exporter import ExcelExporter
from utils.permissions import PermissionManager, Permission
from loguru import logger
from datetime import datetime
from ui.categories.category_list_dialog import CategoryListDialog

from ui.widgets.modern_button import ModernButton
from ui.widgets.action_toolbar import ActionToolbar
from ui.themes.theme_manager import theme_manager
import os


class ProductsPage(QWidget):
    categories_changed = pyqtSignal()
    
    def __init__(self, user_role=None, user_id=None, parent=None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.user_role = user_role
        self.user_id = user_id
        self.current_filter = None
        self.current_page = 1
        self.items_per_page = 50
        self.selected_product_id = None
        self.selected_product_context = None
        self.service = ProductService(self)
        self.setup_ui()
        self.load_categories()
        self.load_products()
        self.update_cards()

        lang.language_changed.connect(self.retranslateUi)
        theme_manager.theme_changed.connect(self._on_theme_changed)
        self.destroyed.connect(lambda *_: self._disconnect_theme_signal())
        self.retranslateUi()
        self._sync_ai_context()
        
        self.apply_permissions()

    def _on_theme_changed(self, theme_name):
        if hasattr(self, "action_toolbar"):
            self.action_toolbar.update_theme()
        if hasattr(self, "ai_chat_panel"):
            self.ai_chat_panel.update_theme()
        if hasattr(self, "ai_chat_dialog"):
            self.ai_chat_dialog.update_theme()

    def _disconnect_theme_signal(self):
        if hasattr(self, "ai_chat_panel"):
            self.ai_chat_panel.shutdown()
        try:
            theme_manager.theme_changed.disconnect(self._on_theme_changed)
        except Exception:
            pass

    def _refresh_inventory_page_if_loaded(self):
        inventory_page = getattr(self.window(), "inventory_page", None)
        if inventory_page is not None and hasattr(inventory_page, "refresh_all"):
            inventory_page.refresh_all()

    def setup_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setSpacing(10)

        # Cards
        self.cards = ProductCards(self)
        main_layout.addWidget(self.cards)

        # Top bar: filters and buttons
        top_layout = QHBoxLayout()
        top_layout.setSpacing(6)  # âœ… Spacing á€€á€­á€¯ á€œá€»á€¾á€±á€¬á€·á€á€»á€œá€­á€¯á€€á€ºá€•á€«
        
        # âœ… ProductFilters - stretch á€™á€•á€±á€¸á€á€±á€¬á€·á€•á€«
        self.filters = ProductFilters(self)
        self.filters.filter_changed.connect(self.on_filter_changed)
        top_layout.addWidget(self.filters)  # âœ… stretch á€–á€šá€ºá€œá€­á€¯á€€á€ºá€•á€«

        self.action_toolbar = ActionToolbar(self)
        self.btn_add = self.action_toolbar.add_primary(" Add Item", self.open_add_dialog, "add", width=112)
        self.btn_edit = self.action_toolbar.add_primary(" Edit", self.edit_product, "edit", ModernButton.SECONDARY, width=86)
        self.btn_ai_chat = self.action_toolbar.add_primary(
            " AI Assistant", self.toggle_ai_chat, "smart_toy",
            ModernButton.SECONDARY, width=126
        )
        self.btn_ai_chat.setCheckable(False)
        self.btn_ai_chat.setAutoExclusive(False)
        self.action_delete = self.action_toolbar.add_more_action("Delete", self.delete_product, "delete")
        self.action_toolbar.add_separator()
        self.action_manage_cat = self.action_toolbar.add_more_action("Manage Categories", self.open_manage_categories, "category")
        self.action_manage_groups = self.action_toolbar.add_more_action("Manage Groups", self.open_manage_groups, "groups")
        self.action_print_barcode = self.action_toolbar.add_more_action("Print Barcode", self.print_barcode, "barcode")
        self.action_toolbar.add_separator()
        self.action_toolbar.finalize()
        top_layout.addWidget(self.action_toolbar)
        main_layout.addLayout(top_layout)

        # Product table; AI assistant lives in a non-modal floating dialog.
        self.table = ProductTable(self)
        self.table.product_selected.connect(self.on_product_selected)
        self.table.service_selected.connect(self.on_service_selected)
        main_layout.addWidget(self.table, 1)

        self.ai_chat_dialog = ProductAIChatDialog(
            self,
            user_id=self.user_id,
            can_view_sensitive=self._can_view_sensitive_ai_data(),
        )
        self.ai_chat_panel = self.ai_chat_dialog.panel
        self.ai_chat_panel.product_action_requested.connect(self._handle_ai_product_action)
        self.ai_chat_panel.audit_event.connect(self._handle_ai_audit_event)
        self.ai_chat_dialog.visibility_changed.connect(self._on_ai_dialog_visibility_changed)
        self.ai_chat_dialog.hide()

        # Bottom bar: Export Buttons (Excel + CSV)
        bottom_layout = QHBoxLayout()
        bottom_layout.setSpacing(6)  # âœ… Spacing á€€á€­á€¯ á€œá€»á€¾á€±á€¬á€·á€á€»á€œá€­á€¯á€€á€ºá€•á€«
        
        self.total_label = QLabel("Total Products: 0")
        
        self.action_export_list = self.action_toolbar.add_more_action("Export Excel", self.export_products_to_excel, "file_export")
        self.action_export_price = self.action_toolbar.add_more_action("Export Price List", self.export_price_list_to_excel, "attach_money")
        self.action_export_barcode = self.action_toolbar.add_more_action("Export Barcode Data", self.export_barcode_data_to_excel, "barcode")
        self.action_export = self.action_toolbar.add_more_action("Export CSV", self.export_products, "upload_file")
        self.action_import = self.action_toolbar.add_more_action("Import CSV", self.import_products, "download_done")
        
        bottom_layout.addWidget(self.total_label)
        bottom_layout.addStretch()
        main_layout.addLayout(bottom_layout)

        self.setLayout(main_layout)

    def toggle_ai_chat(self):
        """Show or hide the floating Products AI assistant."""
        if self.ai_chat_dialog.isVisible():
            self.hide_ai_chat()
            return
        self._position_ai_dialog()
        self.ai_chat_dialog.show()
        self.ai_chat_dialog.raise_()
        self.ai_chat_dialog.activateWindow()
        self.ai_chat_panel.chat.input_field.setFocus()

    def hide_ai_chat(self):
        self.ai_chat_dialog.hide()

    def _on_ai_dialog_visibility_changed(self, visible):
        self.btn_ai_chat.setText(" Hide AI" if visible else " AI Assistant")

    def _position_ai_dialog(self):
        anchor = self.mapToGlobal(self.rect().bottomRight())
        screen = QApplication.screenAt(anchor) or QApplication.primaryScreen()
        available = screen.availableGeometry() if screen else None
        x = anchor.x() - self.ai_chat_dialog.width() - 18
        y = anchor.y() - self.ai_chat_dialog.height() - 18
        if available:
            x = max(available.left() + 8, min(x, available.right() - self.ai_chat_dialog.width() - 8))
            y = max(available.top() + 8, min(y, available.bottom() - self.ai_chat_dialog.height() - 8))
        self.ai_chat_dialog.move(x, y)

    def _set_button_icon(self, button, icon_name, size=18):
        """Set SVG icon for a button"""
        try:
            icon_paths = [
                f"assets/icons/{icon_name}.svg",
                f"assets/icons/{icon_name}.png",
            ]
            
            for path in icon_paths:
                if os.path.exists(path):
                    pixmap = QPixmap(path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(
                            size, size,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        # Color the icon for the button
                        colored = scaled.copy()
                        painter = QPainter(colored)
                        painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
                        
                        # Check button type for color
                        if hasattr(button, '_style') and button._style == "primary":
                            painter.fillRect(colored.rect(), QColor("white"))
                        else:
                            painter.fillRect(colored.rect(), QColor("#5865f2"))
                        painter.end()
                        
                        icon = QIcon(colored)
                        button.setIcon(icon)
                        button.setIconSize(Qt.QSize(size, size))
                        return
        except Exception as e:
            logger.debug(f"Could not set icon {icon_name}: {e}")

    def apply_permissions(self):
        """Apply permissions to buttons based on user role"""
        if self.user_id:
            if not PermissionManager.user_has_permission(self.user_id, Permission.EDIT_PRODUCT):
                self.btn_edit.setEnabled(False)
                self.btn_edit.setToolTip("You don't have permission to edit products")
            
            if not PermissionManager.user_has_permission(self.user_id, Permission.DELETE_PRODUCT):
                self.action_delete.setEnabled(False)
                self.action_delete.setToolTip("You don't have permission to delete products")
            
            if not PermissionManager.user_has_permission(self.user_id, Permission.ADD_PRODUCT):
                self.btn_add.setEnabled(False)
                self.btn_add.setToolTip("You don't have permission to add products")
            if not PermissionManager.user_has_permission(self.user_id, Permission.VIEW_AI_PAGES):
                self.btn_ai_chat.setEnabled(False)
                self.btn_ai_chat.setToolTip("You don't have permission to use AI features")
                self.ai_chat_dialog.hide()

    def _can_view_sensitive_ai_data(self):
        if not self.user_id:
            return True
        return (
            PermissionManager.user_has_permission(self.user_id, Permission.EDIT_PRODUCT)
            or PermissionManager.user_has_permission(self.user_id, Permission.VIEW_REPORTS)
        )

    def on_card_filter(self, key):
        self.filters.reset()
        self.selected_product_id = None
        self.selected_product_context = None
        if key == "total_cost":
            self.current_filter = None
        else:
            self.current_filter = key
        self.current_page = 1
        self.apply_filter()
        self._sync_ai_context()

    def on_filter_changed(self):
        self.current_filter = None
        self.selected_product_id = None
        self.selected_product_context = None
        self.current_page = 1
        self.apply_filter()
        self._sync_ai_context()

    def on_barcode_scanned(self, keyword):
        main_window = self.window()
        if hasattr(main_window, 'sales_page'):
            main_window.switch_to_page(5)
            main_window.sales_page.product_grid.barcode_scanned.emit(keyword)

    def apply_filter(self):
        if self.current_filter == "out_stock":
            rows, total = self.service.filter_by_type(
                'out_of_stock', self.current_page, self.items_per_page
            )
        elif self.current_filter == "low_stock":
            rows, total = self.service.filter_by_type(
                'low_stock', self.current_page, self.items_per_page
            )
        elif self.current_filter == "expiring_soon":
            rows, total = self.service.filter_by_type(
                'expiring_soon', self.current_page, self.items_per_page
            )
        elif self.current_filter == "expired":
            rows, total = self.service.filter_by_type(
                'expired', self.current_page, self.items_per_page
            )
        else:
            rows, total = self.service.load_products(
                self.current_page, self.items_per_page,
                self.filters.get_search_text(),
                self.filters.get_category()
            )
        self.table.set_pagination_total(total)
        self.table.populate_table(rows)
        self.update_total_label(total)

    def load_products(self):
        self.apply_filter()

    def load_categories(self):
        self.filters.load_categories()

    def update_cards(self):
        self.cards.update_cards()

    def update_total_label(self, count):
        if lang.get_current() == "my":
            self.total_label.setText(f"á€…á€¯á€…á€¯á€•á€±á€«á€„á€ºá€¸á€•á€…á€¹á€…á€Šá€ºá€¸: {count}")
        else:
            self.total_label.setText(f"Total Products: {count}")

    def on_product_selected(self, prod_id, name, price, stock):
        self.selected_product_id = prod_id
        self.selected_product_context = {
            "id": prod_id, "name": name, "price": price, "stock": stock
        }
        self._sync_ai_context()

    def on_service_selected(self, prod_id, name, price):
        self.selected_product_id = prod_id
        self.selected_product_context = {
            "id": prod_id, "name": name, "price": price, "stock": None,
            "is_service": True,
        }
        self._sync_ai_context()

    def _sync_ai_context(self):
        """Keep the assistant aware of the current Products page state."""
        if not hasattr(self, "ai_chat_panel") or not hasattr(self, "filters"):
            return
        self.ai_chat_panel.set_product_context(
            product=self.selected_product_context,
            search_text=self.filters.get_search_text(),
            category=self.filters.get_category(),
            active_filter=self.current_filter,
        )

    def _handle_ai_product_action(self, action, product):
        """Apply a safe AI result-card action to the Products page."""
        if action == "prefill_add":
            self._ai_prefill_add_product(product)
            return
        if action == "bulk_category":
            self._ai_bulk_assign_category(product.get("category", ""))
            return
        if action == "export_filtered":
            if self._confirm_ai_action(
                "Export filtered products",
                f"Export the {len(self.table.get_current_rows())} products currently shown in the table.",
                "This opens a file chooser and does not change product data.",
            ):
                self.export_products()
                self._log_ai_action("AI Export Products", "Exported current filtered product rows")
            return
        if action == "filter":
            term = product.get("search_term") or product.get("category") or product.get("name") or ""
            self.filters.search_widget.set_text(str(term))
            return

        product_id = product.get("id")
        if not product_id:
            return
        if not self._select_ai_product(product_id, product.get("name", "")):
            QMessageBox.warning(self, "Product Not Found", "The selected AI product is no longer available.")
            return
        if action == "apply_reorder":
            self._ai_apply_reorder_alert(product)
        elif action == "edit":
            self.edit_product()
        elif action == "barcode":
            self.print_barcode()
        elif action == "view":
            self.table.on_cell_double_clicked(self.table.table.currentRow(), 2)

    def _select_ai_product(self, product_id, product_name=""):
        """Select an AI result in the table, filtering first when it is off-page."""
        def find_row():
            for row in range(self.table.table.rowCount()):
                item = self.table.table.item(row, 0)
                if item and str(item.text()) == str(product_id):
                    return row
            return -1

        row = find_row()
        if row < 0 and product_name:
            self.filters.search_widget.set_text(str(product_name))
            row = find_row()
        if row < 0:
            return False

        self.table.table.selectRow(row)
        name_item = self.table.table.item(row, 2)
        stock_item = self.table.table.item(row, 7)
        self.table.table.scrollToItem(name_item)
        self.selected_product_id = int(product_id)
        self.selected_product_context = {
            "id": int(product_id),
            "name": name_item.text() if name_item else str(product_name),
            "price": 0,
            "stock": int(stock_item.text()) if stock_item and stock_item.text().isdigit() else None,
        }
        self._sync_ai_context()
        return True

    def _ai_prefill_add_product(self, suggestion):
        if not self._ai_has_permission(Permission.ADD_PRODUCT, "add products"):
            return
        name = str(suggestion.get("name") or "").strip()
        if not name or len(name) > 200:
            QMessageBox.warning(self, "Invalid Product Name", "Product names must be between 1 and 200 characters.")
            return
        dialog = ProductFormDialog(parent=self)
        dialog.name_input.setText(name)
        category = str(suggestion.get("category") or "").strip()
        category_index = dialog.category_combo.findText(category)
        if category_index >= 0:
            dialog.category_combo.setCurrentIndex(category_index)
        self.ai_chat_panel.chat._add_bot_message(
            f"Prepared a new-product form for “{name}”. Review every field and click Save Product to confirm."
        )
        if dialog.exec():
            self.load_categories()
            self.load_products()
            self.update_cards()
            self._refresh_inventory_page_if_loaded()
            self.categories_changed.emit()
            self.refresh_current_stock_categories()
            self._log_ai_action("AI Add Product", f"Created product from AI-prefilled form: {name}")

    def _ai_apply_reorder_alert(self, suggestion):
        if not self._ai_has_permission(Permission.EDIT_PRODUCT, "edit reorder alerts"):
            return
        product_id = suggestion.get("id")
        proposed = int(suggestion.get("recommended_low_stock") or 0)
        if not product_id or proposed < 1:
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name, COALESCE(low_stock, 0) FROM products WHERE id=?", (product_id,))
        row = cursor.fetchone()
        conn.close()
        if not row:
            QMessageBox.warning(self, "Product Not Found", "This product no longer exists.")
            return
        name, current = row
        if not self._confirm_ai_action(
            "Apply reorder alert",
            f"Product: {name}\nCurrent low-stock alert: {current}\nSuggested alert: {proposed}",
            "Only the low-stock alert level will change. Stock quantity will not be changed.",
        ):
            return
        conn = connect_db()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "UPDATE products SET low_stock=?, last_updated=CURRENT_TIMESTAMP WHERE id=?",
                (proposed, product_id),
            )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            logger.exception(f"AI reorder alert update failed: {exc}")
            QMessageBox.critical(self, "Update Failed", "The reorder alert was not changed.")
            return
        finally:
            conn.close()
        self.load_products()
        self.update_cards()
        self._refresh_inventory_page_if_loaded()
        self._log_ai_action(
            "AI Reorder Alert", f"Product {name} (ID {product_id}): low_stock {current} -> {proposed}"
        )
        self.ai_chat_panel.chat._add_bot_message(
            f"Updated “{name}” low-stock alert from {current} to {proposed}. Stock quantity was not changed."
        )

    def _ai_bulk_assign_category(self, category):
        if not self._ai_has_permission(Permission.EDIT_PRODUCT, "bulk-edit product categories"):
            return
        category = str(category or "").strip()
        if len(category) > 100:
            QMessageBox.warning(self, "Invalid Category", "Category names cannot exceed 100 characters.")
            return
        rows = self.table.get_current_rows()
        product_ids = [int(row[0]) for row in rows if row]
        if not category or not product_ids:
            QMessageBox.warning(self, "Nothing to Update", "Choose a category and filter products first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM categories WHERE LOWER(name)=LOWER(?)", (category,))
        category_row = cursor.fetchone()
        conn.close()
        if not category_row:
            QMessageBox.warning(self, "Unknown Category", f"Category “{category}” does not exist.")
            return
        if not self._confirm_ai_action(
            "Bulk category assignment",
            f"Assign {len(product_ids)} currently visible product(s) to category “{category}”.",
            "This changes only the current table page. You can edit products individually afterward.",
        ):
            return
        placeholders = ",".join("?" for _ in product_ids)
        conn = connect_db()
        try:
            cursor = conn.cursor()
            cursor.execute(
                f"UPDATE products SET category=?, category_id=?, last_updated=CURRENT_TIMESTAMP "
                f"WHERE id IN ({placeholders})",
                [category, category_row[0], *product_ids],
            )
            changed = cursor.rowcount
            conn.commit()
        except Exception as exc:
            conn.rollback()
            logger.exception(f"AI bulk category update failed: {exc}")
            QMessageBox.critical(self, "Update Failed", "No category changes were committed.")
            return
        finally:
            conn.close()
        self.load_categories()
        self.load_products()
        self.update_cards()
        self.categories_changed.emit()
        self._refresh_inventory_page_if_loaded()
        self._log_ai_action(
            "AI Bulk Category", f"Assigned {changed} products to category {category}; IDs={product_ids}"
        )
        self.ai_chat_panel.chat._add_bot_message(
            f"Assigned {changed} visible product(s) to category “{category}”."
        )

    def _ai_has_permission(self, permission, action_label):
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, permission):
            QMessageBox.warning(self, "Access Denied", f"You do not have permission to {action_label}.")
            return False
        return True

    def _confirm_ai_action(self, title, preview, impact):
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Question)
        box.setWindowTitle("Confirm AI Action")
        box.setText(title)
        box.setInformativeText(preview)
        box.setDetailedText(f"Impact\n{impact}\n\nNothing will change unless you choose Apply.")
        apply_button = box.addButton("Apply", QMessageBox.ButtonRole.AcceptRole)
        box.addButton("Cancel", QMessageBox.ButtonRole.RejectRole)
        box.exec()
        return box.clickedButton() is apply_button

    def _log_ai_action(self, action, details):
        main_window = self.window()
        current_user = getattr(main_window, "current_user", None) or {}
        user_id = current_user.get("id") or self.user_id
        username = current_user.get("username") or self.user_role or "Unknown"
        if user_id:
            self.service.log_activity(user_id, username, action, details)
        logger.info(f"{action}: {details}")

    def _handle_ai_audit_event(self, event_type, payload):
        """Persist concise product-AI audit events without storing business result data."""
        safe_payload = {
            key: value for key, value in dict(payload or {}).items()
            if key in {"command", "focus", "success", "elapsed_ms", "feature", "rating"}
        }
        self._log_ai_action(f"Product AI {event_type}", str(safe_payload))

    def open_add_dialog(self):
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.ADD_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to add products.")
            return
            
        dialog = ProductFormDialog()
        if dialog.exec():
            product_name = dialog.name_input.text()
            self.load_categories()
            self.load_products()
            self.update_cards()
            main_window = self.window()
            self._refresh_inventory_page_if_loaded()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Add Product", f"Product: {product_name}")
            self.categories_changed.emit()
            self.refresh_current_stock_categories()
            logger.info("New product added")

    def edit_product(self):
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.EDIT_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to edit products.")
            return
            
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        product_name = row[0] if row else "Unknown"
        conn.close()
        dialog = ProductFormDialog(prod_id)
        if dialog.exec():
            self.load_categories()
            self.load_products()
            self.update_cards()
            main_window = self.window()
            self._refresh_inventory_page_if_loaded()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Edit Product", f"Product: {product_name}")
            self.refresh_current_stock_categories()
            logger.info(f"Product edited: ID {prod_id}")

    def delete_product(self):
        if self.user_id and not PermissionManager.user_has_permission(self.user_id, Permission.DELETE_PRODUCT):
            QMessageBox.warning(self, "Access Denied", "You don't have permission to delete products.")
            return
            
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        product_name = row[0] if row else "Unknown"
        conn.close()
        reply = QMessageBox.question(self, "Confirm Delete", "Delete this product permanently?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM products WHERE id=?", (prod_id,))
            conn.commit()
            conn.close()
            QMessageBox.information(self, "Deleted", "Product Deleted")
            main_window = self.window()
            if hasattr(main_window, 'current_user'):
                self.service.log_activity(main_window.current_user["id"], main_window.current_user["username"],
                                         "Delete Product", f"Product: {product_name}")
            logger.info(f"Product deleted: ID {prod_id}")
            self.load_categories()
            self.load_products()
            self.update_cards()
            self._refresh_inventory_page_if_loaded()
            self.refresh_current_stock_categories()

    def print_barcode(self):
        prod_id = self.table.get_selected_product_id()
        if not prod_id:
            QMessageBox.warning(self, "No Selection", "Please select a product first.")
            return
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("SELECT name, barcode FROM products WHERE id=?", (prod_id,))
        row = cursor.fetchone()
        conn.close()
        if row:
            name, barcode = row
            if not barcode:
                QMessageBox.warning(self, "No Barcode", "This product does not have a barcode number.\nPlease edit the product and add a barcode.")
                return
            dialog = PrintBarcodeDialog(prod_id, name, barcode, self)
            dialog.exec()
        else:
            QMessageBox.warning(self, "Error", "Product not found.")

    def open_manage_categories(self):
        """Open enhanced category management dialog"""
        dialog = CategoryListDialog(self)
        dialog.categories_changed.connect(self.on_categories_changed)
        dialog.exec()
        self.load_categories()
        self.load_products()
        self.update_cards()
        self.categories_changed.emit()

    def open_manage_groups(self):
        dialog = ManageCategoryGroupsDialog(self)
        dialog.groups_changed.connect(self.on_groups_changed)
        dialog.exec()
        self.load_categories()
        self.load_products()
        self.update_cards()
        self.categories_changed.emit()

    def on_groups_changed(self):
        self.load_categories()
        self.load_products()
        self.update_cards()
        self.categories_changed.emit()

    def on_categories_changed(self):
        self.load_categories()
        self.load_products()
        self.update_cards()
        self.categories_changed.emit()
        self.refresh_current_stock_categories()

    def refresh_current_stock_categories(self):
        main_window = self.window()
        if hasattr(main_window, 'inventory_page'):
            inventory = main_window.inventory_page
            if hasattr(inventory, 'current_stock_tab'):
                inventory.current_stock_tab.load_categories()
                inventory.current_stock_tab.refresh()
                logger.info("Current stock tab categories refreshed")

    def export_products(self):
        rows = self.table.get_current_rows()
        self.service.export_products(rows, self)

    def import_products(self):
        self.service.import_products(self, self.refresh_after_import)

    def refresh_after_import(self):
        self.load_categories()
        self.load_products()
        self.update_cards()
        self._refresh_inventory_page_if_loaded()
        self.refresh_current_stock_categories()

    def get_all_products_data(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, sku, name, category, barcode, price, cost, stock, 
                   low_stock, sold_by, expire_date
            FROM products
            ORDER BY name
        """)
        rows = cursor.fetchall()
        conn.close()
        return rows

    def export_products_to_excel(self):
        lang_code = lang.get_current()
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"product_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Product List" if lang_code != "my" else "á€•á€…á€¹á€…á€Šá€ºá€¸á€…á€¬á€›á€„á€ºá€¸ á€‘á€¯á€á€ºá€›á€”á€º"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            rows = self.get_all_products_data()
            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            ws.merge_cells('A1:K1')
            ws['A1'] = "PRODUCT LIST REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            if lang_code == "my":
                headers = ["SKU", "á€•á€…á€¹á€…á€Šá€ºá€¸á€¡á€™á€Šá€º", "á€¡á€™á€»á€­á€¯á€¸á€¡á€…á€¬á€¸", "á€¡á€¯á€•á€ºá€…á€¯", "á€˜á€¬á€¸á€€á€¯á€’á€º", 
                          "á€›á€±á€¬á€„á€ºá€¸á€ˆá€±á€¸", "á€€á€¯á€”á€ºá€€á€»á€…á€›á€­á€á€º", "á€€á€»á€”á€º", 
                          "á€žá€á€­á€•á€±á€¸á€•á€™á€¬á€", "á€›á€±á€¬á€„á€ºá€¸á€•á€¯á€¶á€…á€¶", "á€žá€€á€ºá€á€™á€ºá€¸á€€á€¯á€”á€ºá€›á€€á€º"]
            else:
                headers = ["SKU", "Product Name", "Category", "Category Group", "Barcode", 
                          "Selling Price", "Cost", "Stock", 
                          "Low Stock Alert", "Sold By", "Expiry Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            total_stock = 0
            total_stock_value = 0
            total_cost_value = 0
            for row_idx, row_data in enumerate(rows, start=6):
                pid, sku, name, category, barcode, price, cost, stock, low_stock, sold_by, expire_date = row_data
                conn = connect_db()
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT cg.name 
                    FROM categories c
                    LEFT JOIN category_groups cg ON c.group_id = cg.id
                    WHERE c.name = ?
                """, (category,))
                group_row = cursor.fetchone()
                conn.close()
                category_group = group_row[0] if group_row and group_row[0] else ""
                price_val = float(price) if price else 0
                cost_val = float(cost) if cost else 0
                stock_val = int(stock) if stock else 0
                stock_value = price_val * stock_val
                cost_value = cost_val * stock_val
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=category or "")
                ws.cell(row=row_idx, column=4, value=category_group or "")
                ws.cell(row=row_idx, column=5, value=barcode or "")
                ws.cell(row=row_idx, column=6, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=7, value=format_money(cost_val, symbol))
                ws.cell(row=row_idx, column=8, value=stock_val)
                ws.cell(row=row_idx, column=9, value=low_stock or 0)
                ws.cell(row=row_idx, column=10, value=sold_by or "Each")
                ws.cell(row=row_idx, column=11, value=expire_date or "")
                total_stock += stock_val
                total_stock_value += stock_value
                total_cost_value += cost_value
            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=6, value="TOTAL STOCK VALUE:").font = Font(bold=True)
            ws.cell(row=summary_row, column=7, value=format_money(total_stock_value, symbol))
            ws.cell(row=summary_row + 1, column=6, value="TOTAL COST VALUE:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=7, value=format_money(total_cost_value, symbol))
            ws.cell(row=summary_row + 2, column=6, value="POTENTIAL PROFIT:").font = Font(bold=True)
            ws.cell(row=summary_row + 2, column=7, value=format_money(total_stock_value - total_cost_value, symbol))
            ws.cell(row=summary_row + 3, column=6, value="TOTAL QUANTITY:").font = Font(bold=True)
            ws.cell(row=summary_row + 3, column=7, value=total_stock)
            for col in range(1, 12):
                col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                ws.column_dimensions[col_letter].width = 18
            ws.column_dimensions['B'].width = 30
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def export_price_list_to_excel(self):
        lang_code = lang.get_current()
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"price_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Price List" if lang_code != "my" else "á€…á€»á€±á€¸á€”á€¾á€¯á€”á€ºá€¸á€…á€¬á€›á€„á€ºá€¸ á€‘á€¯á€á€ºá€›á€”á€º"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sku, name, category, price, sold_by, barcode
                FROM products
                ORDER BY name
            """)
            rows = cursor.fetchall()
            conn.close()
            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Price List"
            ws.merge_cells('A1:F1')
            ws['A1'] = "PRICE LIST REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            if lang_code == "my":
                headers = ["SKU", "á€•á€…á€¹á€…á€Šá€ºá€¸á€¡á€™á€Šá€º", "á€¡á€™á€»á€­á€¯á€¸á€¡á€…á€¬á€¸", "á€…á€»á€±á€¸á€”á€¾á€¯á€”á€ºá€¸", "á€›á€±á€¬á€„á€ºá€¸á€•á€¯á€¶á€…á€¶", "á€˜á€¬á€¸á€€á€¯á€’á€º"]
            else:
                headers = ["SKU", "Product Name", "Category", "Price", "Sold By", "Barcode"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row_data in enumerate(rows, start=6):
                sku, name, category, price, sold_by, barcode = row_data
                price_val = float(price) if price else 0
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=category or "")
                ws.cell(row=row_idx, column=4, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=5, value=sold_by or "Each")
                ws.cell(row=row_idx, column=6, value=barcode or "")
            for col in range(1, 7):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 30
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def export_barcode_data_to_excel(self):
        lang_code = lang.get_current()
        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"barcode_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Export Barcode Data" if lang_code != "my" else "á€˜á€¬á€¸á€€á€¯á€’á€ºá€’á€±á€á€¬ á€‘á€¯á€á€ºá€›á€”á€º"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sku, name, barcode, price, category
                FROM products
                WHERE barcode IS NOT NULL AND barcode != ''
                ORDER BY name
            """)
            rows = cursor.fetchall()
            conn.close()
            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Barcode Data"
            ws.merge_cells('A1:E1')
            ws['A1'] = "BARCODE DATA REPORT"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            ws['A3'] = f"Total Products with Barcode: {len(rows)}"
            ws['A3'].font = Font(size=10, color="7f8c8d")
            ws['A4'] = "NOTE: This file can be used for barcode label printing"
            ws['A4'].font = Font(size=9, italic=True, color="7f8c8d")
            if lang_code == "my":
                headers = ["SKU", "á€•á€…á€¹á€…á€Šá€ºá€¸á€¡á€™á€Šá€º", "á€˜á€¬á€¸á€€á€¯á€’á€º", "á€…á€»á€±á€¸á€”á€¾á€¯á€”á€ºá€¸", "á€¡á€™á€»á€­á€¯á€¸á€¡á€…á€¬á€¸"]
            else:
                headers = ["SKU", "Product Name", "Barcode", "Price", "Category"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row_data in enumerate(rows, start=7):
                sku, name, barcode, price, category = row_data
                price_val = float(price) if price else 0
                ws.cell(row=row_idx, column=1, value=sku or "")
                ws.cell(row=row_idx, column=2, value=name or "")
                ws.cell(row=row_idx, column=3, value=barcode or "")
                ws.cell(row=row_idx, column=4, value=format_money(price_val, symbol))
                ws.cell(row=row_idx, column=5, value=category or "")
            tips_row = len(rows) + 8
            ws.cell(row=tips_row, column=1, value="BARCODE FORMAT TIPS:").font = Font(bold=True, size=11)
            ws.cell(row=tips_row + 1, column=1, value="- Code128: Supports alphanumeric, variable length")
            ws.cell(row=tips_row + 2, column=1, value="- EAN13: 13 digits, for retail products")
            ws.cell(row=tips_row + 3, column=1, value="- UPC-A: 12 digits, for North American products")
            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 18
            ws.column_dimensions['B'].width = 30
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)

    def retranslateUi(self):
        self.filters.retranslateUi()
        self.cards.retranslateUi()
        self.table.retranslateUi()

        if lang.get_current() == "my":
            self.btn_add.setText(" Add Item")
            self.btn_edit.setText(" Edit")
        else:
            self.btn_add.setText(" Add Item")
            self.btn_edit.setText(" Edit")

        self.action_delete.setText("Delete")
        self.action_manage_cat.setText("Manage Categories")
        self.action_manage_groups.setText("Manage Groups")
        self.action_print_barcode.setText("Print Barcode")
        self.action_export_list.setText("Export Excel")
        self.action_export_price.setText("Export Price List")
        self.action_export_barcode.setText("Export Barcode Data")
        self.action_export.setText("Export CSV")
        self.action_import.setText("Import CSV")
    def showEvent(self, event):
        self.load_products()
        self.update_cards()
        super().showEvent(event)

