# ui/products_page/category_cost_dialog.py
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QPushButton, QLabel, QMessageBox, QFrame, QComboBox
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QColor, QPixmap, QIcon
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from utils.language import lang
from utils.excel_exporter import ExcelExporter
from utils.translations import tr
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.search_widget import ModernSearchWidget
from ui.widgets.modern_button import ModernButton
from ui.widgets.summary_card_widget import SummaryCardWidget
from ui.themes.theme_manager import get_theme_colors, is_dark_theme, theme_manager
from datetime import datetime
from loguru import logger
import os


class CategoryCostDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Category Cost Breakdown")
        self.setMinimumSize(900, 550)
        self.setModal(True)
        
        # Store data for export
        self.category_data = []
        self.filtered_data = []
        self.total_cost_all = 0
        
        # Pagination variables
        self.current_page = 1
        self.page_size = 25
        
        # Theme state
        self._is_dark = is_dark_theme()
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # Top bar with export button
        top_layout = QHBoxLayout()
        
        # Export button - with SVG icon
        self.btn_export = ModernButton(" Export Excel")
        self._set_button_icon(self.btn_export, "file_export")
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addStretch()
        top_layout.addWidget(self.btn_export)
        
        layout.addLayout(top_layout)
        
        # ✅ Summary Cards with SVG icons
        card_layout = QHBoxLayout()
        card_layout.setSpacing(15)
        
        # Card 1: Total Categories - groups.svg
        self.card_categories = SummaryCardWidget(
            title="Total Categories",
            value="0",
            icon="groups",
            color="#3498db",
            icon_is_svg=True
        )
        card_layout.addWidget(self.card_categories, 1)
        
        # Card 2: Total Products - package.svg
        self.card_products = SummaryCardWidget(
            title="Total Products",
            value="0",
            icon="package",
            color="#2ecc71",
            icon_is_svg=True
        )
        card_layout.addWidget(self.card_products, 1)
        
        # Card 3: Total Cost - attach_money.svg
        self.card_cost = SummaryCardWidget(
            title="Total Cost",
            value="0",
            icon="attach_money",
            color="#f39c12",
            icon_is_svg=True
        )
        card_layout.addWidget(self.card_cost, 1)
        
        # Card 4: Total Stock - inventory.svg
        self.card_stock = SummaryCardWidget(
            title="Total Stock",
            value="0",
            icon="inventory",
            color="#9b59b6",
            icon_is_svg=True
        )
        card_layout.addWidget(self.card_stock, 1)
        
        layout.addLayout(card_layout)
        
        # Filter section - ModernSearchWidget
        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)
        
        # ModernSearchWidget with search.svg
        self.search_widget = ModernSearchWidget("Search category...")
        self.search_widget.search_changed.connect(self.on_filter_changed)
        filter_layout.addWidget(self.search_widget, 2)
        
        # Min percentage filter
        self.min_percentage_filter = QComboBox()
        self.min_percentage_filter.addItems(["All", "> 10%", "> 25%", "> 50%"])
        self.min_percentage_filter.currentTextChanged.connect(self.on_filter_changed)
        self.min_percentage_label = QLabel("Min %:")
        filter_layout.addWidget(self.min_percentage_label)
        filter_layout.addWidget(self.min_percentage_filter, 1)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Category", "Product Count", "Total Cost", 
            "Total Stock", "Stock Value (Cost)", "Percentage"
        ])
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        
        # Set column stretch modes
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        
        # Set uniform row height
        self.table.verticalHeader().setDefaultSectionSize(30)
        self.table.verticalHeader().setMinimumSectionSize(25)
        self.table.verticalHeader().setMaximumSectionSize(40)
        self.table.setWordWrap(False)
        
        # Apply scrollbar style
        self._apply_table_scrollbar_style()
        
        layout.addWidget(self.table)
        
        # Pagination
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        self.setLayout(layout)
        
        # Load data
        self.load_data()
        
        # Language support
        lang.language_changed.connect(self.retranslateUi)
        self.retranslateUi()
        
        # Theme support
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        # Set window flags
        self.setWindowFlags(
            Qt.WindowType.Window |
            Qt.WindowType.WindowCloseButtonHint |
            Qt.WindowType.WindowMinimizeButtonHint |
            Qt.WindowType.WindowMaximizeButtonHint
        )
    
    def _set_button_icon(self, button, icon_name, size=18):
        """Set SVG icon for a button"""
        try:
            icon_paths = [
                f"assets/icons/{icon_name}.svg",
                f"assets/icons/{icon_name}.png",
            ]
            
            for path in icon_paths:
                if os.path.exists(path):
                    pixmap = QPixmap(path)
                    if not pixmap.isNull():
                        scaled = pixmap.scaled(
                            size, size,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation
                        )
                        # Color the icon white
                        from PyQt6.QtGui import QPainter, QColor
                        colored = scaled.copy()
                        painter = QPainter(colored)
                        painter.setCompositionMode(QPainter.CompositionMode.CompositionMode_SourceIn)
                        painter.fillRect(colored.rect(), QColor("white"))
                        painter.end()
                        
                        icon = QIcon(colored)
                        button.setIcon(icon)
                        button.setIconSize(Qt.QSize(size, size))
                        return
        except Exception as e:
            logger.debug(f"Could not set icon {icon_name}: {e}")
    
    def _apply_table_scrollbar_style(self):
        """Apply smaller scrollbar style to table"""
        is_dark = is_dark_theme()
        
        if is_dark:
            scrollbar_style = """
                QTableWidget {
                    background-color: #2f3136;
                    alternate-background-color: #36393f;
                    selection-background-color: #40444b;
                    selection-color: #dcddde;
                    gridline-color: transparent;
                    border: 1px solid #40444b;
                    border-radius: 6px;
                    color: #dcddde;
                }
                QHeaderView::section {
                    background-color: #202225;
                    padding: 8px;
                    border: none;
                    border-bottom: 2px solid #40444b;
                    font-weight: 600;
                    color: #b9bbbe;
                }
                QTableWidget::item {
                    padding: 6px 12px;
                    border: none;
                    border-bottom: 1px solid #40444b;
                    color: #dcddde;
                }
                QTableWidget::item:selected {
                    background-color: #40444b;
                    color: #dcddde;
                }
                QTableWidget::item:hover {
                    background-color: #40444b;
                }
                QScrollBar:vertical {
                    background: #2f3136;
                    width: 6px;
                    border-radius: 3px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background: #40444b;
                    border-radius: 3px;
                    min-height: 16px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #5865f2;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                    border: none;
                    background: transparent;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: transparent;
                }
                QScrollBar:horizontal {
                    background: #2f3136;
                    height: 6px;
                    border-radius: 3px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background: #40444b;
                    border-radius: 3px;
                    min-width: 16px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #5865f2;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                    border: none;
                    background: transparent;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: transparent;
                }
            """
        else:
            scrollbar_style = """
                QTableWidget {
                    background-color: #ffffff;
                    alternate-background-color: #f8f9fa;
                    selection-background-color: #e9ecef;
                    selection-color: #212529;
                    gridline-color: transparent;
                    border: 1px solid #dee2e6;
                    border-radius: 6px;
                    color: #212529;
                }
                QHeaderView::section {
                    background-color: #f1f3f5;
                    padding: 8px;
                    border: none;
                    border-bottom: 2px solid #dee2e6;
                    font-weight: 600;
                    color: #495057;
                }
                QTableWidget::item {
                    padding: 6px 12px;
                    border: none;
                    border-bottom: 1px solid #dee2e6;
                    color: #212529;
                }
                QTableWidget::item:selected {
                    background-color: #e9ecef;
                    color: #212529;
                }
                QTableWidget::item:hover {
                    background-color: #f1f3f5;
                }
                QScrollBar:vertical {
                    background: #f8f9fa;
                    width: 6px;
                    border-radius: 3px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background: #ced4da;
                    border-radius: 3px;
                    min-height: 16px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #5865f2;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                    border: none;
                    background: transparent;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: transparent;
                }
                QScrollBar:horizontal {
                    background: #f8f9fa;
                    height: 6px;
                    border-radius: 3px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background: #ced4da;
                    border-radius: 3px;
                    min-width: 16px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #5865f2;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                    border: none;
                    background: transparent;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: transparent;
                }
            """
        
        self.table.setStyleSheet(scrollbar_style)
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_table_scrollbar_style()
        # Refresh table
        if self.filtered_data:
            self.apply_filters_and_pagination()
    
    def retranslateUi(self):
        """Update UI text when language changes"""
        self.setWindowTitle(tr("category_cost_breakdown"))
        self.btn_export.setText(" " + tr("export_excel"))
        
        lang_code = lang.get_current()
        if lang_code == "my":
            self.search_widget.retranslateUi("my")
        else:
            self.search_widget.retranslateUi("en")
        
        self.min_percentage_label.setText(tr("min_percent"))
        self.min_percentage_filter.setItemText(0, tr("all"))
        self.min_percentage_filter.setItemText(1, "> 10%")
        self.min_percentage_filter.setItemText(2, "> 25%")
        self.min_percentage_filter.setItemText(3, "> 50%")
        
        # Update card titles
        self.card_categories.set_title(tr("total_categories"))
        self.card_products.set_title(tr("total_products_card"))
        self.card_cost.set_title(tr("total_cost"))
        self.card_stock.set_title(tr("total_stock"))
        
        self.table.setHorizontalHeaderLabels([
            tr("category"), tr("product_count"), tr("total_cost"),
            tr("total_stock"), tr("stock_value_cost"), tr("percentage_header")
        ])
        
        # Update card values
        self.update_summary_cards()
    
    def on_filter_changed(self, text=None):
        """Handle filter changes"""
        self.current_page = 1
        self.apply_filters_and_pagination()
    
    def on_page_changed(self, page: int, page_size: int):
        """Handle page changes from pagination widget"""
        self.current_page = page
        self.page_size = page_size
        self.apply_filters_and_pagination()
    
    def load_data(self):
        """Load category cost data from database"""
        conn = connect_db()
        cursor = conn.cursor()
        
        # Get total cost for all products (for percentage calculation)
        cursor.execute("""
            SELECT SUM(COALESCE(cost, 0) * COALESCE(stock, 0))
            FROM products
            WHERE (sold_by IS NULL OR sold_by != 'Service')
              AND COALESCE(stock, 0) > 0
        """)
        self.total_cost_all = cursor.fetchone()[0] or 0
        
        # Get category breakdown
        cursor.execute("""
            SELECT 
                COALESCE(category, ?) as category,
                COUNT(*) as product_count,
                SUM(COALESCE(cost, 0) * COALESCE(stock, 0)) as total_cost,
                SUM(COALESCE(stock, 0)) as total_stock,
                SUM(COALESCE(cost, 0) * COALESCE(stock, 0)) as stock_value
            FROM products
            WHERE (sold_by IS NULL OR sold_by != 'Service')
              AND COALESCE(stock, 0) > 0
            GROUP BY category
            ORDER BY total_cost DESC
        """, (tr("uncategorized"),))
        rows = cursor.fetchall()
        conn.close()
        
        self.category_data = rows
        self.apply_filters_and_pagination()
        self.update_summary_cards()
    
    def apply_filters_and_pagination(self):
        """Apply search filter, percentage filter and pagination"""
        search_text = self.search_widget.get_text().lower().strip()
        min_percentage = self.min_percentage_filter.currentText()
        
        # Parse min percentage value
        min_percent_value = 0
        if min_percentage == "> 10%":
            min_percent_value = 10
        elif min_percentage == "> 25%":
            min_percent_value = 25
        elif min_percentage == "> 50%":
            min_percent_value = 50
        
        # Filter rows
        filtered_rows = []
        for row in self.category_data:
            category, product_count, total_cost, total_stock, stock_value = row
            
            # Search filter
            if search_text:
                if search_text not in category.lower():
                    continue
            
            # Percentage filter
            if min_percent_value > 0 and self.total_cost_all > 0:
                percentage = (total_cost / self.total_cost_all) * 100
                if percentage <= min_percent_value:
                    continue
            
            filtered_rows.append(row)
        
        self.filtered_data = filtered_rows
        
        # Update total items
        total_items = len(filtered_rows)
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        # Calculate pagination
        start_idx = (self.current_page - 1) * self.page_size
        end_idx = min(start_idx + self.page_size, total_items)
        page_rows = filtered_rows[start_idx:end_idx]
        
        # Populate table
        self.populate_table(page_rows)
    
    def populate_table(self, rows):
        """Populate table with category data"""
        self.table.setRowCount(len(rows))
        symbol = get_currency_symbol()
        
        grand_total = self.total_cost_all
        
        for row_idx, row in enumerate(rows):
            category, product_count, total_cost, total_stock, stock_value = row
            
            # Category
            category_item = QTableWidgetItem(category)
            category_item.setFlags(category_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 0, category_item)
            
            # Product Count
            count_item = QTableWidgetItem(str(product_count))
            count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            count_item.setFlags(count_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 1, count_item)
            
            # Total Cost
            cost_item = QTableWidgetItem(format_money(total_cost, symbol))
            cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            cost_item.setFlags(cost_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 2, cost_item)
            
            # Total Stock
            stock_item = QTableWidgetItem(str(total_stock))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            stock_item.setFlags(stock_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 3, stock_item)
            
            # Stock Value (Cost)
            value_item = QTableWidgetItem(format_money(stock_value, symbol))
            value_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            value_item.setFlags(value_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_idx, 4, value_item)
            
            # Percentage
            if grand_total > 0:
                percentage = (total_cost / grand_total) * 100
                percentage_text = f"{percentage:.1f}%"
            else:
                percentage_text = "0%"
            
            percentage_item = QTableWidgetItem(percentage_text)
            percentage_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            percentage_item.setFlags(percentage_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            
            # Color coding
            if percentage > 50:
                percentage_item.setForeground(QColor(200, 50, 50))
            elif percentage > 25:
                percentage_item.setForeground(QColor(200, 140, 30))
            elif percentage > 10:
                percentage_item.setForeground(QColor(40, 100, 200))
            else:
                percentage_item.setForeground(QColor(40, 160, 60))
            
            self.table.setItem(row_idx, 5, percentage_item)
        
        # Set uniform row height for all rows
        for row in range(self.table.rowCount()):
            self.table.setRowHeight(row, 30)
    
    def update_summary_cards(self):
        """Update summary card values"""
        symbol = get_currency_symbol()
        
        total_categories = len(self.category_data)
        total_products = sum(row[1] for row in self.category_data)
        total_cost = sum(row[2] for row in self.category_data)
        total_stock = sum(row[3] for row in self.category_data)
        
        self.card_categories.set_value(str(total_categories))
        self.card_products.set_value(str(total_products))
        self.card_cost.set_value(format_money(total_cost, symbol))
        self.card_stock.set_value(str(total_stock))
    
    def export_to_excel(self):
        """Export category cost breakdown to Excel file"""
        if not self.category_data:
            QMessageBox.warning(self, tr("no_data"), tr("no_data_to_export"))
            return
        
        # Use filtered data if available, otherwise use all data
        export_data = self.filtered_data if self.filtered_data else self.category_data
        
        if not export_data:
            QMessageBox.warning(self, tr("no_data"), tr("no_data_to_export"))
            return

        file_path = ExcelExporter.save_file_dialog(
            self, 
            f"category_cost_breakdown_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            tr("export_category_cost_breakdown")
        )
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            symbol = get_currency_symbol()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Category Cost Breakdown"
            
            # Title
            ws.merge_cells('A1:F1')
            title_text = tr("category_cost_export_title")
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Subtitle with filter info
            ws['A2'] = f"{tr('generated')} {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = Font(size=10, color="7f8c8d")
            
            # Filter info
            filter_info = []
            search_text = self.search_widget.get_text().strip()
            if search_text:
                filter_info.append(f"Search: {search_text}")
            min_percentage = self.min_percentage_filter.currentText()
            if min_percentage != tr("all"):
                filter_info.append(f"{tr('min_percent')} {min_percentage}")
            
            if filter_info:
                ws['A3'] = " | ".join(filter_info)
                ws['A3'].font = Font(size=9, italic=True, color="7f8c8d")
                start_row = 5
            else:
                start_row = 4
            
            # Headers
            headers = [
                tr("category"), tr("product_count"), tr("total_cost"),
                tr("total_stock"), tr("stock_value_cost"), tr("percentage_header")
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row_idx, row_data in enumerate(export_data, start=start_row + 1):
                category, product_count, total_cost, total_stock, stock_value = row_data
                
                ws.cell(row=row_idx, column=1, value=category)
                ws.cell(row=row_idx, column=2, value=product_count)
                ws.cell(row=row_idx, column=3, value=format_money(total_cost, symbol))
                ws.cell(row=row_idx, column=4, value=total_stock)
                ws.cell(row=row_idx, column=5, value=format_money(stock_value, symbol))
                
                if self.total_cost_all > 0:
                    percentage = (total_cost / self.total_cost_all) * 100
                    ws.cell(row=row_idx, column=6, value=f"{percentage:.1f}%")
                else:
                    ws.cell(row=row_idx, column=6, value="0%")
            
            # Summary section
            summary_row = len(export_data) + start_row + 2
            total_products = sum(row[1] for row in export_data)
            total_cost = sum(row[2] for row in export_data)
            total_stock = sum(row[3] for row in export_data)
            
            ws.cell(row=summary_row, column=1, value=tr("summary")).font = Font(bold=True, size=12)
            
            summary_data = [
                (tr("total_categories"), len(export_data)),
                (tr("total_products_card"), total_products),
                (tr("total_cost"), format_money(total_cost, symbol)),
                (tr("total_stock"), total_stock)
            ]
            
            for i, (label, value) in enumerate(summary_data):
                row = summary_row + 2 + i
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
            
            # Auto adjust columns
            for col in range(1, 7):
                column_letter = chr(64 + col)
                ws.column_dimensions[column_letter].width = 20
            ws.column_dimensions['A'].width = 30
            
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
            
        except Exception as e:
            logger.error(f"Export category cost breakdown failed: {e}")
            ExcelExporter.show_error_message(self, e)
    
    def keyPressEvent(self, event):
        """Handle keyboard shortcuts"""
        if event.key() == Qt.Key.Key_Escape:
            self.accept()
        elif event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.key() == Qt.Key.Key_E:
                self.export_to_excel()
            elif event.key() == Qt.Key.Key_F:
                self.search_widget.focus_search()
        super().keyPressEvent(event)