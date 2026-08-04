# ui/receipts_page/discount_tab.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.search_widget import SearchWidget
from ui.widgets.toast_notification_widget import ToastNotificationWidget
from ui.widgets.modern_button import ModernButton
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime


class DiscountTab(QWidget):
    """Discounted Receipts Tab - Theme-aware (uses parent's date range)"""
    
    def __init__(self, user_id=None, user_role=None, parent=None):
        super().__init__(parent)
        self.user_id = user_id
        self.user_role = user_role
        self.parent_page = parent
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # ====== Search and Filters (No DateRangeWidget) ======
        top_layout = QHBoxLayout()
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(0, 8, 0, 8)
        
        self.search_widget = SearchWidget(
            placeholder="Search by invoice no...",
            show_label=True
        )
        self.search_widget.search_changed.connect(self.reset_and_load)
        top_layout.addWidget(self.search_widget, 2)
        
        # ✅ Export button with SVG icon
        self.btn_export = ModernButton(" Export Excel", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(True)
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addWidget(self.btn_export)
        
        top_layout.addStretch()
        layout.addLayout(top_layout)
        
        # ====== Table - NO custom style, use PyQt6 default ======
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setColumnHidden(0, True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.cellDoubleClicked.connect(self.show_receipt_detail)
        self.table.setAlternatingRowColors(True)
        
        # ✅ NO custom table style
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.table.horizontalHeader()
        for col in range(1, 6):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.table)
        
        # ====== Pagination ======
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        # ====== Toast Notification ======
        self.toast = ToastNotificationWidget(self)
        
        self.setLayout(layout)
        
        # Apply initial theme
        self._apply_theme()
        
        self.retranslateUi()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
        self._update_button_icons()
        self.load_discounted_receipts()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_export.set_icon("file_export", size=(16, 16))
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        # ✅ NO table style update
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        # Update button icons
        self._update_button_icons()
    
    def get_lang(self):
        from utils.language import lang
        return lang.get_current()
    
    def retranslateUi(self):
        lang = self.get_lang()
        
        self.search_widget.retranslateUi(lang)
        
        if lang == "my":
            self.search_widget.search_input.setPlaceholderText("ပြေစာအမှတ်ဖြင့် ရှာရန်...")
            self.btn_export.setText(" Excel ထုတ်မည်")
            headers = ["ID", "ပြေစာအမှတ်", "ရက်စွဲ", "စုစုပေါင်း", "လျှော့စျေး", "အသားတင်"]
        else:
            self.search_widget.search_input.setPlaceholderText("Search by invoice no...")
            self.btn_export.setText(" Export Excel")
            headers = ["ID", "Invoice No", "Date", "Total", "Discount", "Net Total"]
        
        self.table.setHorizontalHeaderLabels(headers)
        
        # Update button icons
        self._update_button_icons()
        
        # Apply theme after language change
        self._apply_theme()
    
    def on_page_changed(self, page: int, page_size: int):
        self.load_discounted_receipts(page, page_size)
    
    def reset_and_load(self):
        self.pagination.set_current_page(1)
        self.load_discounted_receipts()
    
    def load_discounted_receipts(self, from_date=None, to_date=None, page=1, page_size=50):
        """Load discounted receipts with date range from parent"""
        symbol = get_currency_symbol()
        search_text = self.search_widget.get_text()
        
        # Get date range from parent if not provided
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        conn = connect_db()
        cursor = conn.cursor()
        like = f'%{search_text}%'
        
        if search_text:
            cursor.execute("""
                SELECT COUNT(*) FROM sales
                WHERE status = 'completed'
                  AND discount_amount > 0
                  AND date(created_at) BETWEEN ? AND ?
                  AND invoice_no LIKE ?
            """, (from_date, to_date, like))
        else:
            cursor.execute("""
                SELECT COUNT(*) FROM sales
                WHERE status = 'completed'
                  AND discount_amount > 0
                  AND date(created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        offset = (page - 1) * page_size
        
        if search_text:
            cursor.execute("""
                SELECT id, invoice_no, created_at, total, discount_amount,
                       (total - discount_amount) as net_total
                FROM sales
                WHERE status = 'completed'
                  AND discount_amount > 0
                  AND date(created_at) BETWEEN ? AND ?
                  AND invoice_no LIKE ?
                ORDER BY created_at DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, like, page_size, offset))
        else:
            cursor.execute("""
                SELECT id, invoice_no, created_at, total, discount_amount,
                       (total - discount_amount) as net_total
                FROM sales
                WHERE status = 'completed'
                  AND discount_amount > 0
                  AND date(created_at) BETWEEN ? AND ?
                ORDER BY created_at DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, page_size, offset))
        rows = cursor.fetchall()
        conn.close()
        
        self.table.setRowCount(0)
        for row_data in rows:
            sale_id, invoice_no, created_at, total, discount_amount, net_total = row_data
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # ✅ Use PyQt6 default colors - no custom text color
            # ID (hidden)
            id_item = QTableWidgetItem(str(sale_id))
            self.table.setItem(row, 0, id_item)
            
            # Invoice No
            inv_item = QTableWidgetItem(invoice_no)
            self.table.setItem(row, 1, inv_item)
            
            # Date
            date_item = QTableWidgetItem(str(created_at)[:16])
            self.table.setItem(row, 2, date_item)
            
            # Total
            total_item = QTableWidgetItem(format_money(total, symbol))
            self.table.setItem(row, 3, total_item)
            
            # Discount - keep orange color for emphasis
            discount_item = QTableWidgetItem(format_money(discount_amount, symbol))
            if discount_amount > 0:
                discount_item.setForeground(QColor("#f39c12"))  # Orange
            self.table.setItem(row, 4, discount_item)
            
            # Net Total
            net_item = QTableWidgetItem(format_money(net_total, symbol))
            self.table.setItem(row, 5, net_item)
    
    def show_receipt_detail(self, row, column):
        id_item = self.table.item(row, 0)
        if id_item:
            try:
                sale_id = int(id_item.text())
                from ui.receipt_detail_dialog import ReceiptDetailDialog
                dialog = ReceiptDetailDialog(sale_id)
                dialog.exec()
            except ValueError as e:
                QMessageBox.warning(self, "Error", f"Invalid sale ID: {e}")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not open receipt details: {e}")
    
    def load_discounts(self):
        self.load_discounted_receipts()

    def _get_export_rows(self, from_date=None, to_date=None):
        """Get export rows with date range from parent"""
        search_text = self.search_widget.get_text()
        
        # Get date range from parent if not provided
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        like = f'%{search_text}%'

        conn = connect_db()
        cursor = conn.cursor()
        try:
            if search_text:
                cursor.execute("""
                    SELECT invoice_no, created_at, total, discount_amount,
                           (total - discount_amount) as net_total
                    FROM sales
                    WHERE status = 'completed'
                      AND discount_amount > 0
                      AND date(created_at) BETWEEN ? AND ?
                      AND invoice_no LIKE ?
                    ORDER BY created_at DESC
                """, (from_date, to_date, like))
            else:
                cursor.execute("""
                    SELECT invoice_no, created_at, total, discount_amount,
                           (total - discount_amount) as net_total
                    FROM sales
                    WHERE status = 'completed'
                      AND discount_amount > 0
                      AND date(created_at) BETWEEN ? AND ?
                    ORDER BY created_at DESC
                """, (from_date, to_date))
            return cursor.fetchall()
        finally:
            conn.close()
    
    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from utils.excel_exporter import ExcelExporter

        # Get date range from parent
        if hasattr(self.parent_page, 'get_current_date_range'):
            from_date, to_date = self.parent_page.get_current_date_range()
        else:
            today = QDate.currentDate().toString("yyyy-MM-dd")
            from_date, to_date = today, today
            
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"discounted_receipts_{from_date}_to_{to_date}_{datetime.now().strftime('%H%M%S')}.xlsx",
            "Export Discounted Receipts"
        )
        if not file_path:
            return

        try:
            rows = self._get_export_rows(from_date, to_date)
            if not rows:
                QMessageBox.information(self, "Export", "No discounted receipts to export.")
                return

            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Discounted Receipts"

            ws.merge_cells("A1:F1")
            ws["A1"] = "DISCOUNTED RECEIPTS"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws["A2"] = f"Date range: {from_date} to {to_date}"
            ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            headers = ["Invoice No", "Date", "Total", "Discount", "Net Total"]
            ExcelExporter.apply_header_style(ws, 5, headers)

            total_amount = 0.0
            total_discount = 0.0
            total_net = 0.0
            for row_idx, row_data in enumerate(rows, start=6):
                invoice_no, created_at, total, discount_amount, net_total = row_data
                total_amount += float(total or 0)
                total_discount += float(discount_amount or 0)
                total_net += float(net_total or 0)
                values = [
                    invoice_no,
                    str(created_at)[:16] if created_at else "",
                    format_money(total or 0, symbol),
                    format_money(discount_amount or 0, symbol),
                    format_money(net_total or 0, symbol),
                ]
                for col_idx, value in enumerate(values, start=1):
                    align = "right" if col_idx in (3, 4, 5) else "left"
                    ExcelExporter.apply_cell_style(ws, row_idx, col_idx, value, align)

            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=2, value="Totals").font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=format_money(total_amount, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=format_money(total_discount, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=format_money(total_net, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value="Record Count").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=3, value=len(rows)).font = Font(bold=True)

            ExcelExporter.auto_adjust_columns(ws, rows, start_row=6)
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def showEvent(self, event):
        self.load_discounted_receipts()
        super().showEvent(event)