# ui/receipts_page/credit_tab.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView,
    QComboBox
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from ui.receipt_detail_dialog import ReceiptDetailDialog
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.search_widget import SearchWidget
from ui.widgets.toast_notification_widget import ToastNotificationWidget
from ui.widgets.modern_button import ModernButton
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime


class CreditTab(QWidget):
    """Credit Receipts Tab - Theme-aware (uses parent's date range)"""
    
    def __init__(self, user_id=None, user_role=None, parent=None):
        super().__init__(parent)
        self.user_id = user_id
        self.user_role = user_role
        self.parent_page = parent
        self._is_dark = is_dark_theme()
        
        # Connect theme change
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # ====== Search and Filters (No DateRangeWidget) ======
        top_layout = QHBoxLayout()
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(0, 8, 0, 8)
        
        # SearchWidget
        self.search_widget = SearchWidget(
            placeholder="Search by invoice no or customer...",
            show_label=True
        )
        self.search_widget.search_changed.connect(self.reset_and_load)
        top_layout.addWidget(self.search_widget, 2)
        
        # Status ComboBox
        status_label = QLabel("Status:")
        status_label.setStyleSheet(self._get_label_style())
        top_layout.addWidget(status_label)
        
        self.status_filter = QComboBox()
        self.status_filter.addItems(["All", "Pending", "Partial", "Paid", "Overdue"])
        self.status_filter.currentTextChanged.connect(self.reset_and_load)
        self.status_filter.setFixedWidth(120)
        self.status_filter.setStyleSheet(self._get_combobox_style())
        top_layout.addWidget(self.status_filter)
        
        # ✅ Export button with SVG icon
        self.btn_export = ModernButton(" Export Excel", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(True)
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addWidget(self.btn_export)
        
        top_layout.addStretch()
        layout.addLayout(top_layout)
        
        # ====== Table - NO custom style, use PyQt6 default ======
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setColumnHidden(0, True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.cellDoubleClicked.connect(self.show_receipt_detail)
        self.table.setAlternatingRowColors(True)
        
        # ✅ NO custom table style
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.table.horizontalHeader()
        for col in range(1, 8):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.table)
        
        # ====== Pagination ======
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        # ====== Toast Notification ======
        self.toast = ToastNotificationWidget(self)
        
        self.setLayout(layout)
        
        # Apply initial theme
        self._apply_theme()
        
        self.retranslateUi()
    
    def _on_theme_changed(self, theme_name):
        """Handle theme change"""
        self._is_dark = is_dark_theme()
        self._apply_theme()
        self._update_button_icons()
        self.load_credit_receipts()
    
    def _update_button_icons(self):
        """Update button icons when theme changes"""
        self.btn_export.set_icon("file_export", size=(16, 16))
    
    def _apply_theme(self):
        """Apply theme-aware styles"""
        # ✅ NO table style update
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        # Update combobox
        if hasattr(self, 'status_filter'):
            self.status_filter.setStyleSheet(self._get_combobox_style())
        
        # Update labels
        for child in self.findChildren(QLabel):
            child.setStyleSheet(self._get_label_style())
        
        # Update button icons
        self._update_button_icons()
    
    def _get_label_style(self):
        colors = get_theme_colors()
        return f"color: {colors['text']}; font-size: 10pt;"
    
    def _get_combobox_style(self):
        colors = get_theme_colors()
        return f"""
            QComboBox {{
                padding: 6px 12px;
                border: 1px solid {colors['border']};
                border-radius: 6px;
                background: {colors['card_bg']};
                color: {colors['text']};
                font-size: 10pt;
            }}
            QComboBox:focus {{
                border-color: #5865f2;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox QAbstractItemView {{
                background-color: {colors['card_bg']};
                border: 1px solid {colors['border']};
                border-radius: 4px;
                color: {colors['text']};
                selection-background-color: #5865f2;
                selection-color: white;
                padding: 4px;
            }}
            QComboBox QAbstractItemView::item {{
                padding: 4px 8px;
                border-radius: 2px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background-color: {colors['bg_hover']};
            }}
            QComboBox QAbstractItemView::item:selected {{
                background-color: #5865f2;
                color: white;
            }}
        """
    
    def get_lang(self):
        from utils.language import lang
        return lang.get_current()
    
    def retranslateUi(self):
        lang = self.get_lang()
        
        self.search_widget.retranslateUi(lang)
        
        if lang == "my":
            self.search_widget.search_input.setPlaceholderText("ပြေစာအမှတ် သို့မဟုတ် ဝယ်ယူသူဖြင့် ရှာရန်...")
            self.btn_export.setText(" Excel ထုတ်မည်")
            self.status_filter.setItemText(0, "အားလုံး")
            self.status_filter.setItemText(1, "ဆိုင်းငံ့")
            self.status_filter.setItemText(2, "တစ်ပိုင်း")
            self.status_filter.setItemText(3, "ပြီးစီး")
            self.status_filter.setItemText(4, "သက်တမ်းလွန်")
            headers = ["ID", "ပြေစာအမှတ်", "ရက်စွဲ", "စုစုပေါင်း", 
                      "ဝယ်ယူသူ", "ကျန်ငွေ", "အခြေအနေ", "သတ်မှတ်ရက်"]
        else:
            self.search_widget.search_input.setPlaceholderText("Search by invoice no or customer...")
            self.btn_export.setText(" Export Excel")
            self.status_filter.setItemText(0, "All")
            self.status_filter.setItemText(1, "Pending")
            self.status_filter.setItemText(2, "Partial")
            self.status_filter.setItemText(3, "Paid")
            self.status_filter.setItemText(4, "Overdue")
            headers = ["ID", "Invoice No", "Date", "Total", 
                      "Customer", "Balance", "Status", "Due Date"]
        
        self.table.setHorizontalHeaderLabels(headers)
        
        # Update button icons
        self._update_button_icons()
        
        # Apply theme after language change
        self._apply_theme()
    
    def on_page_changed(self, page: int, page_size: int):
        self.load_credit_receipts(page, page_size)
    
    def reset_and_load(self):
        self.pagination.set_current_page(1)
        self.load_credit_receipts()
    
    def load_credit_receipts(self, from_date=None, to_date=None, page=1, page_size=25):
        """Load credit receipts with date range from parent"""
        symbol = get_currency_symbol()
        search_text = self.search_widget.get_text()
        status_filter = self.status_filter.currentText()
        lang = self.get_lang()
        
        # Get date range from parent if not provided
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        conn = connect_db()
        cursor = conn.cursor()
        like = f'%{search_text}%'
        
        status_condition = ""
        if status_filter == "Pending" or status_filter == "ဆိုင်းငံ့":
            status_condition = "AND cs.status = 'pending'"
        elif status_filter == "Partial" or status_filter == "တစ်ပိုင်း":
            status_condition = "AND cs.status = 'partial'"
        elif status_filter == "Paid" or status_filter == "ပြီးစီး":
            status_condition = "AND cs.status = 'paid'"
        elif status_filter == "Overdue" or status_filter == "သက်တမ်းလွန်":
            status_condition = "AND date(cs.due_date) < date('now') AND cs.balance_amount > 0"
        
        if search_text:
            cursor.execute(f"""
                SELECT COUNT(*) FROM credit_sales cs
                LEFT JOIN customers c ON cs.customer_id = c.id
                WHERE cs.sale_date BETWEEN ? AND ?
                  AND (cs.invoice_no LIKE ? OR c.name LIKE ?)
                  {status_condition}
            """, (from_date, to_date, like, like))
        else:
            cursor.execute(f"""
                SELECT COUNT(*) FROM credit_sales cs
                WHERE cs.sale_date BETWEEN ? AND ?
                  {status_condition}
            """, (from_date, to_date))
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        offset = (page - 1) * page_size
        
        if search_text:
            cursor.execute(f"""
                SELECT cs.id, cs.invoice_no, cs.sale_date, cs.total_amount,
                       c.name, cs.balance_amount, cs.status, cs.due_date
                FROM credit_sales cs
                LEFT JOIN customers c ON cs.customer_id = c.id
                WHERE cs.sale_date BETWEEN ? AND ?
                  AND (cs.invoice_no LIKE ? OR c.name LIKE ?)
                  {status_condition}
                ORDER BY cs.sale_date DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, like, like, page_size, offset))
        else:
            cursor.execute(f"""
                SELECT cs.id, cs.invoice_no, cs.sale_date, cs.total_amount,
                       c.name, cs.balance_amount, cs.status, cs.due_date
                FROM credit_sales cs
                LEFT JOIN customers c ON cs.customer_id = c.id
                WHERE cs.sale_date BETWEEN ? AND ?
                  {status_condition}
                ORDER BY cs.sale_date DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, page_size, offset))
        rows = cursor.fetchall()
        conn.close()
        
        self.table.setRowCount(0)
        today = QDate.currentDate()
        
        # ✅ Use hardcoded colors
        red_color = "#dc3545"
        green_color = "#28a745"
        orange_color = "#f39c12"
        
        for row_data in rows:
            sale_id, invoice_no, sale_date, total, customer_name, balance, status, due_date = row_data
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # ✅ Use PyQt6 default colors - no custom text color
            # ID (hidden)
            id_item = QTableWidgetItem(str(sale_id))
            self.table.setItem(row, 0, id_item)
            
            # Invoice No
            inv_item = QTableWidgetItem(invoice_no)
            self.table.setItem(row, 1, inv_item)
            
            # Date
            date_item = QTableWidgetItem(str(sale_date)[:16])
            self.table.setItem(row, 2, date_item)
            
            # Total
            total_item = QTableWidgetItem(format_money(total, symbol))
            self.table.setItem(row, 3, total_item)
            
            # Customer
            cust_name = customer_name if customer_name else "Walk-in"
            cust_item = QTableWidgetItem(cust_name)
            self.table.setItem(row, 4, cust_item)
            
            # Balance
            balance_item = QTableWidgetItem(format_money(balance, symbol))
            if balance > 0:
                balance_item.setForeground(QColor(red_color))
            else:
                balance_item.setForeground(QColor(green_color))
            self.table.setItem(row, 5, balance_item)
            
            # Status
            status_display = status
            if lang == "my":
                if status == "pending":
                    status_display = "ဆိုင်းငံ့"
                elif status == "partial":
                    status_display = "တစ်ပိုင်း"
                elif status == "paid":
                    status_display = "ပြီးစီး"
            
            status_item = QTableWidgetItem(status_display)
            
            is_overdue = False
            if due_date and balance > 0:
                try:
                    due_qdate = QDate.fromString(due_date, "yyyy-MM-dd")
                    if due_qdate.isValid() and due_qdate < today:
                        is_overdue = True
                except:
                    pass
            
            if status == "paid":
                status_item.setForeground(QColor(green_color))
            elif is_overdue:
                status_item.setForeground(QColor(red_color))
                status_display = "Overdue" if lang != "my" else "သက်တမ်းလွန်"
                status_item.setText(status_display)
            elif status == "pending" or status == "partial":
                status_item.setForeground(QColor(orange_color))
            
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 6, status_item)
            
            # Due Date
            due_item = QTableWidgetItem(due_date if due_date else "-")
            if is_overdue:
                due_item.setForeground(QColor(red_color))
            self.table.setItem(row, 7, due_item)
    
    def show_receipt_detail(self, row, column):
        id_item = self.table.item(row, 0)
        if id_item:
            try:
                credit_sale_id = int(id_item.text())
                dialog = ReceiptDetailDialog(credit_sale_id, self, is_credit=True)
                dialog.exec()
            except ValueError as e:
                QMessageBox.warning(self, "Error", f"Invalid sale ID: {e}")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not open receipt details: {e}")
    
    def load_credits(self):
        self.load_credit_receipts()

    def _get_status_condition(self):
        status_filter = self.status_filter.currentText()
        if status_filter == "Pending" or status_filter == "ဆိုင်းငံ့":
            return "AND cs.status = 'pending'"
        if status_filter == "Partial" or status_filter == "တစ်ပိုင်း":
            return "AND cs.status = 'partial'"
        if status_filter == "Paid" or status_filter == "ပြီးစီး":
            return "AND cs.status = 'paid'"
        if status_filter == "Overdue" or status_filter == "သက်တမ်းလွန်":
            return "AND date(cs.due_date) < date('now') AND cs.balance_amount > 0"
        return ""

    def _get_export_rows(self, from_date=None, to_date=None):
        """Get export rows with date range from parent"""
        search_text = self.search_widget.get_text()
        status_condition = self._get_status_condition()
        
        # Get date range from parent if not provided
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        like = f'%{search_text}%'

        conn = connect_db()
        cursor = conn.cursor()
        try:
            if search_text:
                cursor.execute(f"""
                    SELECT cs.invoice_no, cs.sale_date, cs.total_amount,
                           c.name, cs.paid_amount, cs.balance_amount,
                           cs.status, cs.due_date
                    FROM credit_sales cs
                    LEFT JOIN customers c ON cs.customer_id = c.id
                    WHERE cs.sale_date BETWEEN ? AND ?
                      AND (cs.invoice_no LIKE ? OR c.name LIKE ?)
                      {status_condition}
                    ORDER BY cs.sale_date DESC
                """, (from_date, to_date, like, like))
            else:
                cursor.execute(f"""
                    SELECT cs.invoice_no, cs.sale_date, cs.total_amount,
                           c.name, cs.paid_amount, cs.balance_amount,
                           cs.status, cs.due_date
                    FROM credit_sales cs
                    LEFT JOIN customers c ON cs.customer_id = c.id
                    WHERE cs.sale_date BETWEEN ? AND ?
                      {status_condition}
                    ORDER BY cs.sale_date DESC
                """, (from_date, to_date))
            return cursor.fetchall()
        finally:
            conn.close()

    def _display_status(self, status, balance, due_date):
        if due_date and float(balance or 0) > 0:
            due_qdate = QDate.fromString(str(due_date), "yyyy-MM-dd")
            if due_qdate.isValid() and due_qdate < QDate.currentDate():
                return "Overdue"
        if status == "pending":
            return "Pending"
        if status == "partial":
            return "Partial"
        if status == "paid":
            return "Paid"
        return status or ""
    
    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from utils.excel_exporter import ExcelExporter

        # Get date range from parent
        if hasattr(self.parent_page, 'get_current_date_range'):
            from_date, to_date = self.parent_page.get_current_date_range()
        else:
            today = QDate.currentDate().toString("yyyy-MM-dd")
            from_date, to_date = today, today
            
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"credit_receipts_{from_date}_to_{to_date}_{datetime.now().strftime('%H%M%S')}.xlsx",
            "Export Credit Receipts"
        )
        if not file_path:
            return

        try:
            rows = self._get_export_rows(from_date, to_date)
            if not rows:
                QMessageBox.information(self, "Export", "No credit receipts to export.")
                return

            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Credit Receipts"

            ws.merge_cells("A1:H1")
            ws["A1"] = "CREDIT RECEIPTS"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws["A2"] = f"Date range: {from_date} to {to_date}"
            ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            headers = ["Invoice No", "Date", "Total", "Customer", "Paid", "Balance", "Status", "Due Date"]
            ExcelExporter.apply_header_style(ws, 5, headers)

            total_amount = 0.0
            total_paid = 0.0
            total_balance = 0.0
            for row_idx, row_data in enumerate(rows, start=6):
                invoice_no, sale_date, total, customer_name, paid, balance, status, due_date = row_data
                total_amount += float(total or 0)
                total_paid += float(paid or 0)
                total_balance += float(balance or 0)
                values = [
                    invoice_no,
                    str(sale_date)[:16] if sale_date else "",
                    format_money(total or 0, symbol),
                    customer_name if customer_name else "Walk-in",
                    format_money(paid or 0, symbol),
                    format_money(balance or 0, symbol),
                    self._display_status(status, balance, due_date),
                    due_date if due_date else "-",
                ]
                for col_idx, value in enumerate(values, start=1):
                    align = "right" if col_idx in (3, 5, 6) else "left"
                    ExcelExporter.apply_cell_style(ws, row_idx, col_idx, value, align)

            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=2, value="Totals").font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=format_money(total_amount, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row, column=5, value=format_money(total_paid, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row, column=6, value=format_money(total_balance, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value="Record Count").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=3, value=len(rows)).font = Font(bold=True)

            ExcelExporter.auto_adjust_columns(ws, rows, start_row=6)
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def refresh(self):
        self.load_credit_receipts()
    
    def showEvent(self, event):
        self.load_credit_receipts()
        super().showEvent(event)
