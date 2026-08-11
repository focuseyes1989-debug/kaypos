# ui/receipts_page/refund_tab.py
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor
from models.database import connect_db
from utils.currency import get_currency_symbol, format_money
from ui.receipt_detail_dialog import ReceiptDetailDialog
from ui.widgets.pagination_widget import PaginationWidget
from ui.widgets.search_widget import SearchWidget
from ui.widgets.toast_notification_widget import ToastNotificationWidget
from ui.widgets.modern_button import ModernButton
from ui.themes.theme_manager import theme_manager, get_theme_colors, is_dark_theme
from datetime import datetime


class RefundTab(QWidget):
    """Refunded Receipts Tab - Theme-aware (uses parent's date range)"""
    
    def __init__(self, user_id=None, user_role=None, parent=None):
        super().__init__(parent)
        self.user_id = user_id
        self.user_role = user_role
        self.parent_page = parent
        self._is_dark = is_dark_theme()
        
        theme_manager.theme_changed.connect(self._on_theme_changed)
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        
        # ====== Search and Filters ======
        top_layout = QHBoxLayout()
        top_layout.setSpacing(10)
        top_layout.setContentsMargins(0, 8, 0, 8)
        
        self.search_widget = SearchWidget(
            placeholder="Search by invoice no or customer...",
            show_label=True
        )
        self.search_widget.search_changed.connect(self.reset_and_load)
        top_layout.addWidget(self.search_widget, 2)
        
        self.btn_export = ModernButton(" Export Excel", ModernButton.SECONDARY)
        self.btn_export.set_icon("file_export", size=(16, 16))
        self.btn_export.set_compact(True)
        self.btn_export.clicked.connect(self.export_to_excel)
        top_layout.addWidget(self.btn_export)
        
        top_layout.addStretch()
        layout.addLayout(top_layout)
        
        # ====== Table - NO custom style, use PyQt6 default ======
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setColumnHidden(0, True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.cellDoubleClicked.connect(self.show_receipt_detail)
        self.table.setAlternatingRowColors(True)
        
        # ✅ NO custom table style
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        
        header = self.table.horizontalHeader()
        for col in range(1, 7):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.table)
        
        # ====== Pagination ======
        self.pagination = PaginationWidget()
        self.pagination.page_changed.connect(self.on_page_changed)
        layout.addWidget(self.pagination)
        
        self.toast = ToastNotificationWidget(self)
        
        self.setLayout(layout)
        self._apply_theme()
        self.retranslateUi()
    
    def _on_theme_changed(self, theme_name):
        self._is_dark = is_dark_theme()
        self._apply_theme()
        self._update_button_icons()
        self.load_refunded_sales()
    
    def _update_button_icons(self):
        self.btn_export.set_icon("file_export", size=(16, 16))
    
    def _apply_theme(self):
        # ✅ NO table style update
        # self._update_table_style(colors)  <-- ဒီ line ကို ဖယ်ရှားပါ
        self._update_button_icons()
    
    def get_lang(self):
        from utils.language import lang
        return lang.get_current()
    
    def retranslateUi(self):
        lang = self.get_lang()
        
        self.search_widget.retranslateUi(lang)
        
        if lang == "my":
            self.search_widget.search_input.setPlaceholderText("ပြေစာအမှတ် သို့မဟုတ် ဝယ်ယူသူဖြင့် ရှာရန်...")
            self.btn_export.setText(" Excel ထုတ်မည်")
            headers = ["ID", "ပြေစာအမှတ်", "ရက်စွဲ", "စုစုပေါင်း", "ဝယ်ယူသူ", 
                      "ငွေပေးချေမှုအမျိုးအစား", "ပြန်အမ်းချိန်"]
        else:
            self.search_widget.search_input.setPlaceholderText("Search by invoice no or customer...")
            self.btn_export.setText(" Export Excel")
            headers = ["ID", "Invoice No", "Date", "Total", "Customer", 
                      "Payment Type", "Refunded At"]
        
        self.table.setHorizontalHeaderLabels(headers)
        self._update_button_icons()
        self._apply_theme()
    
    def on_page_changed(self, page: int, page_size: int):
        self.load_refunded_sales(page, page_size)
    
    def reset_and_load(self):
        self.pagination.set_current_page(1)
        self.load_refunded_sales()
    
    def load_refunded_sales(self, from_date=None, to_date=None, page=1, page_size=25):
        """Load refunded sales - ✅ FIXED: Use sale_items for total"""
        symbol = get_currency_symbol()
        search_text = self.search_widget.get_text()
        
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        conn = connect_db()
        cursor = conn.cursor()
        like = f'%{search_text}%'
        
        if search_text:
            cursor.execute("""
                SELECT COUNT(*) FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                WHERE s.status = 'refunded'
                  AND date(s.created_at) BETWEEN ? AND ?
                  AND (s.invoice_no LIKE ? OR c.name LIKE ?)
            """, (from_date, to_date, like, like))
        else:
            cursor.execute("""
                SELECT COUNT(*) FROM sales s
                WHERE s.status = 'refunded'
                  AND date(s.created_at) BETWEEN ? AND ?
            """, (from_date, to_date))
        total_items = cursor.fetchone()[0]
        self.pagination.set_total_items(total_items, emit_signal=False)
        
        offset = (page - 1) * page_size
        
        # ✅ FIXED: Using sale_items for total
        if search_text:
            cursor.execute("""
                SELECT 
                    s.id, 
                    s.invoice_no, 
                    s.created_at, 
                    COALESCE(SUM(si.qty * si.price), 0) as total,
                    c.name, 
                    s.payment_type
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE s.status = 'refunded'
                  AND date(s.created_at) BETWEEN ? AND ?
                  AND (s.invoice_no LIKE ? OR c.name LIKE ?)
                GROUP BY s.id, s.invoice_no, s.created_at, c.name, s.payment_type
                ORDER BY s.created_at DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, like, like, page_size, offset))
        else:
            cursor.execute("""
                SELECT 
                    s.id, 
                    s.invoice_no, 
                    s.created_at, 
                    COALESCE(SUM(si.qty * si.price), 0) as total,
                    c.name, 
                    s.payment_type
                FROM sales s
                LEFT JOIN customers c ON s.customer_id = c.id
                LEFT JOIN sale_items si ON s.id = si.sale_id
                WHERE s.status = 'refunded'
                  AND date(s.created_at) BETWEEN ? AND ?
                GROUP BY s.id, s.invoice_no, s.created_at, c.name, s.payment_type
                ORDER BY s.created_at DESC
                LIMIT ? OFFSET ?
            """, (from_date, to_date, page_size, offset))
        rows = cursor.fetchall()
        conn.close()
        
        self.table.setRowCount(0)
        for row_data in rows:
            sale_id, invoice_no, created_at, total, customer_name, payment_type = row_data
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # ✅ Use PyQt6 default colors - no custom text color
            id_item = QTableWidgetItem(str(sale_id))
            self.table.setItem(row, 0, id_item)
            
            inv_item = QTableWidgetItem(invoice_no)
            self.table.setItem(row, 1, inv_item)
            
            date_item = QTableWidgetItem(str(created_at)[:16])
            self.table.setItem(row, 2, date_item)
            
            total_item = QTableWidgetItem(format_money(total, symbol))
            self.table.setItem(row, 3, total_item)
            
            cust_name = customer_name if customer_name else "Walk-in"
            cust_item = QTableWidgetItem(cust_name)
            self.table.setItem(row, 4, cust_item)
            
            ptype_item = QTableWidgetItem(payment_type if payment_type else "-")
            self.table.setItem(row, 5, ptype_item)
            
            refund_item = QTableWidgetItem(str(created_at)[:16] if created_at else "-")
            self.table.setItem(row, 6, refund_item)
    
    def show_receipt_detail(self, row, column):
        id_item = self.table.item(row, 0)
        if id_item:
            try:
                sale_id = int(id_item.text())
                dialog = ReceiptDetailDialog(sale_id)
                dialog.exec()
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Could not open receipt details: {e}")

    def _get_export_rows(self, from_date=None, to_date=None):
        search_text = self.search_widget.get_text()
        
        if from_date is None or to_date is None:
            if hasattr(self.parent_page, 'get_current_date_range'):
                from_date, to_date = self.parent_page.get_current_date_range()
            else:
                today = QDate.currentDate().toString("yyyy-MM-dd")
                from_date, to_date = today, today
        
        like = f'%{search_text}%'

        conn = connect_db()
        cursor = conn.cursor()
        try:
            # ✅ FIXED: Using sale_items for total
            if search_text:
                cursor.execute("""
                    SELECT 
                        s.invoice_no, 
                        s.created_at, 
                        COALESCE(SUM(si.qty * si.price), 0) as total,
                        c.name, 
                        s.payment_type
                    FROM sales s
                    LEFT JOIN customers c ON s.customer_id = c.id
                    LEFT JOIN sale_items si ON s.id = si.sale_id
                    WHERE s.status = 'refunded'
                      AND date(s.created_at) BETWEEN ? AND ?
                      AND (s.invoice_no LIKE ? OR c.name LIKE ?)
                    GROUP BY s.id, s.invoice_no, s.created_at, c.name, s.payment_type
                    ORDER BY s.created_at DESC
                """, (from_date, to_date, like, like))
            else:
                cursor.execute("""
                    SELECT 
                        s.invoice_no, 
                        s.created_at, 
                        COALESCE(SUM(si.qty * si.price), 0) as total,
                        c.name, 
                        s.payment_type
                    FROM sales s
                    LEFT JOIN customers c ON s.customer_id = c.id
                    LEFT JOIN sale_items si ON s.id = si.sale_id
                    WHERE s.status = 'refunded'
                      AND date(s.created_at) BETWEEN ? AND ?
                    GROUP BY s.id, s.invoice_no, s.created_at, c.name, s.payment_type
                    ORDER BY s.created_at DESC
                """, (from_date, to_date))
            return cursor.fetchall()
        finally:
            conn.close()
    
    def export_to_excel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from utils.excel_exporter import ExcelExporter

        if hasattr(self.parent_page, 'get_current_date_range'):
            from_date, to_date = self.parent_page.get_current_date_range()
        else:
            today = QDate.currentDate().toString("yyyy-MM-dd")
            from_date, to_date = today, today
            
        file_path = ExcelExporter.save_file_dialog(
            self,
            f"refunded_receipts_{from_date}_to_{to_date}_{datetime.now().strftime('%H%M%S')}.xlsx",
            "Export Refunded Receipts"
        )
        if not file_path:
            return

        try:
            rows = self._get_export_rows(from_date, to_date)
            if not rows:
                QMessageBox.information(self, "Export", "No refunded receipts to export.")
                return

            symbol = get_currency_symbol()
            wb = Workbook()
            ws = wb.active
            ws.title = "Refunded Receipts"

            ws.merge_cells("A1:F1")
            ws["A1"] = "REFUNDED RECEIPTS"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")
            ws["A2"] = f"Date range: {from_date} to {to_date}"
            ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

            headers = ["Invoice No", "Date", "Total Refunded", "Customer", "Payment Type", "Refunded At"]
            ExcelExporter.apply_header_style(ws, 5, headers)

            total_refunded = 0.0
            for row_idx, row_data in enumerate(rows, start=6):
                invoice_no, created_at, total, customer_name, payment_type = row_data
                total_refunded += float(total or 0)
                values = [
                    invoice_no,
                    str(created_at)[:16] if created_at else "",
                    format_money(total or 0, symbol),
                    customer_name if customer_name else "Walk-in",
                    payment_type if payment_type else "-",
                    str(created_at)[:16] if created_at else "",
                ]
                for col_idx, value in enumerate(values, start=1):
                    align = "right" if col_idx == 3 else "left"
                    ExcelExporter.apply_cell_style(ws, row_idx, col_idx, value, align)

            summary_row = len(rows) + 7
            ws.cell(row=summary_row, column=2, value="Total Refunded").font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=format_money(total_refunded, symbol)).font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=2, value="Record Count").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=3, value=len(rows)).font = Font(bold=True)

            ExcelExporter.auto_adjust_columns(ws, rows, start_row=6)
            wb.save(file_path)
            ExcelExporter.show_success_message(self, file_path)
        except Exception as e:
            ExcelExporter.show_error_message(self, e)
    
    def refresh(self):
        self.load_refunded_sales()
    
    def showEvent(self, event):
        self.load_refunded_sales()
        super().showEvent(event)
