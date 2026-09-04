"""Pure report serializers. Exports always use the captured report snapshot."""
from contextlib import contextmanager
import csv
from html import escape
import os
from pathlib import Path
import re
import tempfile


def display(value, kind='text'):
    if value is None: return '—'
    if kind == 'money': return f'{float(value):,.2f}'
    if kind == 'percent': return f'{float(value):,.2f}%'
    if kind == 'integer': return f'{int(value):,}'
    if kind == 'number' and isinstance(value, (int, float)): return f'{value:,.2f}'.rstrip('0').rstrip('.')
    return str(value)


def csv_value(value):
    if value is None: return ''
    if isinstance(value, str) and value.lstrip().startswith(('=', '+', '-', '@')): return "'" + value
    return value


@contextmanager
def atomic_output(path):
    target = Path(path)
    descriptor, temporary = tempfile.mkstemp(dir=target.parent, prefix=target.stem + '-', suffix=target.suffix)
    os.close(descriptor); temporary = Path(temporary)
    try:
        yield temporary
        temporary.replace(target)
    finally:
        temporary.unlink(missing_ok=True)


def write_csv(path, report, table, filter_text=''):
    with atomic_output(path) as temporary:
        with temporary.open('w', newline='', encoding='utf-8-sig') as stream:
            writer = csv.writer(stream)
            writer.writerow(['Period start', 'Period end', 'Snapshot time', 'Currency', 'Table', 'Row filter', *[c['label'] for c in table['columns']]])
            for row in table['rows']:
                values = [report['start'], report['end'], report['as_of'], report.get('currency',''), table['title'], filter_text, *[row.get(c['key']) for c in table['columns']]]
                writer.writerow([csv_value(v) for v in values])


def write_xlsx(path, report, tables, filter_text=''):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    workbook = Workbook(); metadata = workbook.active; metadata.title = 'Report information'
    rows = [['KAY POS Native report', report['section'] + ' / ' + report['view']],
        ['Period start', report['start']], ['Period end', report['end']], ['Snapshot time', report['as_of']], ['Currency', report.get('currency','')],
        ['Row filter (tables only)', filter_text], ['Summary scope', 'All records in selected dates'],
        *[[key, value] for key, value in report.get('metrics', {}).items()],
        *[[key + ' (current snapshot)', value] for key, value in report.get('snapshot', {}).items()],
        *[['Definition', note] for note in report.get('notes', [])]]
    if report.get('previous_period'): rows.append(['Comparison period', report['previous_period']['start'] + ' to ' + report['previous_period']['end']])
    for row in rows: metadata.append(row)
    for table in tables:
        title = re.sub(r'[\\/*?:\[\]]', '_', table['title'])[:31] or 'Report'
        sheet = workbook.create_sheet(title)
        sheet.append([c['label'] for c in table['columns']])
        for row in table['rows']: sheet.append([row.get(c['key']) for c in table['columns']])
        sheet.freeze_panes = 'A2'; sheet.auto_filter.ref = sheet.dimensions
        for column, definition in enumerate(table['columns'], 1):
            for cells in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
                if definition['kind'] in {'money', 'number', 'percent'}: cells[0].number_format = '#,##0.00'
                elif definition['kind'] == 'integer': cells[0].number_format = '#,##0'
    for sheet in workbook:
        for cell in sheet[1]: cell.font = Font(bold=True)
        for row in sheet:
            for cell in row:
                if isinstance(cell.value, str): cell.data_type = 's'
        for column in sheet.columns:
            length = max((len(str(cell.value or '')) for cell in column), default=10)
            sheet.column_dimensions[column[0].column_letter].width = min(65, max(14, length + 2))
    try:
        with atomic_output(path) as temporary: workbook.save(temporary)
    finally: workbook.close()


def report_html(report, table, filter_text=''):
    out = ['<html><body><h2>KAY POS Native · ' + escape(table['title']) + '</h2>',
        '<p>' + escape(report['start'] + ' to ' + report['end'] + ' · snapshot ' + report['as_of']) + '</p>',
        '<p>Currency: ' + escape(report.get('currency','')) + ' · Row filter: ' + escape(filter_text or 'None') + f" · {len(table['rows']):,} rows</p>"]
    if report.get('previous_period'):
        out.append('<p>Comparison period: ' + escape(report['previous_period']['start'] + ' to ' + report['previous_period']['end']) + '</p>')
    out.append('<table border="1" cellspacing="0" cellpadding="4"><thead><tr>')
    out.extend('<th>' + escape(c['label']) + '</th>' for c in table['columns']); out.append('</tr></thead><tbody>')
    for row in table['rows']:
        out.append('<tr>')
        out.extend('<td>' + escape(display(row.get(c['key']), c['kind'])) + '</td>' for c in table['columns'])
        out.append('</tr>')
    out.append('</tbody></table>')
    out.extend('<p>' + escape(note) + '</p>' for note in report.get('notes', []))
    out.append('</body></html>'); return ''.join(out)
